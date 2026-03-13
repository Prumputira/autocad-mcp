"""AutoCAD MCP Server v5.0 — 22 tools with multi-CAD connection, batch ops, NLP, and Excel export.

Tools: drawing, entity, layer, block, annotation, pid, view, system, query, search, geometry,
       select, modify, validate, export, xref, layout, electrical, connection, batch, nlp,
       excel_export
"""

from __future__ import annotations

import structlog
from mcp.server.fastmcp import FastMCP

from autocad_mcp.client import (
    _error,
    _json,
    _safe,
    add_screenshot_if_available,
    get_backend,
)

# FastMCP validates return types via Pydantic. Tools that may return
# ImageContent (screenshot) alongside TextContent need a union return type.
ToolResult = str | list

log = structlog.get_logger()

mcp = FastMCP("autocad-mcp")


# ==========================================================================
# 1. drawing — File/drawing management
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD Drawing Operations", "readOnlyHint": False})
@_safe("drawing")
async def drawing(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Drawing file management.

    Operations:
      create     — Create a new empty drawing. data: {name?}
      open       — Open an existing drawing. data: {path}
      info       — Get drawing extents, entity count, layers, blocks.
      save       — Save current drawing. data: {path?} (saves to path if given, else QSAVE)
      save_as_dxf — Export as DXF. data: {path}
      plot_pdf   — Plot to PDF. data: {path}
      purge      — Purge unused objects.
      get_variables — Get system variables. data: {names: [...]}
      undo       — Undo last operation.
      redo       — Redo last undone operation.
      audit      — Audit drawing for errors. data: {fix?}
      units      — Set drawing units. data: {units}
      limits     — Set drawing limits. data: {x1, y1, x2, y2}
      wblock     — Write block to file. data: {handles, path}
    """
    data = data or {}
    backend = await get_backend()

    if operation == "create":
        result = await backend.drawing_create(data.get("name"))
    elif operation == "info":
        result = await backend.drawing_info()
    elif operation == "save":
        result = await backend.drawing_save(data.get("path"))
    elif operation == "save_as_dxf":
        result = await backend.drawing_save_as_dxf(data["path"])
    elif operation == "plot_pdf":
        result = await backend.drawing_plot_pdf(data["path"])
    elif operation == "purge":
        result = await backend.drawing_purge()
    elif operation == "get_variables":
        result = await backend.drawing_get_variables(data.get("names"))
    elif operation == "open":
        result = await backend.drawing_open(data["path"])
    elif operation == "undo":
        result = await backend.undo()
    elif operation == "redo":
        result = await backend.redo()
    elif operation == "audit":
        result = await backend.drawing_audit(data.get("fix", False))
    elif operation == "units":
        result = await backend.drawing_units(data.get("units"))
    elif operation == "limits":
        result = await backend.drawing_limits(data.get("x1"), data.get("y1"), data.get("x2"), data.get("y2"))
    elif operation == "wblock":
        result = await backend.drawing_wblock(data.get("handles"), data.get("path"))
    else:
        return _json({"error": f"Unknown drawing operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# 2. entity — Entity CRUD + modification
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD Entity Operations", "readOnlyHint": False})
@_safe("entity")
async def entity(
    operation: str,
    x1: float | None = None,
    y1: float | None = None,
    x2: float | None = None,
    y2: float | None = None,
    points: list[list[float]] | None = None,
    layer: str | None = None,
    entity_id: str | None = None,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Entity creation, querying, and modification.

    Create operations:
      create_line       — x1, y1, x2, y2, layer?
      create_circle     — data: {cx, cy, radius}, layer?
      create_polyline   — points: [[x,y],...], data: {closed?}, layer?
      create_rectangle  — x1, y1, x2, y2, layer?
      create_arc        — data: {cx, cy, radius, start_angle, end_angle}, layer?
      create_ellipse    — data: {cx, cy, major_x, major_y, ratio}, layer?
      create_mtext      — data: {x, y, width, text, height?}, layer?
      create_hatch      — entity_id, data: {pattern?}

    Read operations:
      list              — layer? → list entities
      count             — layer? → count entities
      get               — entity_id → entity details

    Modify operations:
      copy    — entity_id, data: {dx, dy}
      move    — entity_id, data: {dx, dy}
      rotate  — entity_id, data: {cx, cy, angle}
      scale   — entity_id, data: {cx, cy, factor}
      mirror  — entity_id, x1, y1, x2, y2
      offset  — entity_id, data: {distance}
      array   — entity_id, data: {rows, cols, row_dist, col_dist}
      fillet  — data: {id1, id2, radius}
      chamfer — data: {id1, id2, dist1, dist2}
      erase   — entity_id
      explode — entity_id → explode entity into components
      join    — data: {entity_ids} → join entities
      extend  — entity_id, data: {boundary_id}
      trim    — entity_id, data: {boundary_id}
      break_at — entity_id, x1, y1 → break entity at point

    Composite operations:
      place_equipment_tag — data: {cx, cy, cz?, tag, cube_size?, direction?, text_height?}
          Places a complete equipment tag group (3D cube + leader + MTEXT + underline).
          Returns handles for all 4 entities, center, tag, and bounding box.
    """
    data = data or {}
    backend = await get_backend()

    # --- Create ---
    if operation == "create_line":
        result = await backend.create_line(x1, y1, x2, y2, layer)
    elif operation == "create_circle":
        result = await backend.create_circle(data["cx"], data["cy"], data["radius"], layer)
    elif operation == "create_polyline":
        result = await backend.create_polyline(points or [], data.get("closed", False), layer)
    elif operation == "create_rectangle":
        result = await backend.create_rectangle(x1, y1, x2, y2, layer)
    elif operation == "create_arc":
        result = await backend.create_arc(data["cx"], data["cy"], data["radius"], data["start_angle"], data["end_angle"], layer)
    elif operation == "create_ellipse":
        result = await backend.create_ellipse(data["cx"], data["cy"], data["major_x"], data["major_y"], data["ratio"], layer)
    elif operation == "create_mtext":
        result = await backend.create_mtext(data["x"], data["y"], data["width"], data["text"], data.get("height", 2.5), layer)
    elif operation == "create_hatch":
        result = await backend.create_hatch(entity_id, data.get("pattern", "ANSI31"))
    # --- Read ---
    elif operation == "list":
        result = await backend.entity_list(layer)
    elif operation == "count":
        result = await backend.entity_count(layer)
    elif operation == "get":
        result = await backend.entity_get(entity_id)
    # --- Modify ---
    elif operation == "copy":
        result = await backend.entity_copy(entity_id, data["dx"], data["dy"])
    elif operation == "move":
        result = await backend.entity_move(entity_id, data["dx"], data["dy"])
    elif operation == "rotate":
        result = await backend.entity_rotate(entity_id, data["cx"], data["cy"], data["angle"])
    elif operation == "scale":
        result = await backend.entity_scale(entity_id, data["cx"], data["cy"], data["factor"])
    elif operation == "mirror":
        result = await backend.entity_mirror(entity_id, x1, y1, x2, y2)
    elif operation == "offset":
        result = await backend.entity_offset(entity_id, data["distance"])
    elif operation == "array":
        result = await backend.entity_array(entity_id, data["rows"], data["cols"], data["row_dist"], data["col_dist"])
    elif operation == "fillet":
        result = await backend.entity_fillet(data["id1"], data["id2"], data["radius"])
    elif operation == "chamfer":
        result = await backend.entity_chamfer(data["id1"], data["id2"], data["dist1"], data["dist2"])
    elif operation == "erase":
        result = await backend.entity_erase(entity_id)
    elif operation == "explode":
        result = await backend.entity_explode(entity_id)
    elif operation == "join":
        entity_ids = data.get("entity_ids")
        result = await backend.entity_join(entity_ids)
    elif operation == "extend":
        boundary_id = data.get("boundary_id")
        result = await backend.entity_extend(entity_id, boundary_id)
    elif operation == "trim":
        boundary_id = data.get("boundary_id")
        result = await backend.entity_trim(entity_id, boundary_id)
    elif operation == "break_at":
        result = await backend.entity_break_at(entity_id, x1, y1)
    # --- Composite ---
    elif operation == "place_equipment_tag":
        result = await backend.place_equipment_tag(
            data["cx"], data["cy"], data.get("cz", 0.0),
            data["tag"], data.get("cube_size", 24.0),
            data.get("direction", "right"), data.get("text_height", 8.0),
        )
    else:
        return _json({"error": f"Unknown entity operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# 3. layer — Layer management
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD Layer Operations", "readOnlyHint": False})
@_safe("layer")
async def layer(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Layer creation and management.

    Operations:
      list            — List all layers with properties.
      create          — data: {name, color?, linetype?}
      set_current     — data: {name}
      set_properties  — data: {name, color?, linetype?, lineweight?}
      freeze          — data: {name}
      thaw            — data: {name}
      lock            — data: {name}
      unlock          — data: {name}
    """
    data = data or {}
    backend = await get_backend()

    if operation == "list":
        result = await backend.layer_list()
    elif operation == "create":
        result = await backend.layer_create(data["name"], data.get("color", "white"), data.get("linetype", "CONTINUOUS"))
    elif operation == "set_current":
        result = await backend.layer_set_current(data["name"])
    elif operation == "set_properties":
        result = await backend.layer_set_properties(data["name"], data.get("color"), data.get("linetype"), data.get("lineweight"))
    elif operation == "freeze":
        result = await backend.layer_freeze(data["name"])
    elif operation == "thaw":
        result = await backend.layer_thaw(data["name"])
    elif operation == "lock":
        result = await backend.layer_lock(data["name"])
    elif operation == "unlock":
        result = await backend.layer_unlock(data["name"])
    else:
        return _json({"error": f"Unknown layer operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# 4. block — Block operations
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD Block Operations", "readOnlyHint": False})
@_safe("block")
async def block(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Block definition, insertion, and attribute management.

    Operations:
      list                 — List all block definitions.
      insert               — data: {name, x, y, scale?, rotation?, block_id?}
      insert_with_attributes — data: {name, x, y, scale?, rotation?, attributes: {tag: value}}
      get_attributes       — data: {entity_id}
      update_attribute     — data: {entity_id, tag, value}
      define               — data: {name, entities: [{type, ...}]}
    """
    data = data or {}
    backend = await get_backend()

    if operation == "list":
        result = await backend.block_list()
    elif operation == "insert":
        result = await backend.block_insert(
            data["name"], data["x"], data["y"],
            data.get("scale", 1.0), data.get("rotation", 0.0), data.get("block_id"),
        )
    elif operation == "insert_with_attributes":
        result = await backend.block_insert_with_attributes(
            data["name"], data["x"], data["y"],
            data.get("scale", 1.0), data.get("rotation", 0.0), data.get("attributes"),
        )
    elif operation == "get_attributes":
        result = await backend.block_get_attributes(data["entity_id"])
    elif operation == "update_attribute":
        result = await backend.block_update_attribute(data["entity_id"], data["tag"], data["value"])
    elif operation == "define":
        result = await backend.block_define(data["name"], data.get("entities", []))
    else:
        return _json({"error": f"Unknown block operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# 5. annotation — Text, dimensions, leaders
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD Annotation Operations", "readOnlyHint": False})
@_safe("annotation")
async def annotation(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Annotation: text, dimensions, and leaders.

    Operations:
      create_text             — data: {x, y, text, height?, rotation?, layer?}
      create_dimension_linear — data: {x1, y1, x2, y2, dim_x, dim_y}
      create_dimension_aligned — data: {x1, y1, x2, y2, offset}
      create_dimension_angular — data: {cx, cy, x1, y1, x2, y2}
      create_dimension_radius — data: {cx, cy, radius, angle}
      create_leader           — data: {points: [[x,y],...], text}
    """
    data = data or {}
    backend = await get_backend()

    if operation == "create_text":
        result = await backend.create_text(
            data["x"], data["y"], data["text"],
            data.get("height", 2.5), data.get("rotation", 0.0), data.get("layer"),
        )
    elif operation == "create_dimension_linear":
        result = await backend.create_dimension_linear(
            data["x1"], data["y1"], data["x2"], data["y2"], data["dim_x"], data["dim_y"],
        )
    elif operation == "create_dimension_aligned":
        result = await backend.create_dimension_aligned(
            data["x1"], data["y1"], data["x2"], data["y2"], data["offset"],
        )
    elif operation == "create_dimension_angular":
        result = await backend.create_dimension_angular(
            data["cx"], data["cy"], data["x1"], data["y1"], data["x2"], data["y2"],
        )
    elif operation == "create_dimension_radius":
        result = await backend.create_dimension_radius(
            data["cx"], data["cy"], data["radius"], data["angle"],
        )
    elif operation == "create_leader":
        result = await backend.create_leader(data["points"], data["text"])
    else:
        return _json({"error": f"Unknown annotation operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# 6. pid — P&ID operations (CTO library)
# ==========================================================================


@mcp.tool(annotations={"title": "P&ID Operations (CTO Library)", "readOnlyHint": False})
@_safe("pid")
async def pid(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """P&ID drawing with CTO symbol library.

    Operations:
      setup_layers     — Create standard P&ID layers.
      insert_symbol    — data: {category, symbol, x, y, scale?, rotation?}
      list_symbols     — data: {category}
      draw_process_line — data: {x1, y1, x2, y2}
      connect_equipment — data: {x1, y1, x2, y2}
      add_flow_arrow   — data: {x, y, rotation?}
      add_equipment_tag — data: {x, y, tag, description?}
      add_line_number  — data: {x, y, line_num, spec}
      insert_valve     — data: {x, y, valve_type, rotation?, attributes?}
      insert_instrument — data: {x, y, instrument_type, rotation?, tag_id?, range_value?}
      insert_pump      — data: {x, y, pump_type, rotation?, attributes?}
      insert_tank      — data: {x, y, tank_type, scale?, attributes?}
    """
    data = data or {}
    backend = await get_backend()

    if operation == "setup_layers":
        result = await backend.pid_setup_layers()
    elif operation == "insert_symbol":
        result = await backend.pid_insert_symbol(
            data["category"], data["symbol"], data["x"], data["y"],
            data.get("scale", 1.0), data.get("rotation", 0.0),
        )
    elif operation == "list_symbols":
        result = await backend.pid_list_symbols(data["category"])
    elif operation == "draw_process_line":
        result = await backend.pid_draw_process_line(data["x1"], data["y1"], data["x2"], data["y2"])
    elif operation == "connect_equipment":
        result = await backend.pid_connect_equipment(data["x1"], data["y1"], data["x2"], data["y2"])
    elif operation == "add_flow_arrow":
        result = await backend.pid_add_flow_arrow(data["x"], data["y"], data.get("rotation", 0.0))
    elif operation == "add_equipment_tag":
        result = await backend.pid_add_equipment_tag(data["x"], data["y"], data["tag"], data.get("description", ""))
    elif operation == "add_line_number":
        result = await backend.pid_add_line_number(data["x"], data["y"], data["line_num"], data["spec"])
    elif operation == "insert_valve":
        result = await backend.pid_insert_valve(
            data["x"], data["y"], data["valve_type"],
            data.get("rotation", 0.0), data.get("attributes"),
        )
    elif operation == "insert_instrument":
        result = await backend.pid_insert_instrument(
            data["x"], data["y"], data["instrument_type"],
            data.get("rotation", 0.0), data.get("tag_id", ""), data.get("range_value", ""),
        )
    elif operation == "insert_pump":
        result = await backend.pid_insert_pump(
            data["x"], data["y"], data["pump_type"],
            data.get("rotation", 0.0), data.get("attributes"),
        )
    elif operation == "insert_tank":
        result = await backend.pid_insert_tank(
            data["x"], data["y"], data["tank_type"],
            data.get("scale", 1.0), data.get("attributes"),
        )
    else:
        return _json({"error": f"Unknown pid operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# 7. view — Viewport and screenshot
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD View Operations", "readOnlyHint": True})
@_safe("view")
async def view(
    operation: str,
    x1: float | None = None,
    y1: float | None = None,
    x2: float | None = None,
    y2: float | None = None,
    data: dict | None = None,
) -> ToolResult:
    """Viewport control and screenshot capture.

    Operations:
      zoom_extents   — Zoom to show all entities.
      zoom_window    — Zoom to window: x1, y1, x2, y2
      zoom_center    — Zoom to center: data.x, data.y, data.height
      layer_visibility — Toggle layer on/off: data.name, data.visible
      pan            — Pan view by offset: data.dx, data.dy
      zoom_scale     — Zoom by scale factor: data.factor
      get_screenshot — Capture current view as PNG image.
    """
    data = data or {}
    backend = await get_backend()

    if operation == "zoom_extents":
        result = await backend.zoom_extents()
        return _json(result.to_dict())
    elif operation == "zoom_window":
        result = await backend.zoom_window(x1, y1, x2, y2)
        return _json(result.to_dict())
    elif operation == "zoom_center":
        result = await backend.zoom_center(
            data.get("x", 0), data.get("y", 0), data.get("height", 100))
        return _json(result.to_dict())
    elif operation == "layer_visibility":
        name = data.get("name", "")
        visible = data.get("visible", True)
        if not name:
            return _json({"error": "layer name is required"})
        result = await backend.layer_visibility(name, visible)
        return _json(result.to_dict())
    elif operation == "pan":
        result = await backend.pan(data.get("dx"), data.get("dy"))
        return _json(result.to_dict())
    elif operation == "zoom_scale":
        result = await backend.zoom_scale(data.get("factor"))
        return _json(result.to_dict())
    elif operation == "get_screenshot":
        result = await backend.get_screenshot()
        if result.ok and result.payload:
            from mcp.types import ImageContent, TextContent

            return [
                TextContent(type="text", text=_json({"ok": True, "screenshot": "attached"})),
                ImageContent(type="image", data=result.payload, mimeType="image/png"),
            ]
        return _json(result.to_dict())
    else:
        return _json({"error": f"Unknown view operation: {operation}"})


# ==========================================================================
# 8. system — Server management
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD MCP System", "readOnlyHint": True})
@_safe("system")
async def system(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Server status and management.

    Operations:
      status        — Backend info, capabilities, health check.
      health        — Quick health check (ping backend).
      get_backend   — Return current backend name and capabilities.
      runtime       — Return process/runtime details for spawn diagnostics.
      init          — Re-initialize the backend.
      execute_lisp  — Execute arbitrary AutoLISP code (File IPC only). data: {code}
    """
    data = data or {}

    if operation == "status" or operation == "get_backend":
        backend = await get_backend()
        result = await backend.status()
        return await add_screenshot_if_available(result, include_screenshot)
    elif operation == "health":
        try:
            backend = await get_backend()
            result = await backend.status()
            return _json({"ok": result.ok, "backend": backend.name})
        except Exception as e:
            return _json({"ok": False, "error": str(e)})
    elif operation == "runtime":
        import os
        import sys

        return _json(
            {
                "ok": True,
                "platform": sys.platform,
                "python": sys.executable,
                "cwd": os.getcwd(),
                "backend_env": os.environ.get("AUTOCAD_MCP_BACKEND", "auto"),
                "wsl_interop": bool(os.environ.get("WSL_INTEROP")),
            }
        )
    elif operation == "init":
        # Force re-initialization
        from autocad_mcp import client
        client._backend = None
        backend = await get_backend()
        result = await backend.status()
        return _json(result.to_dict())
    elif operation == "execute_lisp":
        backend = await get_backend()
        if not data.get("code"):
            return _json({"error": "data.code is required"})
        result = await backend.execute_lisp(data["code"])
        return await add_screenshot_if_available(result, include_screenshot)
    else:
        return _json({"error": f"Unknown system operation: {operation}"})


# ==========================================================================
# 9. query — Entity and drawing queries
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD Query Operations", "readOnlyHint": True})
@_safe("query")
async def query(
    operation: str,
    entity_id: str | None = None,
    layer: str | None = None,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Query entity properties, geometry, and drawing statistics.

    Operations:
      entity_properties — Get ALL properties of an entity. entity_id required.
      entity_geometry   — Get precise geometry data (vertices, center, etc). entity_id required.
      drawing_summary   — Get entity count by type/layer, extents, block/style/linetype inventory.
      layer_summary     — Get per-layer stats: entity count by type, bounding box. layer required.
      text_styles       — List all text styles.
      dimension_styles  — List all dimension styles.
      linetypes         — List all linetypes.
      block_tree        — Hierarchical block/nested block tree.
      drawing_metadata  — Drawing metadata (author, dates, custom properties).
    """
    data = data or {}
    backend = await get_backend()

    if operation == "entity_properties":
        eid = entity_id or data.get("entity_id")
        if not eid:
            return _json({"error": "entity_id is required"})
        result = await backend.query_entity_properties(eid)
    elif operation == "entity_geometry":
        eid = entity_id or data.get("entity_id")
        if not eid:
            return _json({"error": "entity_id is required"})
        result = await backend.query_entity_geometry(eid)
    elif operation == "drawing_summary":
        result = await backend.query_drawing_summary()
    elif operation == "layer_summary":
        lyr = layer or data.get("layer")
        if not lyr:
            return _json({"error": "layer is required"})
        result = await backend.query_layer_summary(lyr)
    elif operation == "text_styles":
        result = await backend.query_text_styles()
    elif operation == "dimension_styles":
        result = await backend.query_dimension_styles()
    elif operation == "linetypes":
        result = await backend.query_linetypes()
    elif operation == "block_tree":
        result = await backend.query_block_tree()
    elif operation == "drawing_metadata":
        result = await backend.query_drawing_metadata()
    else:
        return _json({"error": f"Unknown query operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# 10. search — Find entities by various criteria
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD Search Operations", "readOnlyHint": True})
@_safe("search")
async def search(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Search for entities by text, attributes, location, type, or layer.

    Operations:
      text             — Find TEXT/MTEXT containing pattern. data: {pattern, case_sensitive?}
      by_attribute     — Find INSERTs with matching attribute. data: {tag?, value?}
      by_window        — Find entities within rectangle. data: {x1, y1, x2, y2}
      by_proximity     — Find entities near a point. data: {x, y, radius}
      by_type_and_layer — Filter by type/layer/color. data: {entity_type?, layer?, color?}
      by_block_name    — Find block insertions by name. data: {block_name}
      by_handle_list   — Get entities by handle list. data: {handles: [...]}
      equipment_find   — Deep search across modelspace, attributes, and block definitions.
                         data: {pattern, case_sensitive?, search_scope?, zoom_to_first?, zoom_height?, max_results?}
                         search_scope: "all" (default), "modelspace", "blocks", "attributes"
      equipment_inspect — Zoom to area, find nearby entities, infer equipment center.
                         data: {x, y, view_width?, view_height?, infer_center?, handle?}
      find_text        — Deep text search across modelspace AND all block definitions.
                         Searches TEXT, MTEXT, DIMENSION, INSERT ATTRIBs, and text inside
                         block definitions (including nested/xref blocks). Like AutoCAD FIND.
                         data: {pattern, case_sensitive?, max_results?, zoom_to_first?, zoom_height?}
      batch_find_and_tag — Find multiple tags and place equipment tag groups at each location.
                         Single LISP execution for speed. Returns placed count and per-tag results.
                         data: {tags: [...], cube_size?, direction?, text_height?}
    """
    data = data or {}
    backend = await get_backend()

    if operation == "text":
        if "pattern" not in data:
            return _json({"error": "data.pattern is required"})
        result = await backend.search_text(
            data["pattern"],
            data.get("case_sensitive", False),
        )
    elif operation == "by_attribute":
        result = await backend.search_by_attribute(
            data.get("tag"),
            data.get("value"),
        )
    elif operation == "by_window":
        for key in ("x1", "y1", "x2", "y2"):
            if key not in data:
                return _json({"error": f"data.{key} is required"})
        result = await backend.search_by_window(
            data["x1"], data["y1"], data["x2"], data["y2"],
        )
    elif operation == "by_proximity":
        for key in ("x", "y", "radius"):
            if key not in data:
                return _json({"error": f"data.{key} is required"})
        result = await backend.search_by_proximity(
            data["x"], data["y"], data["radius"],
        )
    elif operation == "by_type_and_layer":
        result = await backend.search_by_type_and_layer(
            data.get("entity_type"),
            data.get("layer"),
            data.get("color"),
        )
    elif operation == "by_block_name":
        block_name = data.get("block_name")
        result = await backend.search_by_block_name(block_name)
    elif operation == "by_handle_list":
        handles = data.get("handles")
        result = await backend.search_by_handle_list(handles)
    elif operation == "equipment_find":
        if "pattern" not in data:
            return _json({"error": "data.pattern is required"})
        result = await backend.equipment_find(
            data["pattern"],
            case_sensitive=data.get("case_sensitive", False),
            search_scope=data.get("search_scope", "all"),
            zoom_to_first=data.get("zoom_to_first", True),
            zoom_height=data.get("zoom_height", 600.0),
            max_results=data.get("max_results", 50),
        )
        if not include_screenshot:
            include_screenshot = True  # default on for equipment_find
    elif operation == "equipment_inspect":
        for key in ("x", "y"):
            if key not in data:
                return _json({"error": f"data.{key} is required"})
        result = await backend.equipment_inspect(
            data["x"], data["y"],
            view_width=data.get("view_width", 600.0),
            view_height=data.get("view_height", 600.0),
            infer_center=data.get("infer_center", True),
            handle=data.get("handle"),
        )
        if not include_screenshot:
            include_screenshot = True  # default on for equipment_inspect
    elif operation == "find_text":
        if "pattern" not in data:
            return _json({"error": "data.pattern is required"})
        result = await backend.find_text(
            data["pattern"],
            case_sensitive=data.get("case_sensitive", False),
            max_results=data.get("max_results", 50),
            zoom_to_first=data.get("zoom_to_first", True),
            zoom_height=data.get("zoom_height", 600.0),
        )
        if not include_screenshot:
            include_screenshot = True  # default on for find_text
    elif operation == "batch_find_and_tag":
        tags = data.get("tags")
        if not tags or not isinstance(tags, list):
            return _json({"error": "data.tags (list of strings) is required"})
        result = await backend.batch_find_and_tag(
            tags,
            cube_size=data.get("cube_size", 24.0),
            direction=data.get("direction", "right"),
            text_height=data.get("text_height", 8.0),
        )
    else:
        return _json({"error": f"Unknown search operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# 11. geometry — Measurement and analysis
# ==========================================================================


@mcp.tool(annotations={"title": "AutoCAD Geometry Operations", "readOnlyHint": True})
@_safe("geometry")
async def geometry(
    operation: str,
    entity_id: str | None = None,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Geometry measurement: distance, length, area, bounding box, polyline analysis.

    Operations:
      distance       — Distance between two points. data: {x1, y1, x2, y2}
      length         — Total length of line/polyline/arc/circle. entity_id required.
      area           — Area of closed polyline/circle/ellipse. entity_id required.
      bounding_box   — Axis-aligned bounding box. entity_id or data: {layer?}
      polyline_info  — Vertices, bulges, segment lengths, total length, area. entity_id required.
    """
    data = data or {}
    backend = await get_backend()

    if operation == "distance":
        for key in ("x1", "y1", "x2", "y2"):
            if key not in data:
                return _json({"error": f"data.{key} is required"})
        result = await backend.geometry_distance(
            data["x1"], data["y1"], data["x2"], data["y2"],
        )
    elif operation == "length":
        eid = entity_id or data.get("entity_id")
        if not eid:
            return _json({"error": "entity_id is required"})
        result = await backend.geometry_length(eid)
    elif operation == "area":
        eid = entity_id or data.get("entity_id")
        if not eid:
            return _json({"error": "entity_id is required"})
        result = await backend.geometry_area(eid)
    elif operation == "bounding_box":
        eid = entity_id or data.get("entity_id")
        lyr = data.get("layer")
        result = await backend.geometry_bounding_box(eid, lyr)
    elif operation == "polyline_info":
        eid = entity_id or data.get("entity_id")
        if not eid:
            return _json({"error": "entity_id is required"})
        result = await backend.geometry_polyline_info(eid)
    else:
        return _json({"error": f"Unknown geometry operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 12: select — Filter + Bulk Operations
# ==========================================================================


@mcp.tool(annotations={"title": "Select & Bulk Operations"})
@_safe("select")
async def select(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Filter entities and perform bulk operations.

    Operations: filter, bulk_move, bulk_copy, bulk_erase, bulk_set_property,
    find_replace_text, find_replace_attribute, layer_rename, layer_merge
    """
    data = data or {}
    backend = await get_backend()

    if operation == "filter":
        result = await backend.select_filter(
            entity_type=data.get("entity_type"),
            layer=data.get("layer"),
            color=data.get("color"),
            x1=data.get("x1"), y1=data.get("y1"),
            x2=data.get("x2"), y2=data.get("y2"),
        )
    elif operation == "bulk_move":
        handles = data.get("handles", [])
        if not handles:
            return _json({"error": "handles list is required"})
        result = await backend.bulk_move(handles, data.get("dx", 0), data.get("dy", 0))
    elif operation == "bulk_copy":
        handles = data.get("handles", [])
        if not handles:
            return _json({"error": "handles list is required"})
        result = await backend.bulk_copy(handles, data.get("dx", 0), data.get("dy", 0))
    elif operation == "bulk_erase":
        handles = data.get("handles", [])
        if not handles:
            return _json({"error": "handles list is required"})
        result = await backend.bulk_erase(handles)
    elif operation == "bulk_set_property":
        handles = data.get("handles", [])
        prop = data.get("property")
        value = data.get("value")
        if not handles or not prop:
            return _json({"error": "handles, property, and value are required"})
        result = await backend.bulk_set_property(handles, prop, value)
    elif operation == "find_replace_text":
        find = data.get("find")
        replace = data.get("replace")
        if not find:
            return _json({"error": "find string is required"})
        result = await backend.find_replace_text(find, replace or "", data.get("layer"))
    elif operation == "find_replace_attribute":
        tag = data.get("tag")
        find = data.get("find")
        replace = data.get("replace")
        result = await backend.find_replace_attribute(tag, find, replace)
    elif operation == "layer_rename":
        old_name = data.get("old_name")
        new_name = data.get("new_name")
        result = await backend.layer_rename(old_name, new_name)
    elif operation == "layer_merge":
        source = data.get("source_layer")
        target = data.get("target_layer")
        result = await backend.layer_merge(source, target)
    else:
        return _json({"error": f"Unknown select operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 13: modify — Enhanced Entity Modification
# ==========================================================================


@mcp.tool(annotations={"title": "Entity Modification"})
@_safe("modify")
async def modify(
    operation: str,
    entity_id: str = "",
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Modify individual entity properties and content.

    Operations: set_property, set_text
    """
    data = data or {}
    backend = await get_backend()
    eid = entity_id or data.get("entity_id", "")

    if operation == "set_property":
        if not eid:
            return _json({"error": "entity_id is required"})
        prop = data.get("property")
        value = data.get("value")
        if not prop:
            return _json({"error": "property name is required"})
        result = await backend.entity_set_property(eid, prop, value)
    elif operation == "set_text":
        if not eid:
            return _json({"error": "entity_id is required"})
        text = data.get("text", "")
        result = await backend.entity_set_text(eid, text)
    else:
        return _json({"error": f"Unknown modify operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 14: validate — QC & Standards Checking
# ==========================================================================


@mcp.tool(annotations={"title": "Validate & QC", "readOnlyHint": True})
@_safe("validate")
async def validate(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Automated QC and standards verification.

    Operations: layer_standards, duplicates, zero_length, qc_report,
    text_standards, orphaned_entities, attribute_completeness, connectivity
    """
    data = data or {}
    backend = await get_backend()

    if operation == "layer_standards":
        allowed = data.get("allowed_layers", [])
        if not allowed:
            return _json({"error": "allowed_layers list is required"})
        result = await backend.validate_layer_standards(allowed)
    elif operation == "duplicates":
        result = await backend.validate_duplicates(data.get("tolerance", 0.001))
    elif operation == "zero_length":
        result = await backend.validate_zero_length()
    elif operation == "qc_report":
        result = await backend.validate_qc_report(data.get("allowed_layers"))
    elif operation == "text_standards":
        allowed_styles = data.get("allowed_styles")
        allowed_heights = data.get("allowed_heights")
        result = await backend.validate_text_standards(allowed_styles, allowed_heights)
    elif operation == "orphaned_entities":
        result = await backend.validate_orphaned_entities()
    elif operation == "attribute_completeness":
        required_tags = data.get("required_tags")
        result = await backend.validate_attribute_completeness(required_tags)
    elif operation == "connectivity":
        layer = data.get("layer")
        tolerance = data.get("tolerance", 0.01)
        result = await backend.validate_connectivity(layer, tolerance)
    else:
        return _json({"error": f"Unknown validate operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 15: export — Reporting & Data Extraction
# ==========================================================================


@mcp.tool(annotations={"title": "Export & Reports", "readOnlyHint": True})
@_safe("export")
async def export(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Generate reports, BOMs, and extract structured data.

    Operations: entity_data, bom, data_extract, layer_report, block_count,
    drawing_statistics
    """
    data = data or {}
    backend = await get_backend()

    if operation == "entity_data":
        result = await backend.export_entity_data(
            layer=data.get("layer"), entity_type=data.get("entity_type"))
    elif operation == "bom":
        result = await backend.export_bom(data.get("block_names"))
    elif operation == "data_extract":
        result = await backend.export_data_extract(
            entity_type=data.get("entity_type"),
            layer=data.get("layer"),
            properties=data.get("properties"),
        )
    elif operation == "layer_report":
        result = await backend.export_layer_report()
    elif operation == "block_count":
        result = await backend.export_block_count()
    elif operation == "drawing_statistics":
        result = await backend.export_drawing_statistics()
    else:
        return _json({"error": f"Unknown export operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 16: xref — External Reference Management
# ==========================================================================


@mcp.tool(annotations={"title": "XREF Management"})
@_safe("xref")
async def xref(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """External reference management.

    Operations: list, attach, detach, reload, bind, path_update, query_entities
    """
    data = data or {}
    backend = await get_backend()

    if operation == "list":
        result = await backend.xref_list()
    elif operation == "attach":
        path = data.get("path")
        if not path:
            return _json({"error": "path is required"})
        result = await backend.xref_attach(
            path, data.get("x", 0), data.get("y", 0),
            data.get("scale", 1.0), data.get("overlay", False),
        )
    elif operation == "detach":
        name = data.get("name")
        if not name:
            return _json({"error": "name is required"})
        result = await backend.xref_detach(name)
    elif operation == "reload":
        result = await backend.xref_reload(data.get("name", ""))
    elif operation == "bind":
        result = await backend.xref_bind(data.get("name", ""), data.get("insert", False))
    elif operation == "path_update":
        result = await backend.xref_path_update(data.get("name", ""), data.get("new_path", ""))
    elif operation == "query_entities":
        result = await backend.xref_query_entities(
            data.get("name", ""), data.get("entity_type"), data.get("layer"),
        )
    else:
        return _json({"error": f"Unknown xref operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 17: layout — Layout & Paper Space
# ==========================================================================


@mcp.tool(annotations={"title": "Layout & Paper Space"})
@_safe("layout")
async def layout(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Paper space, viewports, and plotting.

    Operations: list, create, switch, delete, viewport_create,
    viewport_set_scale, viewport_lock, page_setup, titleblock_fill, batch_plot
    """
    data = data or {}
    backend = await get_backend()

    if operation == "list":
        result = await backend.layout_list()
    elif operation == "create":
        name = data.get("name")
        if not name:
            return _json({"error": "name is required"})
        result = await backend.layout_create(name)
    elif operation == "switch":
        result = await backend.layout_switch(data.get("name", ""))
    elif operation == "delete":
        result = await backend.layout_delete(data.get("name", ""))
    elif operation == "viewport_create":
        result = await backend.layout_viewport_create(
            data.get("x", 0), data.get("y", 0),
            data.get("width", 100), data.get("height", 100),
            data.get("scale", 1.0),
        )
    elif operation == "viewport_set_scale":
        result = await backend.layout_viewport_set_scale(
            data.get("viewport_id", ""), data.get("scale", 1.0),
        )
    elif operation == "viewport_lock":
        result = await backend.layout_viewport_lock(
            data.get("viewport_id", ""), data.get("lock", True),
        )
    elif operation == "page_setup":
        result = await backend.layout_page_setup(
            data.get("name", ""), data.get("paper_size"), data.get("orientation"),
        )
    elif operation == "titleblock_fill":
        result = await backend.layout_titleblock_fill(
            data.get("layout_name", ""), data.get("attributes"),
        )
    elif operation == "batch_plot":
        result = await backend.layout_batch_plot(
            data.get("layouts"), data.get("output_path"),
        )
    else:
        return _json({"error": f"Unknown layout operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 18: electrical — Electrical Engineering
# ==========================================================================


@mcp.tool(annotations={"title": "Electrical Engineering"})
@_safe("electrical")
async def electrical(
    operation: str,
    data: dict | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Electrical domain tools: NEC lookups, calculations, symbol insertion.

    Operations: nec_lookup, voltage_drop, conduit_fill, load_calc,
    symbol_insert, circuit_trace, panel_schedule_gen, wire_number_assign
    """
    data = data or {}
    backend = await get_backend()

    if operation == "nec_lookup":
        result = await backend.electrical_nec_lookup(
            data.get("table", ""), data.get("parameters"),
        )
    elif operation == "voltage_drop":
        result = await backend.electrical_voltage_drop(
            data.get("voltage", 120), data.get("current", 0),
            data.get("wire_gauge", "12"), data.get("length", 0),
            data.get("phase", 1), data.get("power_factor", 1.0),
        )
    elif operation == "conduit_fill":
        result = await backend.electrical_conduit_fill(
            data.get("conduit_size", ""), data.get("conduit_type", "EMT"),
            data.get("wire_gauges"),
        )
    elif operation == "load_calc":
        result = await backend.electrical_load_calc(data.get("devices"))
    elif operation == "symbol_insert":
        result = await backend.electrical_symbol_insert(
            data.get("symbol_type", ""), data.get("x", 0), data.get("y", 0),
            data.get("scale", 1.0), data.get("rotation", 0.0), data.get("layer"),
        )
    elif operation == "circuit_trace":
        result = await backend.electrical_circuit_trace(
            data.get("start_entity", ""), data.get("layer"),
        )
    elif operation == "panel_schedule_gen":
        result = await backend.electrical_panel_schedule_gen(
            data.get("panel_block", ""), data.get("x", 0), data.get("y", 0),
        )
    elif operation == "wire_number_assign":
        result = await backend.electrical_wire_number_assign(
            data.get("layer", ""), data.get("prefix", "W"), data.get("start_num", 1),
        )
    else:
        return _json({"error": f"Unknown electrical operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 19: connection — Multi-CAD connection management (COM backend)
# ==========================================================================


@mcp.tool(annotations={"title": "CAD Connection Management", "readOnlyHint": False})
@_safe("connection")
async def connection(
    operation: str,
    data: dict | None = None,
) -> ToolResult:
    """Multi-CAD connection management (COM backend only).

    Operations:
      connect        — Connect to a CAD application. data: {cad_type: "autocad"|"zwcad"|"gcad"|"bricscad"}
      disconnect     — Disconnect from the current CAD application.
      status         — Get connection status for all supported CAD types.
      list_supported — List all supported CAD application types.
      switch_backend — Switch backend: data: {backend: "file_ipc"|"com"|"ezdxf"}
    """
    data = data or {}

    if operation == "list_supported":
        from autocad_mcp.config import get_supported_cads
        return _json({"ok": True, "payload": {"supported_cads": get_supported_cads()}})

    if operation == "switch_backend":
        from autocad_mcp.client import switch_backend
        backend_name = data.get("backend", "ezdxf")
        backend = await switch_backend(backend_name)
        return _json({"ok": True, "payload": {"backend": backend.name, "capabilities": vars(backend.capabilities)}})

    backend = await get_backend()

    if operation == "connect":
        result = await backend.connect(data.get("cad_type", "autocad"))
    elif operation == "disconnect":
        result = await backend.disconnect()
    elif operation == "status":
        result = await backend.get_connection_status()
    else:
        return _json({"error": f"Unknown connection operation: {operation}"})

    return _json(result.to_dict())


# ==========================================================================
# Tool 20: batch — Batch drawing operations (60-70% fewer API calls)
# ==========================================================================


@mcp.tool(annotations={"title": "Batch Drawing Operations", "readOnlyHint": False})
@_safe("batch")
async def batch(
    operation: str,
    items: list[dict] | None = None,
    include_screenshot: bool = False,
) -> ToolResult:
    """Batch drawing operations for creating multiple entities efficiently.

    Operations:
      draw_lines      — Draw multiple lines. items: [{x1, y1, x2, y2, layer?, color?}, ...]
      draw_circles    — Draw multiple circles. items: [{cx, cy, radius, layer?, color?}, ...]
      draw_rectangles — Draw multiple rectangles. items: [{x1, y1, x2, y2, layer?, color?}, ...]
      draw_polylines  — Draw multiple polylines. items: [{points: [[x,y],...], closed?, layer?, color?}, ...]
      draw_texts      — Draw multiple text entities. items: [{x, y, text, height?, rotation?, layer?}, ...]
    """
    items = items or []
    if not items:
        return _json({"error": "No items provided for batch operation"})

    backend = await get_backend()

    if operation == "draw_lines":
        result = await backend.batch_draw_lines(items)
    elif operation == "draw_circles":
        result = await backend.batch_draw_circles(items)
    elif operation == "draw_rectangles":
        result = await backend.batch_draw_rectangles(items)
    elif operation == "draw_polylines":
        result = await backend.batch_draw_polylines(items)
    elif operation == "draw_texts":
        result = await backend.batch_draw_texts(items)
    else:
        return _json({"error": f"Unknown batch operation: {operation}"})

    return await add_screenshot_if_available(result, include_screenshot)


# ==========================================================================
# Tool 21: nlp — Natural language command execution
# ==========================================================================


@mcp.tool(annotations={"title": "Natural Language CAD Commands", "readOnlyHint": False})
@_safe("nlp")
async def nlp(
    command: str,
    include_screenshot: bool = False,
) -> ToolResult:
    """Execute a CAD command from natural language description.

    Parses natural language and maps to the appropriate CAD operation.
    Supports shape drawing, layer management in English and Swedish.

    Examples:
      "draw a red line from 0,0 to 100,100"
      "create a circle at 50,50 with radius 25"
      "rita en cirkel vid 10,20 med radie 30"
      "create layer Sprinkler with color red"
      "skapa lager Ventilation"
    """
    from autocad_mcp.nlp.processor import NLPProcessor

    processor = NLPProcessor(strict_mode=False)
    parsed = processor.parse_command(command)

    if parsed.confidence < 0.2:
        return _json({
            "ok": False,
            "error": f"Could not parse command: {command}",
            "parsed": {"operation": parsed.operation, "confidence": parsed.confidence},
        })

    backend = await get_backend()
    p = parsed.parameters

    # Route to backend based on parsed operation
    if parsed.operation == "create_line":
        result = await backend.create_line(
            p.get("x1", 0), p.get("y1", 0), p.get("x2", 100), p.get("y2", 100),
            layer=p.get("layer"),
        )
    elif parsed.operation == "create_circle":
        result = await backend.create_circle(
            p.get("cx", 0), p.get("cy", 0), p.get("radius", 50),
            layer=p.get("layer"),
        )
    elif parsed.operation == "create_rectangle":
        result = await backend.create_rectangle(
            p.get("x1", 0), p.get("y1", 0), p.get("x2", 100), p.get("y2", 100),
            layer=p.get("layer"),
        )
    elif parsed.operation == "create_arc":
        result = await backend.create_arc(
            p.get("cx", 0), p.get("cy", 0), p.get("radius", 50),
            p.get("start_angle", 0), p.get("end_angle", 90),
            layer=p.get("layer"),
        )
    elif parsed.operation == "create_polyline":
        result = await backend.create_polyline(
            p.get("points", [[0, 0], [100, 0], [100, 100]]),
            closed=p.get("closed", False),
            layer=p.get("layer"),
        )
    elif parsed.operation == "create_ellipse":
        result = await backend.create_ellipse(
            p.get("cx", 0), p.get("cy", 0),
            p.get("major_x", 50), p.get("major_y", 0), p.get("ratio", 0.5),
            layer=p.get("layer"),
        )
    elif parsed.operation == "create_text":
        result = await backend.create_text(
            p.get("x", 0), p.get("y", 0), p.get("text", "Text"),
            height=p.get("height", 2.5), layer=p.get("layer"),
        )
    elif parsed.operation == "layer_create":
        result = await backend.layer_create(
            p.get("name", "NewLayer"), color=p.get("color", "white"),
        )
    elif parsed.operation == "layer_freeze":
        result = await backend.layer_freeze(p.get("name", "0"))
    elif parsed.operation == "layer_thaw":
        result = await backend.layer_thaw(p.get("name", "0"))
    elif parsed.operation == "layer_list":
        result = await backend.layer_list()
    elif parsed.operation == "layer_rename":
        result = await backend.layer_rename(
            p.get("old_name", ""), p.get("new_name", ""),
        )
    else:
        from autocad_mcp.backends.base import CommandResult
        result = CommandResult(
            ok=False,
            error=f"NLP parsed operation '{parsed.operation}' is not yet routed to a backend method",
        )

    # Add NLP metadata to response
    response = result.to_dict()
    response["nlp"] = {
        "parsed_operation": parsed.operation,
        "confidence": parsed.confidence,
        "parameters": parsed.parameters,
    }

    if include_screenshot and result.ok:
        return await add_screenshot_if_available(result, True)

    return _json(response)


# ==========================================================================
# Tool 22: excel_export — Export drawing data to Excel
# ==========================================================================


@mcp.tool(annotations={"title": "Excel Export", "readOnlyHint": True})
@_safe("excel_export")
async def excel_export(
    operation: str,
    data: dict | None = None,
) -> ToolResult:
    """Export drawing data to Excel spreadsheets.

    Operations:
      full_export     — Export all entities to Excel. data: {filename?, output_dir?}
      selected_export — Export selected/filtered entities. data: {filename?, layer?, entity_type?}
    """
    data = data or {}
    backend = await get_backend()

    if operation == "full_export":
        result = await backend.export_to_excel(
            data.get("filename", "drawing_data.xlsx"),
            data.get("output_dir"),
        )
    elif operation == "selected_export":
        # First get filtered entities, then export
        entity_data = await backend.export_entity_data(
            layer=data.get("layer"), entity_type=data.get("entity_type"),
        )
        if not entity_data.ok:
            return _json(entity_data.to_dict())

        # Use openpyxl to write
        try:
            from pathlib import Path
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            filename = data.get("filename", "filtered_export.xlsx")
            output_dir = Path(data.get("output_dir", "~/Documents/AutoCAD MCP Exports")).expanduser()
            output_dir.mkdir(parents=True, exist_ok=True)
            filepath = output_dir / filename

            entities = entity_data.payload if isinstance(entity_data.payload, list) else []

            wb = Workbook()
            ws = wb.active
            ws.title = "Entities"

            columns = ["handle", "type", "layer", "color"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, ent in enumerate(entities, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=ent.get(col_name, ""))

            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

            ws.freeze_panes = "A2"
            wb.save(str(filepath))

            from autocad_mcp.backends.base import CommandResult as CR
            result = CR(ok=True, payload={"file": str(filepath), "count": len(entities)})
        except ImportError:
            from autocad_mcp.backends.base import CommandResult as CR
            result = CR(ok=False, error="openpyxl not installed. Install with: pip install openpyxl")
        except Exception as ex:
            from autocad_mcp.backends.base import CommandResult as CR
            result = CR(ok=False, error=str(ex))
    else:
        return _json({"error": f"Unknown excel_export operation: {operation}"})

    return _json(result.to_dict())


# ==========================================================================
# Main entry point
# ==========================================================================


def main():
    """Run the MCP server on stdio transport."""
    import logging
    import sys

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        handlers=[logging.StreamHandler(sys.stderr)],
    )

    structlog.configure(
        wrapper_class=structlog.make_filtering_bound_logger(logging.INFO),
        logger_factory=structlog.PrintLoggerFactory(file=sys.stderr),
        cache_logger_on_first_use=True,
        processors=[
            structlog.processors.add_log_level,
            structlog.processors.TimeStamper(fmt="iso"),
            structlog.dev.ConsoleRenderer(),
        ],
    )

    log.info("autocad_mcp_starting", version="5.0.0")
    mcp.run(transport="stdio")
