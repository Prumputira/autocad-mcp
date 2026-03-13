"""COM-based backend for AutoCAD MCP.

Supports multiple CAD applications (AutoCAD, ZWCAD, GstarCAD, BricsCAD) via
Windows COM automation (ActiveX / IDispatch).  Falls back gracefully on
non-Windows platforms: the class remains importable but ``initialize()``
returns an error.
"""

from __future__ import annotations

import base64
import io
import math
import re
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

import structlog

from autocad_mcp.backends.base import AutoCADBackend, BackendCapabilities, CommandResult
from autocad_mcp.config import get_cad_config, get_supported_cads

log = structlog.get_logger()

# ---------------------------------------------------------------------------
# Guarded Win32 imports
# ---------------------------------------------------------------------------

_WIN32 = sys.platform == "win32"

if _WIN32:
    try:
        import pythoncom
        import pywintypes
        import win32com.client
        import win32gui

        _WIN32_AVAILABLE = True
    except ImportError:
        _WIN32_AVAILABLE = False
else:
    _WIN32_AVAILABLE = False

# ---------------------------------------------------------------------------
# ACI colour map (AutoCAD Color Index)
# ---------------------------------------------------------------------------

COLOR_MAP: dict[str, int] = {
    "black": 0,
    "red": 1,
    "yellow": 2,
    "green": 3,
    "cyan": 4,
    "blue": 5,
    "magenta": 6,
    "white": 7,
    "gray": 8,
    "orange": 30,
}

# ---------------------------------------------------------------------------
# COM backend
# ---------------------------------------------------------------------------


class COMBackend(AutoCADBackend):
    """COM / ActiveX backend for AutoCAD-compatible CAD applications."""

    def __init__(self) -> None:
        self._app: Any | None = None
        self._doc: Any | None = None
        self._msp: Any | None = None
        self._cad_type: str = "autocad"
        self._prog_id: str = "AutoCAD.Application"
        self._com_initialised: bool = False

    # ------------------------------------------------------------------
    # Abstract property implementations
    # ------------------------------------------------------------------

    @property
    def name(self) -> str:  # noqa: D401
        return "com"

    @property
    def capabilities(self) -> BackendCapabilities:
        return BackendCapabilities(
            can_read_drawing=True,
            can_modify_entities=True,
            can_create_entities=True,
            can_screenshot=True,
            can_save=True,
            can_plot_pdf=True,
            can_zoom=True,
            can_query_entities=True,
            can_file_operations=True,
            can_undo=True,
        )

    # ------------------------------------------------------------------
    # Lifecycle
    # ------------------------------------------------------------------

    async def initialize(self) -> CommandResult:
        """Initialise COM and connect to the default CAD application."""
        if not _WIN32_AVAILABLE:
            return CommandResult(
                ok=False,
                error=(
                    "COM backend requires Windows with pywin32 installed. "
                    "Install with: pip install pywin32"
                ),
            )
        try:
            pythoncom.CoInitialize()
            self._com_initialised = True
        except Exception as exc:
            return CommandResult(ok=False, error=f"COM initialization failed: {exc}")

        return await self.connect()

    async def status(self) -> CommandResult:
        info: dict[str, Any] = {
            "backend": "com",
            "cad_type": self._cad_type,
            "prog_id": self._prog_id,
            "connected": self._app is not None,
            "has_document": self._doc is not None,
            "capabilities": {k: v for k, v in self.capabilities.__dict__.items()},
        }
        if self._doc is not None:
            try:
                info["document_name"] = self._doc.Name
                info["document_path"] = self._doc.FullName
            except Exception:
                pass
        return CommandResult(ok=True, payload=info)

    # ------------------------------------------------------------------
    # COM helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _to_variant_array(point: tuple[float, ...]) -> Any:
        """Convert a Python tuple/list to a COM VARIANT array of doubles."""
        return win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8, list(point)
        )

    @staticmethod
    def _points_to_variant_array(points: list[list[float]]) -> Any:
        """Flatten 2-D point list to a single VARIANT double array.

        Each point ``[x, y]`` is expanded to ``[x, y, 0]`` when only two
        coordinates are given.
        """
        flat: list[float] = []
        for pt in points:
            flat.append(float(pt[0]))
            flat.append(float(pt[1]))
            flat.append(float(pt[2]) if len(pt) > 2 else 0.0)
        return win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8, flat
        )

    @staticmethod
    def _get_color_index(color: str | int | None) -> int | None:
        """Resolve a colour name or int to an ACI index."""
        if color is None:
            return None
        if isinstance(color, int):
            return color
        return COLOR_MAP.get(str(color).lower().strip())

    def _apply_properties(
        self,
        entity: Any,
        layer: str | None = None,
        color: str | int | None = None,
    ) -> None:
        """Set layer and/or colour on a newly-created entity."""
        if layer is not None:
            try:
                # Ensure layer exists
                try:
                    self._doc.Layers.Item(layer)
                except Exception:
                    self._doc.Layers.Add(layer)
                entity.Layer = layer
            except Exception as exc:
                log.warning("apply_layer_failed", layer=layer, error=str(exc))
        ci = self._get_color_index(color)
        if ci is not None:
            try:
                entity.color = ci
            except Exception as exc:
                log.warning("apply_color_failed", color=color, error=str(exc))

    @staticmethod
    def _safe_get_property(entity: Any, prop_name: str, default: Any = None) -> Any:
        """Try to read a COM property, returning *default* on failure."""
        try:
            return getattr(entity, prop_name)
        except Exception:
            return default

    def _get_entity_by_handle(self, handle: str) -> Any | None:
        """Find an entity in ModelSpace by its Handle string."""
        if self._msp is None:
            return None
        try:
            count = self._msp.Count
            for i in range(count):
                ent = self._msp.Item(i)
                try:
                    if ent.Handle == handle:
                        return ent
                except Exception:
                    continue
        except Exception:
            pass
        return None

    def _validate_connection(self) -> CommandResult | None:
        """Return a ``CommandResult`` error if the COM connection is broken,
        otherwise return ``None`` (meaning everything is fine).
        """
        if self._app is None:
            return CommandResult(ok=False, error="Not connected to any CAD application")
        try:
            # Quick liveness check
            _ = self._app.Name
        except Exception:
            self._app = None
            self._doc = None
            self._msp = None
            return CommandResult(
                ok=False,
                error="Lost connection to the CAD application. Call connect() again.",
            )
        if self._doc is None:
            try:
                self._doc = self._app.ActiveDocument
                self._msp = self._doc.ModelSpace
            except Exception:
                return CommandResult(
                    ok=False,
                    error="No active document. Open or create a drawing first.",
                )
        return None

    # ------------------------------------------------------------------
    # Connection management
    # ------------------------------------------------------------------

    async def connect(self, cad_type: str = "autocad") -> CommandResult:
        """Connect to a specific CAD application via COM."""
        if not _WIN32_AVAILABLE:
            return CommandResult(
                ok=False,
                error="COM backend requires Windows with pywin32",
            )

        cfg = get_cad_config(cad_type)
        self._cad_type = cfg.cad_type
        self._prog_id = cfg.prog_id

        try:
            # Try to attach to a running instance first
            try:
                self._app = win32com.client.GetActiveObject(self._prog_id)
                log.info("com_attached", cad_type=self._cad_type, prog_id=self._prog_id)
            except Exception:
                # Launch a new instance
                self._app = win32com.client.Dispatch(self._prog_id)
                self._app.Visible = True
                log.info(
                    "com_launched",
                    cad_type=self._cad_type,
                    prog_id=self._prog_id,
                    wait=cfg.startup_wait_time,
                )
                time.sleep(cfg.startup_wait_time)

            # Grab active document (may not exist yet)
            try:
                self._doc = self._app.ActiveDocument
                self._msp = self._doc.ModelSpace
            except Exception:
                self._doc = None
                self._msp = None
                log.info("com_no_active_document")

            return CommandResult(
                ok=True,
                payload={
                    "backend": "com",
                    "cad_type": self._cad_type,
                    "prog_id": self._prog_id,
                    "has_document": self._doc is not None,
                    "app_name": self._safe_get_property(self._app, "Name", "Unknown"),
                    "app_version": self._safe_get_property(self._app, "Version", "Unknown"),
                },
            )
        except pywintypes.com_error as exc:
            return CommandResult(
                ok=False,
                error=f"COM connection failed for {self._cad_type} ({self._prog_id}): {exc}",
            )
        except Exception as exc:
            return CommandResult(
                ok=False,
                error=f"Failed to connect to {self._cad_type}: {exc}",
            )

    async def disconnect(self) -> CommandResult:
        """Release COM objects and uninitialise COM."""
        cad = self._cad_type
        try:
            self._msp = None
            self._doc = None
            self._app = None
            if self._com_initialised:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass
                self._com_initialised = False
            return CommandResult(ok=True, payload={"disconnected": cad})
        except Exception as exc:
            return CommandResult(ok=False, error=f"Disconnect error: {exc}")

    async def get_connection_status(self) -> CommandResult:
        """Check COM availability for every supported CAD type."""
        statuses: dict[str, Any] = {}
        for cad in get_supported_cads():
            cfg = get_cad_config(cad)
            status: dict[str, Any] = {"prog_id": cfg.prog_id, "running": False}
            try:
                app = win32com.client.GetActiveObject(cfg.prog_id)
                status["running"] = True
                status["version"] = self._safe_get_property(app, "Version", "Unknown")
            except Exception:
                pass
            statuses[cad] = status
        return CommandResult(ok=True, payload=statuses)

    # ==================================================================
    # Drawing management
    # ==================================================================

    async def drawing_info(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            entity_count = self._msp.Count

            layers: list[str] = []
            for i in range(self._doc.Layers.Count):
                layers.append(self._doc.Layers.Item(i).Name)

            blocks: list[str] = []
            for i in range(self._doc.Blocks.Count):
                blk = self._doc.Blocks.Item(i)
                if not blk.Name.startswith("*"):
                    blocks.append(blk.Name)

            return CommandResult(ok=True, payload={
                "entity_count": entity_count,
                "layers": layers,
                "blocks": blocks,
                "document_name": self._doc.Name,
                "document_path": self._doc.FullName,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"drawing_info COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"drawing_info failed: {exc}")

    async def drawing_save(self, path: str | None = None) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            if path:
                self._doc.SaveAs(path)
            else:
                self._doc.Save()
            return CommandResult(ok=True, payload={
                "saved": self._doc.FullName,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"drawing_save COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"drawing_save failed: {exc}")

    async def drawing_save_as_dxf(self, path: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            # acR12_dxf = 1
            self._doc.SaveAs(path, 1)
            return CommandResult(ok=True, payload={"saved_dxf": path})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"drawing_save_as_dxf COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"drawing_save_as_dxf failed: {exc}")

    async def drawing_create(self, name: str | None = None) -> CommandResult:
        err_conn = self._validate_connection()
        # For create, we only need the app, not a doc
        if self._app is None:
            if err_conn:
                return err_conn
        try:
            new_doc = self._app.Documents.Add()
            self._doc = new_doc
            self._msp = self._doc.ModelSpace
            return CommandResult(ok=True, payload={
                "created": self._doc.Name,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"drawing_create COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"drawing_create failed: {exc}")

    async def drawing_open(self, path: str) -> CommandResult:
        if self._app is None:
            return CommandResult(ok=False, error="Not connected to any CAD application")
        try:
            doc = self._app.Documents.Open(path)
            self._doc = doc
            self._msp = self._doc.ModelSpace
            return CommandResult(ok=True, payload={
                "opened": self._doc.FullName,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"drawing_open COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"drawing_open failed: {exc}")

    async def drawing_purge(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            self._doc.PurgeAll()
            return CommandResult(ok=True, payload={"purged": True})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"drawing_purge COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"drawing_purge failed: {exc}")

    async def drawing_plot_pdf(self, path: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            layout = self._doc.ActiveLayout
            plot = self._doc.Plot

            # Configure plot settings
            layout.ConfigName = "DWG To PDF.pc3"
            layout.PlotType = 1  # acExtents
            layout.UseStandardScale = True
            layout.StandardScale = 0  # acScaleToFit
            layout.CenterPlot = True

            plot.PlotToFile(path)
            return CommandResult(ok=True, payload={"plotted_pdf": path})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"drawing_plot_pdf COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"drawing_plot_pdf failed: {exc}")

    async def drawing_get_variables(self, names: list[str] | None = None) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            if not names:
                names = ["ACADVER", "DWGNAME", "INSUNITS", "LTSCALE", "DIMSCALE"]
            variables: dict[str, Any] = {}
            for var_name in names:
                clean_name = var_name.lstrip("$")
                try:
                    variables[clean_name] = self._doc.GetVariable(clean_name)
                except Exception as exc:
                    variables[clean_name] = f"<error: {exc}>"
            return CommandResult(ok=True, payload=variables)
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"drawing_get_variables COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"drawing_get_variables failed: {exc}")

    # ------------------------------------------------------------------
    # Undo / Redo
    # ------------------------------------------------------------------

    async def undo(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            self._doc.SendCommand("_undo 1\n")
            return CommandResult(ok=True, payload={"undone": True})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"undo COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"undo failed: {exc}")

    async def redo(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            self._doc.SendCommand("_redo\n")
            return CommandResult(ok=True, payload={"redone": True})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"redo COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"redo failed: {exc}")

    # ==================================================================
    # Entity creation
    # ==================================================================

    async def create_line(
        self, x1: float, y1: float, x2: float, y2: float, layer: str | None = None
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            start_pt = self._to_variant_array((x1, y1, 0.0))
            end_pt = self._to_variant_array((x2, y2, 0.0))
            entity = self._msp.AddLine(start_pt, end_pt)
            self._apply_properties(entity, layer=layer)
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "LINE",
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_line COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_line failed: {exc}")

    async def create_circle(
        self, cx: float, cy: float, radius: float, layer: str | None = None
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            center = self._to_variant_array((cx, cy, 0.0))
            entity = self._msp.AddCircle(center, radius)
            self._apply_properties(entity, layer=layer)
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "CIRCLE",
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_circle COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_circle failed: {exc}")

    async def create_polyline(
        self,
        points: list[list[float]],
        closed: bool = False,
        layer: str | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            # AddLightWeightPolyline expects a flat array of doubles [x1,y1,x2,y2,...]
            flat: list[float] = []
            for pt in points:
                flat.append(float(pt[0]))
                flat.append(float(pt[1]))
            pts_variant = win32com.client.VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_R8, flat
            )
            entity = self._msp.AddLightWeightPolyline(pts_variant)
            if closed:
                entity.Closed = True
            self._apply_properties(entity, layer=layer)
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "LWPOLYLINE",
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_polyline COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_polyline failed: {exc}")

    async def create_rectangle(
        self,
        x1: float, y1: float,
        x2: float, y2: float,
        layer: str | None = None,
    ) -> CommandResult:
        """Create a closed rectangular polyline from two diagonal corners."""
        points = [[x1, y1], [x2, y1], [x2, y2], [x1, y2]]
        return await self.create_polyline(points, closed=True, layer=layer)

    async def create_arc(
        self,
        cx: float, cy: float,
        radius: float,
        start_angle: float, end_angle: float,
        layer: str | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            center = self._to_variant_array((cx, cy, 0.0))
            entity = self._msp.AddArc(
                center, radius, math.radians(start_angle), math.radians(end_angle)
            )
            self._apply_properties(entity, layer=layer)
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "ARC",
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_arc COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_arc failed: {exc}")

    async def create_ellipse(
        self,
        cx: float, cy: float,
        major_x: float, major_y: float,
        ratio: float,
        layer: str | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            center = self._to_variant_array((cx, cy, 0.0))
            major_axis = self._to_variant_array((major_x, major_y, 0.0))
            entity = self._msp.AddEllipse(center, major_axis, ratio)
            self._apply_properties(entity, layer=layer)
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "ELLIPSE",
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_ellipse COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_ellipse failed: {exc}")

    async def create_mtext(
        self,
        x: float, y: float,
        width: float,
        text: str,
        height: float = 2.5,
        layer: str | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            insertion = self._to_variant_array((x, y, 0.0))
            entity = self._msp.AddMText(insertion, width, text)
            entity.Height = height
            self._apply_properties(entity, layer=layer)
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "MTEXT",
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_mtext COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_mtext failed: {exc}")

    async def create_hatch(
        self, entity_id: str, pattern: str = "ANSI31"
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            boundary = self._get_entity_by_handle(entity_id)
            if boundary is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")

            # 0 = acHatchPatternTypePreDefined
            hatch = self._msp.AddHatch(0, pattern, True)
            outer_loop = win32com.client.VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, [boundary]
            )
            hatch.AppendOuterLoop(outer_loop)
            hatch.Evaluate()
            return CommandResult(ok=True, payload={
                "handle": hatch.Handle,
                "entity_type": "HATCH",
                "pattern": pattern,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_hatch COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_hatch failed: {exc}")

    # ==================================================================
    # Entity operations
    # ==================================================================

    async def entity_list(self, layer: str | None = None) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            entities: list[dict[str, Any]] = []
            count = self._msp.Count
            for i in range(count):
                ent = self._msp.Item(i)
                try:
                    ent_layer = ent.Layer
                    if layer and ent_layer.upper() != layer.upper():
                        continue
                    entities.append({
                        "handle": ent.Handle,
                        "type": ent.EntityName,
                        "layer": ent_layer,
                    })
                except Exception:
                    continue
            return CommandResult(ok=True, payload=entities)
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_list COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_list failed: {exc}")

    async def entity_count(self, layer: str | None = None) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            if not layer:
                return CommandResult(ok=True, payload={"count": self._msp.Count})

            count = 0
            total = self._msp.Count
            for i in range(total):
                try:
                    if self._msp.Item(i).Layer.upper() == layer.upper():
                        count += 1
                except Exception:
                    continue
            return CommandResult(ok=True, payload={"count": count, "layer": layer})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_count COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_count failed: {exc}")

    async def entity_get(self, entity_id: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")

            props: dict[str, Any] = {
                "handle": ent.Handle,
                "type": ent.EntityName,
                "layer": self._safe_get_property(ent, "Layer"),
                "color": self._safe_get_property(ent, "color"),
                "linetype": self._safe_get_property(ent, "Linetype"),
                "lineweight": self._safe_get_property(ent, "Lineweight"),
                "visible": self._safe_get_property(ent, "Visible"),
            }

            # Type-specific properties
            etype = ent.EntityName.upper()
            if etype == "ACDBLINE":
                props["start_point"] = list(ent.StartPoint)
                props["end_point"] = list(ent.EndPoint)
                props["length"] = ent.Length
            elif etype == "ACDBCIRCLE":
                props["center"] = list(ent.Center)
                props["radius"] = ent.Radius
                props["area"] = ent.Area
            elif etype in ("ACDBLWPOLYLINE", "ACDBPOLYLINE", "ACDB2DPOLYLINE"):
                props["closed"] = self._safe_get_property(ent, "Closed", False)
                props["area"] = self._safe_get_property(ent, "Area")
                props["length"] = self._safe_get_property(ent, "Length")
            elif etype == "ACDBARC":
                props["center"] = list(ent.Center)
                props["radius"] = ent.Radius
                props["start_angle"] = math.degrees(ent.StartAngle)
                props["end_angle"] = math.degrees(ent.EndAngle)
            elif etype in ("ACDBTEXT", "ACDBMTEXT"):
                props["text_string"] = self._safe_get_property(ent, "TextString")
                props["height"] = self._safe_get_property(ent, "Height")
                props["insertion_point"] = list(
                    self._safe_get_property(ent, "InsertionPoint", (0, 0, 0))
                )
            elif etype == "ACDBBLOCKREFERENCE":
                props["block_name"] = self._safe_get_property(ent, "Name")
                props["insertion_point"] = list(
                    self._safe_get_property(ent, "InsertionPoint", (0, 0, 0))
                )
                props["rotation"] = self._safe_get_property(ent, "Rotation")
                props["x_scale"] = self._safe_get_property(ent, "XScaleFactor")
                props["y_scale"] = self._safe_get_property(ent, "YScaleFactor")

            return CommandResult(ok=True, payload=props)
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_get COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_get failed: {exc}")

    async def entity_erase(self, entity_id: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            ent.Delete()
            return CommandResult(ok=True, payload={"erased": entity_id})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_erase COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_erase failed: {exc}")

    async def entity_copy(self, entity_id: str, dx: float, dy: float) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            copy = ent.Copy()
            origin = self._to_variant_array((0.0, 0.0, 0.0))
            offset = self._to_variant_array((dx, dy, 0.0))
            copy.Move(origin, offset)
            return CommandResult(ok=True, payload={
                "handle": copy.Handle,
                "entity_type": copy.EntityName,
                "source_handle": entity_id,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_copy COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_copy failed: {exc}")

    async def entity_move(self, entity_id: str, dx: float, dy: float) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            from_pt = self._to_variant_array((0.0, 0.0, 0.0))
            to_pt = self._to_variant_array((dx, dy, 0.0))
            ent.Move(from_pt, to_pt)
            return CommandResult(ok=True, payload={"moved": entity_id, "dx": dx, "dy": dy})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_move COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_move failed: {exc}")

    async def entity_rotate(
        self, entity_id: str, cx: float, cy: float, angle: float
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            base_pt = self._to_variant_array((cx, cy, 0.0))
            ent.Rotate(base_pt, math.radians(angle))
            return CommandResult(ok=True, payload={
                "rotated": entity_id, "angle_deg": angle,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_rotate COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_rotate failed: {exc}")

    async def entity_scale(
        self, entity_id: str, cx: float, cy: float, factor: float
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            base_pt = self._to_variant_array((cx, cy, 0.0))
            ent.ScaleEntity(base_pt, factor)
            return CommandResult(ok=True, payload={
                "scaled": entity_id, "factor": factor,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_scale COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_scale failed: {exc}")

    async def entity_mirror(
        self, entity_id: str, x1: float, y1: float, x2: float, y2: float
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            pt1 = self._to_variant_array((x1, y1, 0.0))
            pt2 = self._to_variant_array((x2, y2, 0.0))
            mirror_ent = ent.Mirror(pt1, pt2)
            return CommandResult(ok=True, payload={
                "handle": mirror_ent.Handle,
                "entity_type": mirror_ent.EntityName,
                "source_handle": entity_id,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_mirror COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_mirror failed: {exc}")

    async def entity_offset(self, entity_id: str, distance: float) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            result = ent.Offset(distance)
            handles = []
            if result:
                for obj in result:
                    handles.append(self._safe_get_property(obj, "Handle", "unknown"))
            return CommandResult(ok=True, payload={
                "offset_handles": handles,
                "source_handle": entity_id,
                "distance": distance,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_offset COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_offset failed: {exc}")

    async def entity_array(
        self,
        entity_id: str,
        rows: int, cols: int,
        row_dist: float, col_dist: float,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            result = ent.ArrayRectangular(rows, cols, 1, row_dist, col_dist, 0)
            handles = []
            if result:
                for obj in result:
                    handles.append(self._safe_get_property(obj, "Handle", "unknown"))
            return CommandResult(ok=True, payload={
                "array_handles": handles,
                "source_handle": entity_id,
                "rows": rows,
                "cols": cols,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"entity_array COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"entity_array failed: {exc}")

    # ==================================================================
    # Layer operations
    # ==================================================================

    async def layer_list(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            layers: list[dict[str, Any]] = []
            for i in range(self._doc.Layers.Count):
                lyr = self._doc.Layers.Item(i)
                layers.append({
                    "name": lyr.Name,
                    "color": self._safe_get_property(lyr, "color"),
                    "linetype": self._safe_get_property(lyr, "Linetype"),
                    "frozen": self._safe_get_property(lyr, "Freeze", False),
                    "locked": self._safe_get_property(lyr, "Lock", False),
                    "on": self._safe_get_property(lyr, "LayerOn", True),
                })
            return CommandResult(ok=True, payload=layers)
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"layer_list COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"layer_list failed: {exc}")

    async def layer_create(
        self, name: str, color: str | int = "white", linetype: str = "CONTINUOUS"
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            lyr = self._doc.Layers.Add(name)
            ci = self._get_color_index(color)
            if ci is not None:
                lyr.color = ci
            try:
                lyr.Linetype = linetype
            except Exception:
                log.warning("layer_linetype_not_found", linetype=linetype)
            return CommandResult(ok=True, payload={
                "name": name, "color": ci, "linetype": linetype,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"layer_create COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"layer_create failed: {exc}")

    async def layer_set_current(self, name: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            lyr = self._doc.Layers.Item(name)
            self._doc.ActiveLayer = lyr
            return CommandResult(ok=True, payload={"current_layer": name})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"layer_set_current COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"layer_set_current failed: {exc}")

    async def layer_set_properties(
        self,
        name: str,
        color: str | int | None = None,
        linetype: str | None = None,
        lineweight: str | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            lyr = self._doc.Layers.Item(name)
            if color is not None:
                ci = self._get_color_index(color)
                if ci is not None:
                    lyr.color = ci
            if linetype is not None:
                lyr.Linetype = linetype
            if lineweight is not None:
                lyr.Lineweight = int(lineweight)
            return CommandResult(ok=True, payload={"layer": name, "updated": True})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"layer_set_properties COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"layer_set_properties failed: {exc}")

    async def layer_freeze(self, name: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            lyr = self._doc.Layers.Item(name)
            lyr.Freeze = True
            return CommandResult(ok=True, payload={"layer": name, "frozen": True})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"layer_freeze COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"layer_freeze failed: {exc}")

    async def layer_thaw(self, name: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            lyr = self._doc.Layers.Item(name)
            lyr.Freeze = False
            return CommandResult(ok=True, payload={"layer": name, "frozen": False})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"layer_thaw COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"layer_thaw failed: {exc}")

    async def layer_lock(self, name: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            lyr = self._doc.Layers.Item(name)
            lyr.Lock = True
            return CommandResult(ok=True, payload={"layer": name, "locked": True})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"layer_lock COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"layer_lock failed: {exc}")

    async def layer_unlock(self, name: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            lyr = self._doc.Layers.Item(name)
            lyr.Lock = False
            return CommandResult(ok=True, payload={"layer": name, "locked": False})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"layer_unlock COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"layer_unlock failed: {exc}")

    # ==================================================================
    # Block operations
    # ==================================================================

    async def block_list(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            blocks: list[str] = []
            for i in range(self._doc.Blocks.Count):
                blk = self._doc.Blocks.Item(i)
                name = blk.Name
                # Skip anonymous/internal blocks (start with *)
                if not name.startswith("*"):
                    blocks.append(name)
            return CommandResult(ok=True, payload=blocks)
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"block_list COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"block_list failed: {exc}")

    async def block_insert(
        self,
        name: str,
        x: float, y: float,
        scale: float = 1.0,
        rotation: float = 0.0,
        block_id: str | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            insertion = self._to_variant_array((x, y, 0.0))
            entity = self._msp.InsertBlock(
                insertion, name, scale, scale, scale, math.radians(rotation)
            )
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "INSERT",
                "block_name": name,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"block_insert COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"block_insert failed: {exc}")

    async def block_get_attributes(self, entity_id: str) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")

            attribs_list: list[dict[str, str]] = []
            try:
                attribs = ent.GetAttributes()
                for attr in attribs:
                    attribs_list.append({
                        "tag": attr.TagString,
                        "value": attr.TextString,
                    })
            except Exception:
                return CommandResult(ok=False, error="Entity has no attributes or is not an INSERT")

            return CommandResult(ok=True, payload={
                "handle": entity_id,
                "attributes": attribs_list,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"block_get_attributes COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"block_get_attributes failed: {exc}")

    async def block_update_attribute(
        self, entity_id: str, tag: str, value: str
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")

            attribs = ent.GetAttributes()
            for attr in attribs:
                if attr.TagString.upper() == tag.upper():
                    attr.TextString = value
                    return CommandResult(ok=True, payload={
                        "handle": entity_id,
                        "tag": tag,
                        "new_value": value,
                    })
            return CommandResult(ok=False, error=f"Attribute tag '{tag}' not found")
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"block_update_attribute COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"block_update_attribute failed: {exc}")

    # ==================================================================
    # Annotation
    # ==================================================================

    async def create_text(
        self,
        x: float, y: float,
        text: str,
        height: float = 2.5,
        rotation: float = 0.0,
        layer: str | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            insertion = self._to_variant_array((x, y, 0.0))
            entity = self._msp.AddText(text, insertion, height)
            if rotation != 0.0:
                entity.Rotation = math.radians(rotation)
            self._apply_properties(entity, layer=layer)
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "TEXT",
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_text COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_text failed: {exc}")

    async def create_dimension_linear(
        self,
        x1: float, y1: float,
        x2: float, y2: float,
        dim_x: float, dim_y: float,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            pt1 = self._to_variant_array((x1, y1, 0.0))
            pt2 = self._to_variant_array((x2, y2, 0.0))
            text_pt = self._to_variant_array((dim_x, dim_y, 0.0))
            entity = self._msp.AddDimAligned(pt1, pt2, text_pt)
            return CommandResult(ok=True, payload={
                "handle": entity.Handle,
                "entity_type": "DIMENSION",
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"create_dimension_linear COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"create_dimension_linear failed: {exc}")

    # ==================================================================
    # View
    # ==================================================================

    async def zoom_extents(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            self._app.ZoomExtents()
            return CommandResult(ok=True, payload={"zoomed": "extents"})
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"zoom_extents COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"zoom_extents failed: {exc}")

    async def zoom_window(
        self, x1: float, y1: float, x2: float, y2: float
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            pt1 = self._to_variant_array((x1, y1, 0.0))
            pt2 = self._to_variant_array((x2, y2, 0.0))
            self._app.ZoomWindow(pt1, pt2)
            return CommandResult(ok=True, payload={
                "zoomed": "window",
                "lower_left": [x1, y1],
                "upper_right": [x2, y2],
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"zoom_window COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"zoom_window failed: {exc}")

    async def get_screenshot(self) -> CommandResult:
        """Capture the CAD application window as a base64-encoded PNG."""
        if not _WIN32_AVAILABLE:
            return CommandResult(ok=False, error="Screenshot requires Windows with pywin32")
        try:
            import ctypes
            from ctypes import wintypes

            # Find the CAD window
            hwnd = self._find_cad_window()
            if not hwnd:
                return CommandResult(ok=False, error="CAD application window not found")

            # Get window dimensions
            rect = wintypes.RECT()
            ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
            width = rect.right - rect.left
            height = rect.bottom - rect.top

            if width <= 0 or height <= 0:
                return CommandResult(ok=False, error="Invalid window dimensions")

            # Create device contexts
            hwnd_dc = ctypes.windll.user32.GetDC(hwnd)
            mfc_dc = ctypes.windll.gdi32.CreateCompatibleDC(hwnd_dc)
            bitmap = ctypes.windll.gdi32.CreateCompatibleBitmap(hwnd_dc, width, height)
            old_bitmap = ctypes.windll.gdi32.SelectObject(mfc_dc, bitmap)

            # Capture via PrintWindow (works even if window is partially hidden)
            ctypes.windll.user32.PrintWindow(hwnd, mfc_dc, 0x00000002)  # PW_RENDERFULLCONTENT

            # Convert to PIL Image
            try:
                from PIL import Image

                # Create BITMAPINFOHEADER
                import struct

                bmi_size = 40
                bmi = struct.pack(
                    "IiiHHIIiiII",
                    bmi_size, width, -height, 1, 32, 0, 0, 0, 0, 0, 0,
                )
                buffer = ctypes.create_string_buffer(width * height * 4)
                ctypes.windll.gdi32.GetDIBits(
                    mfc_dc, bitmap, 0, height, buffer, bmi, 0
                )
                img = Image.frombuffer("RGBA", (width, height), buffer, "raw", "BGRA", 0, 1)

                buf = io.BytesIO()
                img.save(buf, format="PNG")
                b64 = base64.b64encode(buf.getvalue()).decode("ascii")
            finally:
                # Clean up GDI resources
                ctypes.windll.gdi32.SelectObject(mfc_dc, old_bitmap)
                ctypes.windll.gdi32.DeleteObject(bitmap)
                ctypes.windll.gdi32.DeleteDC(mfc_dc)
                ctypes.windll.user32.ReleaseDC(hwnd, hwnd_dc)

            return CommandResult(ok=True, payload=b64)

        except ImportError:
            return CommandResult(ok=False, error="Screenshot requires Pillow: pip install Pillow")
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"get_screenshot COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"get_screenshot failed: {exc}")

    def _find_cad_window(self) -> int | None:
        """Find the CAD application's main window handle."""
        if not _WIN32_AVAILABLE:
            return None
        try:
            # Try to get the HWND from the COM application object
            try:
                return self._app.HWND
            except Exception:
                pass

            # Fall back to window title search
            cad_keywords = {
                "autocad": "autocad",
                "zwcad": "zwcad",
                "gcad": "gcad",
                "bricscad": "bricscad",
            }
            keyword = cad_keywords.get(self._cad_type, "autocad")
            windows: list[int] = []

            def callback(hwnd: int, result: list) -> bool:
                if win32gui.IsWindowVisible(hwnd):
                    text = win32gui.GetWindowText(hwnd).lower()
                    if keyword in text:
                        result.append(hwnd)
                return True

            win32gui.EnumWindows(callback, windows)
            return windows[0] if windows else None
        except Exception:
            return None

    # ==================================================================
    # Query operations
    # ==================================================================

    async def query_entity_properties(self, entity_id: str) -> CommandResult:
        """Full property dump for an entity."""
        err = self._validate_connection()
        if err:
            return err
        try:
            ent = self._get_entity_by_handle(entity_id)
            if ent is None:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")

            props: dict[str, Any] = {
                "handle": ent.Handle,
                "type": ent.EntityName,
                "layer": self._safe_get_property(ent, "Layer"),
                "color": self._safe_get_property(ent, "color"),
                "linetype": self._safe_get_property(ent, "Linetype"),
                "linetype_scale": self._safe_get_property(ent, "LinetypeScale"),
                "lineweight": self._safe_get_property(ent, "Lineweight"),
                "visible": self._safe_get_property(ent, "Visible"),
                "plotstyle_name": self._safe_get_property(ent, "PlotStyleName"),
            }

            # Geometry properties
            for prop in (
                "StartPoint", "EndPoint", "Center", "Radius", "Length", "Area",
                "Closed", "Height", "Width", "TextString", "Rotation",
                "InsertionPoint", "Name", "XScaleFactor", "YScaleFactor",
                "ZScaleFactor", "StartAngle", "EndAngle", "Normal",
                "MajorAxis", "MinorAxis", "RadiusRatio",
            ):
                val = self._safe_get_property(ent, prop)
                if val is not None:
                    # Convert COM tuples to lists
                    if isinstance(val, tuple):
                        val = list(val)
                    props[prop.lower()] = val

            # Bounding box
            try:
                min_pt, max_pt = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 0, 0]), \
                                 win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 0, 0])
                ent.GetBoundingBox(min_pt, max_pt)
                props["bounding_box"] = {
                    "min": list(min_pt),
                    "max": list(max_pt),
                }
            except Exception:
                pass

            return CommandResult(ok=True, payload=props)
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"query_entity_properties COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"query_entity_properties failed: {exc}")

    async def query_drawing_summary(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            type_counts: dict[str, int] = defaultdict(int)
            layer_counts: dict[str, int] = defaultdict(int)

            count = self._msp.Count
            for i in range(count):
                try:
                    ent = self._msp.Item(i)
                    etype = ent.EntityName
                    lyr = ent.Layer
                    type_counts[etype] += 1
                    layer_counts[lyr] += 1
                except Exception:
                    continue

            return CommandResult(ok=True, payload={
                "total_entities": count,
                "by_type": dict(type_counts),
                "by_layer": dict(layer_counts),
                "layer_count": self._doc.Layers.Count,
                "block_count": self._doc.Blocks.Count,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"query_drawing_summary COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"query_drawing_summary failed: {exc}")

    # ==================================================================
    # Search operations
    # ==================================================================

    async def search_text(
        self, pattern: str, case_sensitive: bool = False
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            flags = 0 if case_sensitive else re.IGNORECASE
            regex = re.compile(pattern, flags)
            results: list[dict[str, Any]] = []

            count = self._msp.Count
            for i in range(count):
                try:
                    ent = self._msp.Item(i)
                    etype = ent.EntityName.upper()
                    if etype not in ("ACDBTEXT", "ACDBMTEXT"):
                        continue
                    text_str = self._safe_get_property(ent, "TextString", "")
                    if text_str and regex.search(text_str):
                        results.append({
                            "handle": ent.Handle,
                            "type": ent.EntityName,
                            "text": text_str,
                            "layer": self._safe_get_property(ent, "Layer"),
                            "insertion_point": list(
                                self._safe_get_property(ent, "InsertionPoint", (0, 0, 0))
                            ),
                        })
                except Exception:
                    continue

            return CommandResult(ok=True, payload={
                "pattern": pattern,
                "count": len(results),
                "results": results,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"search_text COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"search_text failed: {exc}")

    async def search_by_type_and_layer(
        self,
        entity_type: str | None = None,
        layer: str | None = None,
        color: int | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            results: list[dict[str, Any]] = []
            count = self._msp.Count

            for i in range(count):
                try:
                    ent = self._msp.Item(i)
                    ent_type = ent.EntityName
                    ent_layer = ent.Layer
                    ent_color = self._safe_get_property(ent, "color")

                    if entity_type and ent_type.upper() != entity_type.upper():
                        continue
                    if layer and ent_layer.upper() != layer.upper():
                        continue
                    if color is not None and ent_color != color:
                        continue

                    results.append({
                        "handle": ent.Handle,
                        "type": ent_type,
                        "layer": ent_layer,
                        "color": ent_color,
                    })
                except Exception:
                    continue

            return CommandResult(ok=True, payload={
                "count": len(results),
                "results": results,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"search_by_type_and_layer COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"search_by_type_and_layer failed: {exc}")

    # ==================================================================
    # Bulk operations
    # ==================================================================

    async def bulk_erase(self, handles: list) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        erased: list[str] = []
        errors: list[str] = []
        for h in handles:
            try:
                ent = self._get_entity_by_handle(str(h))
                if ent is None:
                    errors.append(f"Not found: {h}")
                    continue
                ent.Delete()
                erased.append(str(h))
            except Exception as exc:
                errors.append(f"{h}: {exc}")
        return CommandResult(ok=True, payload={
            "erased": erased,
            "errors": errors,
            "erased_count": len(erased),
        })

    async def bulk_set_property(
        self, handles: list, property_name: str, value: Any
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        updated: list[str] = []
        errors: list[str] = []
        for h in handles:
            try:
                ent = self._get_entity_by_handle(str(h))
                if ent is None:
                    errors.append(f"Not found: {h}")
                    continue
                # Map friendly names to COM property names
                prop_map = {
                    "layer": "Layer",
                    "color": "color",
                    "linetype": "Linetype",
                    "lineweight": "Lineweight",
                }
                com_prop = prop_map.get(property_name.lower(), property_name)
                setattr(ent, com_prop, value)
                updated.append(str(h))
            except Exception as exc:
                errors.append(f"{h}: {exc}")
        return CommandResult(ok=True, payload={
            "updated": updated,
            "errors": errors,
            "updated_count": len(updated),
        })

    # ==================================================================
    # Export operations
    # ==================================================================

    async def export_entity_data(
        self, layer: str | None = None, entity_type: str | None = None
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            entities: list[dict[str, Any]] = []
            count = self._msp.Count

            for i in range(count):
                try:
                    ent = self._msp.Item(i)
                    ent_type = ent.EntityName
                    ent_layer = ent.Layer

                    if layer and ent_layer.upper() != layer.upper():
                        continue
                    if entity_type and ent_type.upper() != entity_type.upper():
                        continue

                    data: dict[str, Any] = {
                        "handle": ent.Handle,
                        "type": ent_type,
                        "layer": ent_layer,
                        "color": self._safe_get_property(ent, "color"),
                        "linetype": self._safe_get_property(ent, "Linetype"),
                    }

                    # Add type-specific data
                    for prop in (
                        "StartPoint", "EndPoint", "Center", "Radius",
                        "Length", "Area", "Closed", "Height", "Width",
                        "TextString", "Rotation", "InsertionPoint", "Name",
                    ):
                        val = self._safe_get_property(ent, prop)
                        if val is not None:
                            if isinstance(val, tuple):
                                val = list(val)
                            data[prop.lower()] = val

                    entities.append(data)
                except Exception:
                    continue

            return CommandResult(ok=True, payload={
                "count": len(entities),
                "entities": entities,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"export_entity_data COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"export_entity_data failed: {exc}")

    async def export_drawing_statistics(self) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            type_counts: dict[str, int] = defaultdict(int)
            layer_counts: dict[str, int] = defaultdict(int)

            count = self._msp.Count
            for i in range(count):
                try:
                    ent = self._msp.Item(i)
                    type_counts[ent.EntityName] += 1
                    layer_counts[ent.Layer] += 1
                except Exception:
                    continue

            return CommandResult(ok=True, payload={
                "total_entities": count,
                "by_type": dict(type_counts),
                "by_layer": dict(layer_counts),
                "layer_definitions": self._doc.Layers.Count,
                "block_definitions": self._doc.Blocks.Count,
                "document_name": self._doc.Name,
                "document_path": self._doc.FullName,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"export_drawing_statistics COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"export_drawing_statistics failed: {exc}")

    async def export_to_excel(
        self,
        filename: str = "drawing_data.xlsx",
        output_dir: str | None = None,
    ) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        try:
            from openpyxl import Workbook

            wb = Workbook()

            # --- Sheet 1: Entities ---
            ws_ent = wb.active
            ws_ent.title = "Entities"
            headers = [
                "Handle", "Type", "Layer", "Color", "Linetype",
                "StartPoint", "EndPoint", "Center", "Radius",
                "Length", "Area", "TextString",
            ]
            ws_ent.append(headers)

            count = self._msp.Count
            for i in range(count):
                try:
                    ent = self._msp.Item(i)
                    row = [
                        ent.Handle,
                        ent.EntityName,
                        self._safe_get_property(ent, "Layer", ""),
                        self._safe_get_property(ent, "color", ""),
                        self._safe_get_property(ent, "Linetype", ""),
                    ]
                    for prop in ("StartPoint", "EndPoint", "Center"):
                        val = self._safe_get_property(ent, prop)
                        row.append(str(list(val)) if val is not None else "")
                    for prop in ("Radius", "Length", "Area"):
                        row.append(self._safe_get_property(ent, prop, ""))
                    row.append(self._safe_get_property(ent, "TextString", ""))
                    ws_ent.append(row)
                except Exception:
                    continue

            # --- Sheet 2: Layers ---
            ws_lyr = wb.create_sheet("Layers")
            ws_lyr.append(["Name", "Color", "Linetype", "Frozen", "Locked", "On"])
            for i in range(self._doc.Layers.Count):
                try:
                    lyr = self._doc.Layers.Item(i)
                    ws_lyr.append([
                        lyr.Name,
                        self._safe_get_property(lyr, "color", ""),
                        self._safe_get_property(lyr, "Linetype", ""),
                        self._safe_get_property(lyr, "Freeze", ""),
                        self._safe_get_property(lyr, "Lock", ""),
                        self._safe_get_property(lyr, "LayerOn", ""),
                    ])
                except Exception:
                    continue

            # --- Sheet 3: Summary ---
            ws_sum = wb.create_sheet("Summary")
            ws_sum.append(["Statistic", "Value"])
            ws_sum.append(["Total Entities", count])
            ws_sum.append(["Total Layers", self._doc.Layers.Count])
            ws_sum.append(["Total Blocks", self._doc.Blocks.Count])
            ws_sum.append(["Document", self._doc.Name])

            # Determine output path
            if output_dir:
                out_path = Path(output_dir) / filename
            else:
                out_path = Path(self._doc.FullName).parent / filename

            out_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(out_path))

            return CommandResult(ok=True, payload={
                "file": str(out_path),
                "entity_rows": count,
                "sheets": ["Entities", "Layers", "Summary"],
            })
        except ImportError:
            return CommandResult(
                ok=False,
                error="export_to_excel requires openpyxl: pip install openpyxl",
            )
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"export_to_excel COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"export_to_excel failed: {exc}")

    # ==================================================================
    # Batch drawing operations
    # ==================================================================

    async def batch_draw_lines(self, lines: list[dict]) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        handles: list[str] = []
        errors: list[str] = []
        try:
            self._app.ActiveDocument.SendCommand("_regen\n") if False else None  # placeholder
            for idx, line_def in enumerate(lines):
                try:
                    x1 = float(line_def["x1"])
                    y1 = float(line_def["y1"])
                    x2 = float(line_def["x2"])
                    y2 = float(line_def["y2"])
                    start_pt = self._to_variant_array((x1, y1, 0.0))
                    end_pt = self._to_variant_array((x2, y2, 0.0))
                    entity = self._msp.AddLine(start_pt, end_pt)
                    self._apply_properties(
                        entity,
                        layer=line_def.get("layer"),
                        color=line_def.get("color"),
                    )
                    handles.append(entity.Handle)
                except Exception as exc:
                    errors.append(f"line[{idx}]: {exc}")

            # Force a single regen at the end for performance
            try:
                self._doc.Regen(1)  # acActiveViewport = 1
            except Exception:
                pass

            return CommandResult(ok=True, payload={
                "handles": handles,
                "created": len(handles),
                "errors": errors,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"batch_draw_lines COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"batch_draw_lines failed: {exc}")

    async def batch_draw_circles(self, circles: list[dict]) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        handles: list[str] = []
        errors: list[str] = []
        try:
            for idx, circ_def in enumerate(circles):
                try:
                    cx = float(circ_def["cx"])
                    cy = float(circ_def["cy"])
                    radius = float(circ_def["radius"])
                    center = self._to_variant_array((cx, cy, 0.0))
                    entity = self._msp.AddCircle(center, radius)
                    self._apply_properties(
                        entity,
                        layer=circ_def.get("layer"),
                        color=circ_def.get("color"),
                    )
                    handles.append(entity.Handle)
                except Exception as exc:
                    errors.append(f"circle[{idx}]: {exc}")

            try:
                self._doc.Regen(1)
            except Exception:
                pass

            return CommandResult(ok=True, payload={
                "handles": handles,
                "created": len(handles),
                "errors": errors,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"batch_draw_circles COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"batch_draw_circles failed: {exc}")

    async def batch_draw_rectangles(self, rectangles: list[dict]) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        handles: list[str] = []
        errors: list[str] = []
        try:
            for idx, rect_def in enumerate(rectangles):
                try:
                    x1 = float(rect_def["x1"])
                    y1 = float(rect_def["y1"])
                    x2 = float(rect_def["x2"])
                    y2 = float(rect_def["y2"])
                    pts = [x1, y1, x2, y1, x2, y2, x1, y2]
                    pts_variant = win32com.client.VARIANT(
                        pythoncom.VT_ARRAY | pythoncom.VT_R8, pts
                    )
                    entity = self._msp.AddLightWeightPolyline(pts_variant)
                    entity.Closed = True
                    self._apply_properties(
                        entity,
                        layer=rect_def.get("layer"),
                        color=rect_def.get("color"),
                    )
                    handles.append(entity.Handle)
                except Exception as exc:
                    errors.append(f"rectangle[{idx}]: {exc}")

            try:
                self._doc.Regen(1)
            except Exception:
                pass

            return CommandResult(ok=True, payload={
                "handles": handles,
                "created": len(handles),
                "errors": errors,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"batch_draw_rectangles COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"batch_draw_rectangles failed: {exc}")

    async def batch_draw_polylines(self, polylines: list[dict]) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        handles: list[str] = []
        errors: list[str] = []
        try:
            for idx, pl_def in enumerate(polylines):
                try:
                    points = pl_def["points"]
                    flat: list[float] = []
                    for pt in points:
                        flat.append(float(pt[0]))
                        flat.append(float(pt[1]))
                    pts_variant = win32com.client.VARIANT(
                        pythoncom.VT_ARRAY | pythoncom.VT_R8, flat
                    )
                    entity = self._msp.AddLightWeightPolyline(pts_variant)
                    if pl_def.get("closed", False):
                        entity.Closed = True
                    self._apply_properties(
                        entity,
                        layer=pl_def.get("layer"),
                        color=pl_def.get("color"),
                    )
                    handles.append(entity.Handle)
                except Exception as exc:
                    errors.append(f"polyline[{idx}]: {exc}")

            try:
                self._doc.Regen(1)
            except Exception:
                pass

            return CommandResult(ok=True, payload={
                "handles": handles,
                "created": len(handles),
                "errors": errors,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"batch_draw_polylines COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"batch_draw_polylines failed: {exc}")

    async def batch_draw_texts(self, texts: list[dict]) -> CommandResult:
        err = self._validate_connection()
        if err:
            return err
        handles: list[str] = []
        errors: list[str] = []
        try:
            for idx, txt_def in enumerate(texts):
                try:
                    x = float(txt_def["x"])
                    y = float(txt_def["y"])
                    text = str(txt_def["text"])
                    height = float(txt_def.get("height", 2.5))
                    rotation = float(txt_def.get("rotation", 0.0))

                    insertion = self._to_variant_array((x, y, 0.0))
                    entity = self._msp.AddText(text, insertion, height)
                    if rotation != 0.0:
                        entity.Rotation = math.radians(rotation)
                    self._apply_properties(
                        entity,
                        layer=txt_def.get("layer"),
                        color=txt_def.get("color"),
                    )
                    handles.append(entity.Handle)
                except Exception as exc:
                    errors.append(f"text[{idx}]: {exc}")

            try:
                self._doc.Regen(1)
            except Exception:
                pass

            return CommandResult(ok=True, payload={
                "handles": handles,
                "created": len(handles),
                "errors": errors,
            })
        except pywintypes.com_error as exc:
            return CommandResult(ok=False, error=f"batch_draw_texts COM error: {exc}")
        except Exception as exc:
            return CommandResult(ok=False, error=f"batch_draw_texts failed: {exc}")
