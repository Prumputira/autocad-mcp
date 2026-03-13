"""Headless DXF backend using ezdxf — no AutoCAD needed."""

from __future__ import annotations

import math
import os
from pathlib import Path
from typing import Any

import ezdxf
import structlog

from autocad_mcp.backends.base import AutoCADBackend, BackendCapabilities, CommandResult
from autocad_mcp.screenshot import MatplotlibScreenshotProvider

log = structlog.get_logger()


class EzdxfBackend(AutoCADBackend):
    """Pure-Python DXF generation via ezdxf."""

    def __init__(self):
        self._doc: ezdxf.document.Drawing | None = None
        self._msp = None  # modelspace
        self._save_path: str | None = None
        self._screenshot = MatplotlibScreenshotProvider()
        self._entity_counter = 0

    @property
    def name(self) -> str:
        return "ezdxf"

    @property
    def capabilities(self) -> BackendCapabilities:
        return BackendCapabilities(
            can_read_drawing=True,
            can_modify_entities=True,
            can_create_entities=True,
            can_screenshot=True,
            can_save=True,
            can_plot_pdf=False,
            can_zoom=False,  # No viewport in headless
            can_query_entities=True,
            can_file_operations=True,
            can_undo=False,
        )

    async def initialize(self) -> CommandResult:
        self._doc = ezdxf.new("R2013")
        self._msp = self._doc.modelspace()
        self._screenshot.doc = self._doc
        # Ensure default layer 0 is on
        try:
            self._doc.layers.get("0").on = True
        except Exception:
            pass
        return CommandResult(ok=True, payload={"backend": "ezdxf", "version": ezdxf.__version__})

    async def status(self) -> CommandResult:
        entity_count = len(self._msp) if self._msp else 0
        return CommandResult(ok=True, payload={
            "backend": "ezdxf",
            "version": ezdxf.__version__,
            "has_document": self._doc is not None,
            "entity_count": entity_count,
            "save_path": self._save_path,
            "capabilities": {k: v for k, v in self.capabilities.__dict__.items()},
        })

    def _next_id(self) -> str:
        self._entity_counter += 1
        return f"ezdxf_{self._entity_counter}"

    def _ensure_layer(self, layer: str | None):
        if layer and layer not in self._doc.layers:
            self._doc.layers.add(layer)

    # --- Drawing management ---

    async def drawing_info(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        layers = [l.dxf.name for l in self._doc.layers]
        entity_count = len(self._msp)
        blocks = [b.name for b in self._doc.blocks if not b.name.startswith("*")]
        return CommandResult(ok=True, payload={
            "entity_count": entity_count,
            "layers": layers,
            "blocks": blocks,
            "dxf_version": self._doc.dxfversion,
            "save_path": self._save_path,
        })

    async def drawing_save(self, path: str | None = None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        save_path = path or self._save_path
        if not save_path:
            return CommandResult(ok=False, error="No save path specified")
        self._doc.saveas(save_path)
        self._save_path = save_path
        return CommandResult(ok=True, payload={"path": save_path})

    async def drawing_save_as_dxf(self, path: str) -> CommandResult:
        return await self.drawing_save(path)

    async def drawing_create(self, name: str | None = None) -> CommandResult:
        self._doc = ezdxf.new("R2013")
        self._msp = self._doc.modelspace()
        self._screenshot.doc = self._doc
        self._entity_counter = 0
        self._save_path = f"{name}.dxf" if name else None
        return CommandResult(ok=True, payload={"name": name or "untitled"})

    async def drawing_purge(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        # ezdxf doesn't have a direct purge; just report
        return CommandResult(ok=True, payload={"purged": True})

    async def drawing_open(self, path: str) -> CommandResult:
        try:
            self._doc = ezdxf.readfile(path)
            self._msp = self._doc.modelspace()
            self._screenshot.doc = self._doc
            self._save_path = path
            return CommandResult(ok=True, payload={"path": path})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def drawing_get_variables(self, names: list[str] | None = None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        result = {}
        header = self._doc.header
        for name in (names or []):
            try:
                result[name] = str(header[name])
            except (KeyError, ezdxf.DXFKeyError):
                result[name] = None
        return CommandResult(ok=True, payload=result)

    # --- Entity operations ---

    async def create_line(self, x1, y1, x2, y2, layer=None) -> CommandResult:
        self._ensure_layer(layer)
        e = self._msp.add_line((x1, y1), (x2, y2), dxfattribs={"layer": layer or "0"})
        return CommandResult(ok=True, payload={"entity_type": "LINE", "handle": e.dxf.handle})

    async def create_circle(self, cx, cy, radius, layer=None) -> CommandResult:
        self._ensure_layer(layer)
        e = self._msp.add_circle((cx, cy), radius, dxfattribs={"layer": layer or "0"})
        return CommandResult(ok=True, payload={"entity_type": "CIRCLE", "handle": e.dxf.handle})

    async def create_polyline(self, points, closed=False, layer=None) -> CommandResult:
        self._ensure_layer(layer)
        pts = [(p[0], p[1]) for p in points]
        e = self._msp.add_lwpolyline(pts, close=closed, dxfattribs={"layer": layer or "0"})
        return CommandResult(ok=True, payload={"entity_type": "LWPOLYLINE", "handle": e.dxf.handle})

    async def create_rectangle(self, x1, y1, x2, y2, layer=None) -> CommandResult:
        pts = [(x1, y1), (x2, y1), (x2, y2), (x1, y2)]
        return await self.create_polyline(pts, closed=True, layer=layer)

    async def create_arc(self, cx, cy, radius, start_angle, end_angle, layer=None) -> CommandResult:
        self._ensure_layer(layer)
        e = self._msp.add_arc((cx, cy), radius, start_angle, end_angle, dxfattribs={"layer": layer or "0"})
        return CommandResult(ok=True, payload={"entity_type": "ARC", "handle": e.dxf.handle})

    async def create_ellipse(self, cx, cy, major_x, major_y, ratio, layer=None) -> CommandResult:
        self._ensure_layer(layer)
        e = self._msp.add_ellipse(
            (cx, cy), major_axis=(major_x - cx, major_y - cy, 0), ratio=ratio,
            dxfattribs={"layer": layer or "0"},
        )
        return CommandResult(ok=True, payload={"entity_type": "ELLIPSE", "handle": e.dxf.handle})

    async def create_mtext(self, x, y, width, text, height=2.5, layer=None) -> CommandResult:
        self._ensure_layer(layer)
        e = self._msp.add_mtext(text, dxfattribs={
            "insert": (x, y),
            "char_height": height,
            "width": width,
            "layer": layer or "0",
        })
        return CommandResult(ok=True, payload={"entity_type": "MTEXT", "handle": e.dxf.handle})

    async def entity_list(self, layer=None) -> CommandResult:
        entities = []
        for e in self._msp:
            if layer and e.dxf.get("layer", "0") != layer:
                continue
            entities.append({
                "type": e.dxftype(),
                "handle": e.dxf.handle,
                "layer": e.dxf.get("layer", "0"),
            })
        return CommandResult(ok=True, payload={"entities": entities, "count": len(entities)})

    async def entity_count(self, layer=None) -> CommandResult:
        if layer:
            count = sum(1 for e in self._msp if e.dxf.get("layer", "0") == layer)
        else:
            count = len(self._msp)
        return CommandResult(ok=True, payload={"count": count})

    async def entity_get(self, entity_id) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            info = {"type": e.dxftype(), "handle": e.dxf.handle, "layer": e.dxf.get("layer", "0")}
            # Add type-specific info
            if e.dxftype() == "LINE":
                info["start"] = list(e.dxf.start)[:2]
                info["end"] = list(e.dxf.end)[:2]
            elif e.dxftype() == "CIRCLE":
                info["center"] = list(e.dxf.center)[:2]
                info["radius"] = e.dxf.radius
            return CommandResult(ok=True, payload=info)
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_erase(self, entity_id) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                # Try "last" keyword
                if entity_id == "last" and len(self._msp) > 0:
                    entities = list(self._msp)
                    e = entities[-1]
                else:
                    return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            self._msp.delete_entity(e)
            return CommandResult(ok=True, payload={"erased": entity_id})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_copy(self, entity_id, dx, dy) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            copy = e.copy()
            self._msp.add_entity(copy)
            copy.translate(dx, dy, 0)
            return CommandResult(ok=True, payload={"handle": copy.dxf.handle})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_move(self, entity_id, dx, dy) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            e.translate(dx, dy, 0)
            return CommandResult(ok=True, payload={"moved": entity_id})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_rotate(self, entity_id, cx, cy, angle) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            from ezdxf.math import Matrix44
            m = Matrix44.z_rotate(math.radians(angle))
            # Translate to origin, rotate, translate back
            e.translate(-cx, -cy, 0)
            e.transform(m)
            e.translate(cx, cy, 0)
            return CommandResult(ok=True, payload={"rotated": entity_id})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_scale(self, entity_id, cx, cy, factor) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            from ezdxf.math import Matrix44
            m = Matrix44.scale(factor, factor, factor)
            e.translate(-cx, -cy, 0)
            e.transform(m)
            e.translate(cx, cy, 0)
            return CommandResult(ok=True, payload={"scaled": entity_id})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_mirror(self, entity_id, x1, y1, x2, y2) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            copy = e.copy()
            self._msp.add_entity(copy)
            # Mirror across line (x1,y1)-(x2,y2) using reflection matrix
            dx, dy = x2 - x1, y2 - y1
            length_sq = dx * dx + dy * dy
            if length_sq == 0:
                return CommandResult(ok=False, error="Mirror line has zero length")
            from ezdxf.math import Matrix44
            # Reflect: translate to origin, reflect, translate back
            # Reflection matrix across line through origin with direction (dx, dy):
            #   [[cos2a, sin2a], [sin2a, -cos2a]] where a = atan2(dy, dx)
            a = math.atan2(dy, dx)
            cos2a = math.cos(2 * a)
            sin2a = math.sin(2 * a)
            m = Matrix44([
                cos2a, sin2a, 0, 0,
                sin2a, -cos2a, 0, 0,
                0, 0, 1, 0,
                0, 0, 0, 1,
            ])
            copy.translate(-x1, -y1, 0)
            copy.transform(m)
            copy.translate(x1, y1, 0)
            return CommandResult(ok=True, payload={"handle": copy.dxf.handle})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_offset(self, entity_id, distance) -> CommandResult:
        # ezdxf doesn't have a native offset command; approximate for simple cases
        return CommandResult(ok=False, error="Offset not supported on ezdxf backend")

    async def entity_array(self, entity_id, rows, cols, row_dist, col_dist) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            handles = []
            for r in range(rows):
                for c in range(cols):
                    if r == 0 and c == 0:
                        continue  # Skip original position
                    copy = e.copy()
                    self._msp.add_entity(copy)
                    copy.translate(c * col_dist, r * row_dist, 0)
                    handles.append(copy.dxf.handle)
            return CommandResult(ok=True, payload={"copies": len(handles), "handles": handles})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_fillet(self, entity_id1, entity_id2, radius) -> CommandResult:
        return CommandResult(ok=False, error="Fillet not supported on ezdxf backend")

    async def entity_chamfer(self, entity_id1, entity_id2, dist1, dist2) -> CommandResult:
        return CommandResult(ok=False, error="Chamfer not supported on ezdxf backend")

    async def create_hatch(self, entity_id, pattern="ANSI31") -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            hatch = self._msp.add_hatch()
            hatch.set_pattern_fill(pattern, scale=1.0)
            # Try to use the entity as a boundary path
            hatch.paths.add_polyline_path(
                [(p[0], p[1]) for p in e.get_points(format="xy")],
                is_closed=True,
            )
            return CommandResult(ok=True, payload={"entity_type": "HATCH", "handle": hatch.dxf.handle})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Layer operations ---

    async def layer_list(self) -> CommandResult:
        layers = []
        for l in self._doc.layers:
            layers.append({
                "name": l.dxf.name,
                "color": l.dxf.get("color", 7),
                "linetype": l.dxf.get("linetype", "Continuous"),
                "is_frozen": l.is_frozen(),
                "is_locked": l.is_locked(),
            })
        return CommandResult(ok=True, payload={"layers": layers})

    async def layer_create(self, name, color="white", linetype="CONTINUOUS") -> CommandResult:
        if name in self._doc.layers:
            return CommandResult(ok=True, payload={"name": name, "existed": True})
        color_int = self._color_to_int(color)
        self._doc.layers.add(name, color=color_int, linetype=linetype)
        return CommandResult(ok=True, payload={"name": name, "color": color_int})

    async def layer_set_current(self, name) -> CommandResult:
        if name not in self._doc.layers:
            return CommandResult(ok=False, error=f"Layer '{name}' does not exist")
        self._doc.header["$CLAYER"] = name
        return CommandResult(ok=True, payload={"current_layer": name})

    async def layer_set_properties(self, name, color=None, linetype=None, lineweight=None) -> CommandResult:
        if name not in self._doc.layers:
            return CommandResult(ok=False, error=f"Layer '{name}' does not exist")
        layer = self._doc.layers.get(name)
        if color is not None:
            layer.color = self._color_to_int(color)
        if linetype is not None:
            layer.dxf.linetype = linetype
        return CommandResult(ok=True, payload={"name": name})

    async def layer_freeze(self, name) -> CommandResult:
        if name not in self._doc.layers:
            return CommandResult(ok=False, error=f"Layer '{name}' does not exist")
        self._doc.layers.get(name).freeze()
        return CommandResult(ok=True, payload={"name": name, "frozen": True})

    async def layer_thaw(self, name) -> CommandResult:
        if name not in self._doc.layers:
            return CommandResult(ok=False, error=f"Layer '{name}' does not exist")
        self._doc.layers.get(name).thaw()
        return CommandResult(ok=True, payload={"name": name, "frozen": False})

    async def layer_lock(self, name) -> CommandResult:
        if name not in self._doc.layers:
            return CommandResult(ok=False, error=f"Layer '{name}' does not exist")
        self._doc.layers.get(name).lock()
        return CommandResult(ok=True, payload={"name": name, "locked": True})

    async def layer_unlock(self, name) -> CommandResult:
        if name not in self._doc.layers:
            return CommandResult(ok=False, error=f"Layer '{name}' does not exist")
        self._doc.layers.get(name).unlock()
        return CommandResult(ok=True, payload={"name": name, "locked": False})

    # --- Block operations ---

    async def block_list(self) -> CommandResult:
        blocks = [b.name for b in self._doc.blocks if not b.name.startswith("*")]
        return CommandResult(ok=True, payload={"blocks": blocks})

    async def block_insert(self, name, x, y, scale=1.0, rotation=0.0, block_id=None) -> CommandResult:
        if name not in self._doc.blocks:
            return CommandResult(ok=False, error=f"Block '{name}' not defined")
        e = self._msp.add_blockref(name, (x, y), dxfattribs={
            "xscale": scale, "yscale": scale, "zscale": scale,
            "rotation": rotation,
        })
        if block_id:
            try:
                e.add_attrib("ID", block_id)
            except Exception:
                pass
        return CommandResult(ok=True, payload={"entity_type": "INSERT", "handle": e.dxf.handle})

    async def block_insert_with_attributes(self, name, x, y, scale=1.0, rotation=0.0, attributes=None) -> CommandResult:
        if name not in self._doc.blocks:
            return CommandResult(ok=False, error=f"Block '{name}' not defined")
        block = self._doc.blocks[name]
        e = self._msp.add_blockref(name, (x, y), dxfattribs={
            "xscale": scale, "yscale": scale, "zscale": scale,
            "rotation": rotation,
        })
        if attributes:
            # Try add_auto_attribs first (uses ATTDEF templates)
            try:
                e.add_auto_attribs(attributes)
            except Exception:
                # Fallback: add manual attribs
                for tag, value in attributes.items():
                    try:
                        e.add_attrib(tag, value, (x, y))
                    except Exception:
                        pass
        return CommandResult(ok=True, payload={"entity_type": "INSERT", "handle": e.dxf.handle})

    async def block_get_attributes(self, entity_id) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None or e.dxftype() != "INSERT":
                return CommandResult(ok=False, error="Not an INSERT entity")
            attribs = {}
            for attrib in e.attribs:
                attribs[attrib.dxf.tag] = attrib.dxf.text
            return CommandResult(ok=True, payload={"attributes": attribs})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def block_update_attribute(self, entity_id, tag, value) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None or e.dxftype() != "INSERT":
                return CommandResult(ok=False, error="Not an INSERT entity")
            for attrib in e.attribs:
                if attrib.dxf.tag.upper() == tag.upper():
                    attrib.dxf.text = value
                    return CommandResult(ok=True, payload={"tag": tag, "value": value})
            return CommandResult(ok=False, error=f"Attribute '{tag}' not found")
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def block_define(self, name, entities) -> CommandResult:
        block = self._doc.blocks.new(name=name)
        for ent_def in entities:
            etype = ent_def.get("type", "LINE")
            if etype == "LINE":
                block.add_line(
                    (ent_def.get("x1", 0), ent_def.get("y1", 0)),
                    (ent_def.get("x2", 0), ent_def.get("y2", 0)),
                )
            elif etype == "CIRCLE":
                block.add_circle(
                    (ent_def.get("cx", 0), ent_def.get("cy", 0)),
                    ent_def.get("radius", 1),
                )
            elif etype == "ATTDEF":
                block.add_attdef(
                    ent_def.get("tag", "TAG"),
                    (ent_def.get("x", 0), ent_def.get("y", 0)),
                    dxfattribs={"height": ent_def.get("height", 2.5)},
                )
        return CommandResult(ok=True, payload={"block": name, "entity_count": len(entities)})

    # --- Annotation ---

    async def create_text(self, x, y, text, height=2.5, rotation=0.0, layer=None) -> CommandResult:
        self._ensure_layer(layer)
        e = self._msp.add_text(text, dxfattribs={
            "insert": (x, y),
            "height": height,
            "rotation": rotation,
            "layer": layer or "0",
        })
        return CommandResult(ok=True, payload={"entity_type": "TEXT", "handle": e.dxf.handle})

    async def create_dimension_linear(self, x1, y1, x2, y2, dim_x, dim_y) -> CommandResult:
        try:
            dim = self._msp.add_linear_dim(
                base=(dim_x, dim_y),
                p1=(x1, y1),
                p2=(x2, y2),
            )
            dim.render()
            return CommandResult(ok=True, payload={"entity_type": "DIMENSION"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def create_dimension_aligned(self, x1, y1, x2, y2, offset) -> CommandResult:
        try:
            dim = self._msp.add_aligned_dim(
                p1=(x1, y1),
                p2=(x2, y2),
                distance=offset,
            )
            dim.render()
            return CommandResult(ok=True, payload={"entity_type": "DIMENSION"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def create_dimension_angular(self, cx, cy, x1, y1, x2, y2) -> CommandResult:
        try:
            # Calculate angle arc midpoint for dimension location
            a1 = math.atan2(y1 - cy, x1 - cx)
            a2 = math.atan2(y2 - cy, x2 - cx)
            amid = (a1 + a2) / 2
            r = max(math.hypot(x1 - cx, y1 - cy), math.hypot(x2 - cx, y2 - cy)) * 0.7
            dim = self._msp.add_angular_dim_cra(
                center=(cx, cy),
                radius=r,
                start_angle=math.degrees(a1),
                end_angle=math.degrees(a2),
                distance=r * 1.2,
            )
            dim.render()
            return CommandResult(ok=True, payload={"entity_type": "DIMENSION"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def create_dimension_radius(self, cx, cy, radius, angle) -> CommandResult:
        try:
            rad = math.radians(angle)
            px = cx + radius * math.cos(rad)
            py = cy + radius * math.sin(rad)
            dim = self._msp.add_radius_dim(
                center=(cx, cy),
                mpoint=(px, py),
            )
            dim.render()
            return CommandResult(ok=True, payload={"entity_type": "DIMENSION"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def create_leader(self, points, text) -> CommandResult:
        try:
            pts = [(p[0], p[1]) for p in points]
            leader = self._msp.add_leader(pts)
            # Add text at the last point
            last = pts[-1]
            self._msp.add_mtext(text, dxfattribs={
                "insert": (last[0] + 2, last[1]),
                "char_height": 2.5,
                "width": 30,
            })
            return CommandResult(ok=True, payload={"entity_type": "LEADER"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- P&ID ---

    async def pid_setup_layers(self) -> CommandResult:
        pid_layers = [
            ("PID-EQUIPMENT", 6, "CONTINUOUS"),
            ("PID-PROCESS-PIPING", 4, "CONTINUOUS"),
            ("PID-UTILITY-PIPING", 3, "CONTINUOUS"),
            ("PID-INSTRUMENTS", 5, "CONTINUOUS"),
            ("PID-ELECTRICAL", 1, "CONTINUOUS"),
            ("PID-ANNOTATION", 7, "CONTINUOUS"),
            ("PID-VALVES", 2, "CONTINUOUS"),
        ]
        for name, color, lt in pid_layers:
            if name not in self._doc.layers:
                self._doc.layers.add(name, color=color, linetype=lt)
        return CommandResult(ok=True, payload={"layers_created": len(pid_layers)})

    async def pid_list_symbols(self, category) -> CommandResult:
        """List CTO symbols from disk or built-in catalog."""
        from autocad_mcp.pid.cto_library import CTO_ROOT, list_symbols
        symbols = list_symbols(category)
        return CommandResult(ok=True, payload={"category": category, "symbols": symbols, "count": len(symbols)})

    async def pid_insert_symbol(self, category, symbol, x, y, scale=1.0, rotation=0.0) -> CommandResult:
        """Insert a CTO symbol as a simple block placeholder."""
        self._ensure_layer("PID-EQUIPMENT")
        # In headless mode, create a placeholder rectangle with the symbol name
        half = 5 * scale
        pts = [(x - half, y - half), (x + half, y - half), (x + half, y + half), (x - half, y + half)]
        e = self._msp.add_lwpolyline(pts, close=True, dxfattribs={"layer": "PID-EQUIPMENT"})
        self._msp.add_text(symbol, dxfattribs={
            "insert": (x, y), "height": 1.5 * scale, "layer": "PID-ANNOTATION",
        })
        return CommandResult(ok=True, payload={"symbol": symbol, "handle": e.dxf.handle})

    async def pid_insert_valve(self, x, y, valve_type, rotation=0.0, attributes=None) -> CommandResult:
        """Insert a valve symbol (simplified for headless)."""
        self._ensure_layer("PID-VALVES")
        # Simplified diamond shape for valve
        size = 3.0
        pts = [(x - size, y), (x, y + size), (x + size, y), (x, y - size)]
        e = self._msp.add_lwpolyline(pts, close=True, dxfattribs={"layer": "PID-VALVES"})
        self._msp.add_text(valve_type, dxfattribs={
            "insert": (x, y - size - 2), "height": 1.5, "layer": "PID-ANNOTATION",
        })
        return CommandResult(ok=True, payload={"valve_type": valve_type, "handle": e.dxf.handle})

    async def pid_insert_instrument(self, x, y, instrument_type, rotation=0.0, tag_id="", range_value="") -> CommandResult:
        """Insert an instrument symbol (simplified for headless)."""
        self._ensure_layer("PID-INSTRUMENTS")
        # Circle with crosshair for instrument
        e = self._msp.add_circle((x, y), 4, dxfattribs={"layer": "PID-INSTRUMENTS"})
        self._msp.add_line((x - 4, y), (x + 4, y), dxfattribs={"layer": "PID-INSTRUMENTS"})
        label = tag_id if tag_id else instrument_type
        self._msp.add_text(label, dxfattribs={
            "insert": (x, y - 6), "height": 1.5, "layer": "PID-ANNOTATION",
        })
        return CommandResult(ok=True, payload={"instrument_type": instrument_type, "handle": e.dxf.handle})

    async def pid_insert_pump(self, x, y, pump_type, rotation=0.0, attributes=None) -> CommandResult:
        """Insert a pump symbol (simplified for headless)."""
        self._ensure_layer("PID-EQUIPMENT")
        # Circle with triangle for pump
        e = self._msp.add_circle((x, y), 6, dxfattribs={"layer": "PID-EQUIPMENT"})
        rad = math.radians(rotation)
        tip_x = x + 8 * math.cos(rad)
        tip_y = y + 8 * math.sin(rad)
        self._msp.add_lwpolyline(
            [(x + 6 * math.cos(rad + 0.5), y + 6 * math.sin(rad + 0.5)),
             (tip_x, tip_y),
             (x + 6 * math.cos(rad - 0.5), y + 6 * math.sin(rad - 0.5))],
            close=True,
            dxfattribs={"layer": "PID-EQUIPMENT"},
        )
        self._msp.add_text(pump_type, dxfattribs={
            "insert": (x, y - 8), "height": 1.5, "layer": "PID-ANNOTATION",
        })
        return CommandResult(ok=True, payload={"pump_type": pump_type, "handle": e.dxf.handle})

    async def pid_insert_tank(self, x, y, tank_type, scale=1.0, attributes=None) -> CommandResult:
        """Insert a tank symbol (simplified for headless)."""
        self._ensure_layer("PID-EQUIPMENT")
        w = 10 * scale
        h = 15 * scale
        pts = [(x - w, y), (x + w, y), (x + w, y + h), (x - w, y + h)]
        e = self._msp.add_lwpolyline(pts, close=True, dxfattribs={"layer": "PID-EQUIPMENT"})
        self._msp.add_text(tank_type, dxfattribs={
            "insert": (x, y + h + 2), "height": 2.0 * scale, "layer": "PID-ANNOTATION",
        })
        return CommandResult(ok=True, payload={"tank_type": tank_type, "handle": e.dxf.handle})

    async def pid_draw_process_line(self, x1, y1, x2, y2) -> CommandResult:
        self._ensure_layer("PID-PROCESS-PIPING")
        e = self._msp.add_line((x1, y1), (x2, y2), dxfattribs={"layer": "PID-PROCESS-PIPING"})
        return CommandResult(ok=True, payload={"entity_type": "LINE", "handle": e.dxf.handle})

    async def pid_connect_equipment(self, x1, y1, x2, y2) -> CommandResult:
        """Connect two points with orthogonal routing."""
        self._ensure_layer("PID-PROCESS-PIPING")
        mid_x = (x1 + x2) / 2
        pts = [(x1, y1), (mid_x, y1), (mid_x, y2), (x2, y2)]
        e = self._msp.add_lwpolyline(pts, dxfattribs={"layer": "PID-PROCESS-PIPING"})
        return CommandResult(ok=True, payload={"entity_type": "LWPOLYLINE", "handle": e.dxf.handle})

    async def pid_add_flow_arrow(self, x, y, rotation=0.0) -> CommandResult:
        self._ensure_layer("PID-ANNOTATION")
        # Simple triangle arrow
        rad = math.radians(rotation)
        size = 2.0
        p1 = (x + size * math.cos(rad), y + size * math.sin(rad))
        p2 = (x + size * 0.5 * math.cos(rad + 2.4), y + size * 0.5 * math.sin(rad + 2.4))
        p3 = (x + size * 0.5 * math.cos(rad - 2.4), y + size * 0.5 * math.sin(rad - 2.4))
        e = self._msp.add_lwpolyline([p1, p2, p3], close=True, dxfattribs={"layer": "PID-ANNOTATION"})
        return CommandResult(ok=True, payload={"entity_type": "LWPOLYLINE", "handle": e.dxf.handle})

    async def pid_add_equipment_tag(self, x, y, tag, description="") -> CommandResult:
        self._ensure_layer("PID-ANNOTATION")
        e = self._msp.add_text(tag, dxfattribs={
            "insert": (x, y), "height": 2.5, "layer": "PID-ANNOTATION",
        })
        result = {"entity_type": "TEXT", "handle": e.dxf.handle, "tag": tag}
        if description:
            e2 = self._msp.add_text(description, dxfattribs={
                "insert": (x, y - 3.5), "height": 1.8, "layer": "PID-ANNOTATION",
            })
            result["description_handle"] = e2.dxf.handle
        return CommandResult(ok=True, payload=result)

    async def pid_add_line_number(self, x, y, line_num, spec) -> CommandResult:
        self._ensure_layer("PID-ANNOTATION")
        text = f"{line_num}-{spec}"
        e = self._msp.add_text(text, dxfattribs={
            "insert": (x, y), "height": 2.0, "layer": "PID-ANNOTATION",
        })
        return CommandResult(ok=True, payload={"entity_type": "TEXT", "handle": e.dxf.handle})

    # --- Query operations ---

    async def query_entity_properties(self, entity_id) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            props = {
                "handle": e.dxf.handle,
                "type": e.dxftype(),
                "layer": e.dxf.get("layer", "0"),
            }
            # Common DXF properties — only include if explicitly set
            for attr in ("color", "linetype", "lineweight", "ltscale", "true_color",
                         "plotstyle_name", "shadow_mode", "transparency"):
                if e.dxf.hasattr(attr):
                    val = e.dxf.get(attr)
                    props[attr] = val if not hasattr(val, '__iter__') or isinstance(val, str) else list(val)

            # Type-specific geometry data
            dtype = e.dxftype()
            if dtype == "LINE":
                props["start"] = list(e.dxf.start)
                props["end"] = list(e.dxf.end)
            elif dtype == "CIRCLE":
                props["center"] = list(e.dxf.center)
                props["radius"] = e.dxf.radius
            elif dtype == "ARC":
                props["center"] = list(e.dxf.center)
                props["radius"] = e.dxf.radius
                props["start_angle"] = e.dxf.start_angle
                props["end_angle"] = e.dxf.end_angle
            elif dtype == "LWPOLYLINE":
                props["closed"] = e.closed
                props["vertex_count"] = len(e)
                if e.dxf.hasattr("const_width"):
                    props["const_width"] = e.dxf.const_width
                if e.dxf.hasattr("elevation"):
                    props["elevation"] = e.dxf.elevation
            elif dtype == "POLYLINE":
                props["closed"] = e.is_closed
                props["vertex_count"] = len(list(e.vertices))
            elif dtype == "TEXT":
                props["text"] = e.dxf.text
                props["insert"] = list(e.dxf.insert) if e.dxf.hasattr("insert") else None
                props["height"] = e.dxf.get("height", 2.5)
                props["rotation"] = e.dxf.get("rotation", 0.0)
                if e.dxf.hasattr("style"):
                    props["style"] = e.dxf.style
            elif dtype == "MTEXT":
                props["text"] = e.text
                if e.dxf.hasattr("insert"):
                    props["insert"] = list(e.dxf.insert)
                props["char_height"] = e.dxf.get("char_height", 2.5)
                props["width"] = e.dxf.get("width", 0)
                if e.dxf.hasattr("style"):
                    props["style"] = e.dxf.style
                if e.dxf.hasattr("attachment_point"):
                    props["attachment_point"] = e.dxf.attachment_point
            elif dtype == "INSERT":
                props["block_name"] = e.dxf.name
                props["insert"] = list(e.dxf.insert)
                props["xscale"] = e.dxf.get("xscale", 1.0)
                props["yscale"] = e.dxf.get("yscale", 1.0)
                props["zscale"] = e.dxf.get("zscale", 1.0)
                props["rotation"] = e.dxf.get("rotation", 0.0)
                attribs = {}
                for attrib in e.attribs:
                    attribs[attrib.dxf.tag] = attrib.dxf.text
                if attribs:
                    props["attributes"] = attribs
            elif dtype == "ELLIPSE":
                props["center"] = list(e.dxf.center)
                props["major_axis"] = list(e.dxf.major_axis)
                props["ratio"] = e.dxf.ratio
                props["start_param"] = e.dxf.get("start_param", 0.0)
                props["end_param"] = e.dxf.get("end_param", math.tau)
            elif dtype == "HATCH":
                if e.dxf.hasattr("pattern_name"):
                    props["pattern_name"] = e.dxf.pattern_name
                props["solid_fill"] = e.dxf.get("solid_fill", 0)
            elif dtype == "DIMENSION":
                for attr in ("dimstyle", "defpoint", "defpoint2", "defpoint3",
                             "text_midpoint", "dimtype"):
                    if e.dxf.hasattr(attr):
                        val = e.dxf.get(attr)
                        props[attr] = list(val) if hasattr(val, '__iter__') and not isinstance(val, str) else val
            elif dtype == "LEADER":
                try:
                    props["vertices"] = [list(v) for v in e.vertices]
                except Exception:
                    pass
                if e.dxf.hasattr("dimstyle"):
                    props["dimstyle"] = e.dxf.dimstyle

            return CommandResult(ok=True, payload=props)
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def query_entity_geometry(self, entity_id) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            dtype = e.dxftype()
            geom = {"type": dtype, "handle": e.dxf.handle}

            if dtype == "LINE":
                geom["start"] = list(e.dxf.start)
                geom["end"] = list(e.dxf.end)
            elif dtype == "CIRCLE":
                geom["center"] = list(e.dxf.center)
                geom["radius"] = e.dxf.radius
            elif dtype == "ARC":
                geom["center"] = list(e.dxf.center)
                geom["radius"] = e.dxf.radius
                geom["start_angle"] = e.dxf.start_angle
                geom["end_angle"] = e.dxf.end_angle
            elif dtype == "LWPOLYLINE":
                # get_points returns (x, y, start_width, end_width, bulge)
                vertices = []
                for pt in e.get_points(format="xyseb"):
                    vertices.append([pt[0], pt[1], pt[4]])  # x, y, bulge
                geom["vertices"] = vertices
                geom["closed"] = e.closed
            elif dtype == "POLYLINE":
                vertices = []
                for v in e.vertices:
                    vertices.append(list(v.dxf.location))
                geom["vertices"] = vertices
                geom["closed"] = e.is_closed
            elif dtype == "TEXT":
                geom["content"] = e.dxf.text
                geom["insert"] = list(e.dxf.insert) if e.dxf.hasattr("insert") else [0, 0, 0]
                geom["height"] = e.dxf.get("height", 2.5)
                geom["rotation"] = e.dxf.get("rotation", 0.0)
                geom["style"] = e.dxf.get("style", "Standard")
            elif dtype == "MTEXT":
                geom["content"] = e.text
                geom["insert"] = list(e.dxf.insert) if e.dxf.hasattr("insert") else [0, 0, 0]
                geom["char_height"] = e.dxf.get("char_height", 2.5)
                geom["width"] = e.dxf.get("width", 0)
                geom["style"] = e.dxf.get("style", "Standard")
                geom["attachment_point"] = e.dxf.get("attachment_point", 1)
            elif dtype == "INSERT":
                geom["block_name"] = e.dxf.name
                geom["insert"] = list(e.dxf.insert)
                geom["x_scale"] = e.dxf.get("xscale", 1.0)
                geom["y_scale"] = e.dxf.get("yscale", 1.0)
                geom["z_scale"] = e.dxf.get("zscale", 1.0)
                geom["rotation"] = e.dxf.get("rotation", 0.0)
                attribs = {}
                for attrib in e.attribs:
                    attribs[attrib.dxf.tag] = attrib.dxf.text
                geom["attributes"] = attribs
            elif dtype == "ELLIPSE":
                geom["center"] = list(e.dxf.center)
                geom["major_axis"] = list(e.dxf.major_axis)
                geom["ratio"] = e.dxf.ratio
                geom["start_param"] = e.dxf.get("start_param", 0.0)
                geom["end_param"] = e.dxf.get("end_param", math.tau)
            elif dtype == "LEADER":
                try:
                    geom["vertices"] = [list(v) for v in e.vertices]
                except Exception:
                    geom["vertices"] = []
                geom["dimstyle"] = e.dxf.get("dimstyle", "Standard")
            elif dtype == "DIMENSION":
                for attr in ("dimstyle", "defpoint", "defpoint2", "defpoint3",
                             "text_midpoint", "dimtype", "actual_measurement"):
                    if e.dxf.hasattr(attr):
                        val = e.dxf.get(attr)
                        geom[attr] = list(val) if hasattr(val, '__iter__') and not isinstance(val, str) else val
            else:
                geom["info"] = f"Geometry extraction not implemented for {dtype}"

            return CommandResult(ok=True, payload=geom)
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def query_drawing_summary(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            by_type = {}
            by_layer = {}
            for e in self._msp:
                dtype = e.dxftype()
                layer = e.dxf.get("layer", "0")
                by_type[dtype] = by_type.get(dtype, 0) + 1
                by_layer[layer] = by_layer.get(layer, 0) + 1

            # Drawing extents via bounding box
            extents = None
            try:
                from ezdxf import bbox
                box = bbox.extents(self._msp)
                if box.has_data:
                    extents = {
                        "min": list(box.extmin),
                        "max": list(box.extmax),
                    }
            except Exception:
                pass

            # Inventory
            blocks = [b.name for b in self._doc.blocks if not b.name.startswith("*")]
            styles = [s.dxf.name for s in self._doc.styles]
            linetypes = [lt.dxf.name for lt in self._doc.linetypes]
            layers = [l.dxf.name for l in self._doc.layers]

            return CommandResult(ok=True, payload={
                "total_entities": len(self._msp),
                "by_type": by_type,
                "by_layer": by_layer,
                "extents": extents,
                "blocks": blocks,
                "styles": styles,
                "linetypes": linetypes,
                "layers": layers,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def query_layer_summary(self, layer) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            if layer not in self._doc.layers:
                return CommandResult(ok=False, error=f"Layer '{layer}' does not exist")

            by_type = {}
            layer_entities = []
            for e in self._msp:
                if e.dxf.get("layer", "0") == layer:
                    dtype = e.dxftype()
                    by_type[dtype] = by_type.get(dtype, 0) + 1
                    layer_entities.append(e)

            bbox_data = None
            try:
                from ezdxf import bbox
                box = bbox.extents(layer_entities)
                if box.has_data:
                    bbox_data = {
                        "min": list(box.extmin),
                        "max": list(box.extmax),
                    }
            except Exception:
                pass

            total = sum(by_type.values())
            return CommandResult(ok=True, payload={
                "layer": layer,
                "total_entities": total,
                "by_type": by_type,
                "bounding_box": bbox_data,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Search operations ---

    async def search_text(self, pattern, case_sensitive=False) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            import re
            flags = 0 if case_sensitive else re.IGNORECASE
            regex = re.compile(pattern, flags)
            results = []
            for e in self._msp:
                dtype = e.dxftype()
                text_content = None
                if dtype == "TEXT":
                    text_content = e.dxf.get("text", "")
                elif dtype == "MTEXT":
                    text_content = e.text
                if text_content and regex.search(text_content):
                    entry = {
                        "handle": e.dxf.handle,
                        "type": dtype,
                        "text": text_content,
                        "layer": e.dxf.get("layer", "0"),
                    }
                    if dtype == "TEXT" and e.dxf.hasattr("insert"):
                        entry["position"] = list(e.dxf.insert)
                    elif dtype == "MTEXT" and e.dxf.hasattr("insert"):
                        entry["position"] = list(e.dxf.insert)
                    results.append(entry)
            return CommandResult(ok=True, payload={"matches": results, "count": len(results)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def search_by_attribute(self, tag=None, value=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            results = []
            for e in self._msp:
                if e.dxftype() != "INSERT":
                    continue
                attribs = {}
                matched = False
                for attrib in e.attribs:
                    a_tag = attrib.dxf.tag
                    a_val = attrib.dxf.text
                    attribs[a_tag] = a_val
                    if tag and value:
                        if a_tag.upper() == tag.upper() and value.upper() in a_val.upper():
                            matched = True
                    elif tag:
                        if a_tag.upper() == tag.upper():
                            matched = True
                    elif value:
                        if value.upper() in a_val.upper():
                            matched = True
                    else:
                        matched = True  # No filter, return all INSERTs with attributes

                if matched and attribs:
                    results.append({
                        "handle": e.dxf.handle,
                        "block_name": e.dxf.name,
                        "insert": list(e.dxf.insert),
                        "layer": e.dxf.get("layer", "0"),
                        "attributes": attribs,
                    })
            return CommandResult(ok=True, payload={"matches": results, "count": len(results)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def search_by_window(self, x1, y1, x2, y2) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            from ezdxf import bbox
            min_x, max_x = min(x1, x2), max(x1, x2)
            min_y, max_y = min(y1, y2), max(y1, y2)
            results = []
            for e in self._msp:
                try:
                    box = bbox.extents([e])
                    if not box.has_data:
                        continue
                    emin = box.extmin
                    emax = box.extmax
                    # Entity fully inside window
                    if emin[0] >= min_x and emin[1] >= min_y and emax[0] <= max_x and emax[1] <= max_y:
                        entry = {
                            "handle": e.dxf.handle,
                            "type": e.dxftype(),
                            "layer": e.dxf.get("layer", "0"),
                            "bbox_min": [emin[0], emin[1]],
                            "bbox_max": [emax[0], emax[1]],
                        }
                        results.append(entry)
                except Exception:
                    continue
            return CommandResult(ok=True, payload={"matches": results, "count": len(results)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def search_by_proximity(self, x, y, radius) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            from ezdxf import bbox
            results = []
            for e in self._msp:
                try:
                    box = bbox.extents([e])
                    if not box.has_data:
                        continue
                    # Use center of bounding box as entity reference point
                    emin = box.extmin
                    emax = box.extmax
                    cx = (emin[0] + emax[0]) / 2
                    cy = (emin[1] + emax[1]) / 2
                    dist = math.sqrt((cx - x) ** 2 + (cy - y) ** 2)
                    if dist <= radius:
                        results.append({
                            "handle": e.dxf.handle,
                            "type": e.dxftype(),
                            "layer": e.dxf.get("layer", "0"),
                            "distance": round(dist, 6),
                            "center": [round(cx, 6), round(cy, 6)],
                        })
                except Exception:
                    continue
            # Sort by distance
            results.sort(key=lambda r: r["distance"])
            return CommandResult(ok=True, payload={"matches": results, "count": len(results)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def search_by_type_and_layer(self, entity_type=None, layer=None, color=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            results = []
            for e in self._msp:
                dtype = e.dxftype()
                e_layer = e.dxf.get("layer", "0")
                e_color = e.dxf.get("color", None)

                if entity_type and dtype.upper() != entity_type.upper():
                    continue
                if layer and e_layer != layer:
                    continue
                if color is not None and e_color != color:
                    continue

                entry = {
                    "handle": e.dxf.handle,
                    "type": dtype,
                    "layer": e_layer,
                }
                if e_color is not None:
                    entry["color"] = e_color
                if e.dxf.hasattr("linetype"):
                    entry["linetype"] = e.dxf.linetype
                if e.dxf.hasattr("lineweight"):
                    entry["lineweight"] = e.dxf.lineweight
                results.append(entry)
            return CommandResult(ok=True, payload={"matches": results, "count": len(results)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Geometry operations ---

    async def geometry_distance(self, x1, y1, x2, y2) -> CommandResult:
        try:
            dx = x2 - x1
            dy = y2 - y1
            dist = math.sqrt(dx * dx + dy * dy)
            angle = math.degrees(math.atan2(dy, dx))
            return CommandResult(ok=True, payload={
                "distance": round(dist, 10),
                "dx": round(dx, 10),
                "dy": round(dy, 10),
                "angle": round(angle, 6),
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def geometry_length(self, entity_id) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            dtype = e.dxftype()
            length = 0.0
            if dtype == "LINE":
                s = e.dxf.start
                end = e.dxf.end
                length = math.sqrt((end[0] - s[0]) ** 2 + (end[1] - s[1]) ** 2 + (end[2] - s[2]) ** 2)
            elif dtype == "CIRCLE":
                length = 2 * math.pi * e.dxf.radius
            elif dtype == "ARC":
                sa = math.radians(e.dxf.start_angle)
                ea = math.radians(e.dxf.end_angle)
                sweep = ea - sa
                if sweep < 0:
                    sweep += 2 * math.pi
                length = abs(sweep) * e.dxf.radius
            elif dtype == "LWPOLYLINE":
                pts = list(e.get_points(format="xyseb"))
                closed = e.closed
                n = len(pts)
                for i in range(n - 1):
                    x1p, y1p = pts[i][0], pts[i][1]
                    x2p, y2p = pts[i + 1][0], pts[i + 1][1]
                    bulge = pts[i][4]
                    length += self._segment_length(x1p, y1p, x2p, y2p, bulge)
                if closed and n > 1:
                    x1p, y1p = pts[-1][0], pts[-1][1]
                    x2p, y2p = pts[0][0], pts[0][1]
                    bulge = pts[-1][4]
                    length += self._segment_length(x1p, y1p, x2p, y2p, bulge)
            elif dtype == "POLYLINE":
                verts = [v.dxf.location for v in e.vertices]
                for i in range(len(verts) - 1):
                    dx = verts[i + 1][0] - verts[i][0]
                    dy = verts[i + 1][1] - verts[i][1]
                    dz = verts[i + 1][2] - verts[i][2]
                    length += math.sqrt(dx * dx + dy * dy + dz * dz)
                if e.is_closed and len(verts) > 1:
                    dx = verts[0][0] - verts[-1][0]
                    dy = verts[0][1] - verts[-1][1]
                    dz = verts[0][2] - verts[-1][2]
                    length += math.sqrt(dx * dx + dy * dy + dz * dz)
            elif dtype == "ELLIPSE":
                # Approximate ellipse perimeter using Ramanujan's formula
                a_len = math.sqrt(e.dxf.major_axis[0] ** 2 + e.dxf.major_axis[1] ** 2 + e.dxf.major_axis[2] ** 2)
                b_len = a_len * e.dxf.ratio
                h = ((a_len - b_len) / (a_len + b_len)) ** 2
                length = math.pi * (a_len + b_len) * (1 + 3 * h / (10 + math.sqrt(4 - 3 * h)))
            else:
                return CommandResult(ok=False, error=f"Length not supported for {dtype}")

            return CommandResult(ok=True, payload={"entity_type": dtype, "length": round(length, 10)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def geometry_area(self, entity_id) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            dtype = e.dxftype()
            area = 0.0
            if dtype == "CIRCLE":
                area = math.pi * e.dxf.radius ** 2
            elif dtype == "LWPOLYLINE":
                if not e.closed:
                    return CommandResult(ok=False, error="Polyline is not closed; area undefined")
                # Shoelace formula (approximation for straight segments)
                pts = list(e.get_points(format="xy"))
                n = len(pts)
                for i in range(n):
                    j = (i + 1) % n
                    area += pts[i][0] * pts[j][1]
                    area -= pts[j][0] * pts[i][1]
                area = abs(area) / 2.0
            elif dtype == "POLYLINE":
                if not e.is_closed:
                    return CommandResult(ok=False, error="Polyline is not closed; area undefined")
                verts = [(v.dxf.location[0], v.dxf.location[1]) for v in e.vertices]
                n = len(verts)
                for i in range(n):
                    j = (i + 1) % n
                    area += verts[i][0] * verts[j][1]
                    area -= verts[j][0] * verts[i][1]
                area = abs(area) / 2.0
            elif dtype == "ELLIPSE":
                a_len = math.sqrt(e.dxf.major_axis[0] ** 2 + e.dxf.major_axis[1] ** 2 + e.dxf.major_axis[2] ** 2)
                b_len = a_len * e.dxf.ratio
                area = math.pi * a_len * b_len
            else:
                return CommandResult(ok=False, error=f"Area not supported for {dtype}")

            return CommandResult(ok=True, payload={"entity_type": dtype, "area": round(area, 10)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def geometry_bounding_box(self, entity_id=None, layer=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            from ezdxf import bbox
            if entity_id:
                e = self._doc.entitydb.get(entity_id)
                if e is None:
                    return CommandResult(ok=False, error=f"Entity {entity_id} not found")
                box = bbox.extents([e])
            elif layer:
                entities = [e for e in self._msp if e.dxf.get("layer", "0") == layer]
                if not entities:
                    return CommandResult(ok=False, error=f"No entities on layer '{layer}'")
                box = bbox.extents(entities)
            else:
                box = bbox.extents(self._msp)

            if not box.has_data:
                return CommandResult(ok=False, error="No bounding box data available")

            return CommandResult(ok=True, payload={
                "min": list(box.extmin),
                "max": list(box.extmax),
                "width": round(box.extmax[0] - box.extmin[0], 10),
                "height": round(box.extmax[1] - box.extmin[1], 10),
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def geometry_polyline_info(self, entity_id) -> CommandResult:
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            dtype = e.dxftype()
            if dtype == "LWPOLYLINE":
                pts = list(e.get_points(format="xyseb"))
                closed = e.closed
                vertices = []
                for pt in pts:
                    vertices.append({"x": pt[0], "y": pt[1], "bulge": pt[4]})

                # Segment lengths
                n = len(pts)
                segment_lengths = []
                total_length = 0.0
                count = n if closed else n - 1
                for i in range(count):
                    j = (i + 1) % n
                    x1p, y1p = pts[i][0], pts[i][1]
                    x2p, y2p = pts[j][0], pts[j][1]
                    bulge = pts[i][4]
                    seg_len = self._segment_length(x1p, y1p, x2p, y2p, bulge)
                    segment_lengths.append(round(seg_len, 10))
                    total_length += seg_len

                result = {
                    "entity_type": "LWPOLYLINE",
                    "vertices": vertices,
                    "vertex_count": n,
                    "closed": closed,
                    "segment_lengths": segment_lengths,
                    "total_length": round(total_length, 10),
                }

                if closed and n >= 3:
                    # Compute area using shoelace formula
                    xy = [(pt[0], pt[1]) for pt in pts]
                    area = 0.0
                    for i in range(n):
                        j = (i + 1) % n
                        area += xy[i][0] * xy[j][1]
                        area -= xy[j][0] * xy[i][1]
                    result["area"] = round(abs(area) / 2.0, 10)

                return CommandResult(ok=True, payload=result)

            elif dtype == "POLYLINE":
                verts = list(e.vertices)
                closed = e.is_closed
                vertices = [{"x": v.dxf.location[0], "y": v.dxf.location[1], "z": v.dxf.location[2]} for v in verts]

                segment_lengths = []
                total_length = 0.0
                n = len(verts)
                count = n if closed else n - 1
                for i in range(count):
                    j = (i + 1) % n
                    loc_i = verts[i].dxf.location
                    loc_j = verts[j].dxf.location
                    dx = loc_j[0] - loc_i[0]
                    dy = loc_j[1] - loc_i[1]
                    dz = loc_j[2] - loc_i[2]
                    seg_len = math.sqrt(dx * dx + dy * dy + dz * dz)
                    segment_lengths.append(round(seg_len, 10))
                    total_length += seg_len

                result = {
                    "entity_type": "POLYLINE",
                    "vertices": vertices,
                    "vertex_count": n,
                    "closed": closed,
                    "segment_lengths": segment_lengths,
                    "total_length": round(total_length, 10),
                }

                if closed and n >= 3:
                    coords = [(v.dxf.location[0], v.dxf.location[1]) for v in verts]
                    area = 0.0
                    for i in range(n):
                        j = (i + 1) % n
                        area += coords[i][0] * coords[j][1]
                        area -= coords[j][0] * coords[i][1]
                    result["area"] = round(abs(area) / 2.0, 10)

                return CommandResult(ok=True, payload=result)
            else:
                return CommandResult(ok=False, error=f"Entity is {dtype}, not a polyline")
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Bulk operations ---

    async def bulk_set_property(self, handles, property_name, value) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            updated = 0
            errors = []
            for h in handles:
                e = self._doc.entitydb.get(h)
                if e is None:
                    errors.append(f"{h}: not found")
                    continue
                if property_name == "layer":
                    self._ensure_layer(value)
                    e.dxf.layer = value
                elif property_name == "color":
                    e.dxf.color = int(value) if not isinstance(value, int) else value
                elif property_name == "linetype":
                    e.dxf.linetype = value
                elif property_name == "lineweight":
                    e.dxf.lineweight = int(value) if not isinstance(value, int) else value
                else:
                    errors.append(f"{h}: unsupported property '{property_name}'")
                    continue
                updated += 1
            return CommandResult(ok=True, payload={
                "updated": updated,
                "errors": errors if errors else None,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def bulk_erase(self, handles) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            erased = 0
            errors = []
            for h in handles:
                e = self._doc.entitydb.get(h)
                if e is None:
                    errors.append(f"{h}: not found")
                    continue
                try:
                    self._msp.delete_entity(e)
                    erased += 1
                except Exception as inner_ex:
                    errors.append(f"{h}: {str(inner_ex)}")
            return CommandResult(ok=True, payload={
                "erased": erased,
                "errors": errors if errors else None,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Export ---

    async def export_entity_data(self, layer=None, entity_type=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            entities = []
            for e in self._msp:
                dtype = e.dxftype()
                e_layer = e.dxf.get("layer", "0")
                if layer and e_layer != layer:
                    continue
                if entity_type and dtype.upper() != entity_type.upper():
                    continue

                entry = {
                    "handle": e.dxf.handle,
                    "type": dtype,
                    "layer": e_layer,
                }
                # Common properties
                for attr in ("color", "linetype", "lineweight", "ltscale"):
                    if e.dxf.hasattr(attr):
                        entry[attr] = e.dxf.get(attr)

                # Type-specific data
                if dtype == "LINE":
                    entry["start"] = list(e.dxf.start)
                    entry["end"] = list(e.dxf.end)
                elif dtype == "CIRCLE":
                    entry["center"] = list(e.dxf.center)
                    entry["radius"] = e.dxf.radius
                elif dtype == "ARC":
                    entry["center"] = list(e.dxf.center)
                    entry["radius"] = e.dxf.radius
                    entry["start_angle"] = e.dxf.start_angle
                    entry["end_angle"] = e.dxf.end_angle
                elif dtype == "LWPOLYLINE":
                    entry["vertices"] = [[pt[0], pt[1], pt[4]] for pt in e.get_points(format="xyseb")]
                    entry["closed"] = e.closed
                elif dtype == "TEXT":
                    entry["text"] = e.dxf.get("text", "")
                    if e.dxf.hasattr("insert"):
                        entry["insert"] = list(e.dxf.insert)
                    entry["height"] = e.dxf.get("height", 2.5)
                    entry["rotation"] = e.dxf.get("rotation", 0.0)
                elif dtype == "MTEXT":
                    entry["text"] = e.text
                    if e.dxf.hasattr("insert"):
                        entry["insert"] = list(e.dxf.insert)
                    entry["char_height"] = e.dxf.get("char_height", 2.5)
                    entry["width"] = e.dxf.get("width", 0)
                elif dtype == "INSERT":
                    entry["block_name"] = e.dxf.name
                    entry["insert"] = list(e.dxf.insert)
                    entry["xscale"] = e.dxf.get("xscale", 1.0)
                    entry["yscale"] = e.dxf.get("yscale", 1.0)
                    entry["rotation"] = e.dxf.get("rotation", 0.0)
                    attribs = {}
                    for attrib in e.attribs:
                        attribs[attrib.dxf.tag] = attrib.dxf.text
                    if attribs:
                        entry["attributes"] = attribs

                entities.append(entry)

            return CommandResult(ok=True, payload={"entities": entities, "count": len(entities)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Selection ---

    async def select_filter(self, entity_type=None, layer=None, color=None,
                            x1=None, y1=None, x2=None, y2=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            use_window = all(v is not None for v in (x1, y1, x2, y2))
            if use_window:
                from ezdxf import bbox as bbox_mod
                min_x, max_x = min(x1, x2), max(x1, x2)
                min_y, max_y = min(y1, y2), max(y1, y2)

            handles = []
            for e in self._msp:
                dtype = e.dxftype()
                e_layer = e.dxf.get("layer", "0")
                e_color = e.dxf.get("color", None)

                if entity_type and dtype.upper() != entity_type.upper():
                    continue
                if layer and e_layer != layer:
                    continue
                if color is not None and e_color != color:
                    continue

                if use_window:
                    try:
                        box = bbox_mod.extents([e])
                        if not box.has_data:
                            continue
                        emin = box.extmin
                        emax = box.extmax
                        if not (emin[0] >= min_x and emin[1] >= min_y
                                and emax[0] <= max_x and emax[1] <= max_y):
                            continue
                    except Exception:
                        continue

                handles.append(e.dxf.handle)

            return CommandResult(ok=True, payload={"handles": handles, "count": len(handles)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Bulk move / copy ---

    async def bulk_move(self, handles, dx, dy) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            moved = 0
            errors = []
            for h in handles:
                e = self._doc.entitydb.get(h)
                if e is None:
                    errors.append(f"{h}: not found")
                    continue
                try:
                    e.translate(dx, dy, 0)
                    moved += 1
                except Exception as inner_ex:
                    errors.append(f"{h}: {str(inner_ex)}")
            return CommandResult(ok=True, payload={
                "moved": moved,
                "errors": errors if errors else None,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def bulk_copy(self, handles, dx, dy) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            copied = 0
            new_handles = []
            errors = []
            for h in handles:
                e = self._doc.entitydb.get(h)
                if e is None:
                    errors.append(f"{h}: not found")
                    continue
                try:
                    copy = e.copy()
                    self._msp.add_entity(copy)
                    copy.translate(dx, dy, 0)
                    new_handles.append(copy.dxf.handle)
                    copied += 1
                except Exception as inner_ex:
                    errors.append(f"{h}: {str(inner_ex)}")
            return CommandResult(ok=True, payload={
                "copied": copied,
                "new_handles": new_handles,
                "errors": errors if errors else None,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Find / replace text ---

    async def find_replace_text(self, find, replace, layer=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            replaced = 0
            entities = []
            for e in self._msp:
                dtype = e.dxftype()
                if dtype not in ("TEXT", "MTEXT"):
                    continue
                e_layer = e.dxf.get("layer", "0")
                if layer and e_layer != layer:
                    continue

                if dtype == "TEXT":
                    old_text = e.dxf.get("text", "")
                    if find in old_text:
                        new_text = old_text.replace(find, replace)
                        e.dxf.text = new_text
                        entities.append({
                            "handle": e.dxf.handle,
                            "old_text": old_text,
                            "new_text": new_text,
                        })
                        replaced += 1
                elif dtype == "MTEXT":
                    old_text = e.text
                    if find in old_text:
                        new_text = old_text.replace(find, replace)
                        e.text = new_text
                        entities.append({
                            "handle": e.dxf.handle,
                            "old_text": old_text,
                            "new_text": new_text,
                        })
                        replaced += 1

            return CommandResult(ok=True, payload={"replaced": replaced, "entities": entities})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Entity property / text setters ---

    async def entity_set_property(self, entity_id, property_name, value) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")

            if property_name == "layer":
                self._ensure_layer(value)
                e.dxf.layer = value
            elif property_name == "color":
                e.dxf.color = int(value) if not isinstance(value, int) else value
            elif property_name == "linetype":
                e.dxf.linetype = value
            elif property_name == "lineweight":
                e.dxf.lineweight = int(value) if not isinstance(value, int) else value
            else:
                return CommandResult(ok=False, error=f"Unsupported property '{property_name}'")

            return CommandResult(ok=True, payload={
                "handle": e.dxf.handle,
                "property": property_name,
                "value": value,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_set_text(self, entity_id, text) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")

            dtype = e.dxftype()
            if dtype == "TEXT":
                e.dxf.text = text
            elif dtype == "MTEXT":
                e.text = text
            else:
                return CommandResult(ok=False, error=f"Entity {entity_id} is {dtype}, not TEXT or MTEXT")

            return CommandResult(ok=True, payload={"handle": e.dxf.handle, "new_text": text})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Zoom ---

    async def zoom_center(self, x, y, height) -> CommandResult:
        # The ezdxf backend is headless and has no real viewport zoom.
        # Store the requested view parameters and return success.
        return CommandResult(ok=True, payload={
            "center": [x, y],
            "height": height,
            "note": "Headless backend — zoom stored but has no visual effect",
        })

    # --- Layer visibility ---

    async def layer_visibility(self, name, visible) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            if name not in self._doc.layers:
                return CommandResult(ok=False, error=f"Layer '{name}' does not exist")
            layer = self._doc.layers.get(name)
            layer.is_off = not visible
            return CommandResult(ok=True, payload={"layer": name, "visible": visible})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Validation ---

    async def validate_layer_standards(self, allowed_layers) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            allowed_set = set(allowed_layers)
            violations = []
            for e in self._msp:
                e_layer = e.dxf.get("layer", "0")
                if e_layer not in allowed_set:
                    violations.append({
                        "handle": e.dxf.handle,
                        "layer": e_layer,
                        "type": e.dxftype(),
                    })
            passed = len(violations) == 0
            return CommandResult(ok=True, payload={
                "pass": passed,
                "violations": violations,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def validate_duplicates(self, tolerance=0.001) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            duplicates = []
            # Group entities by type for efficient comparison
            by_type: dict[str, list] = {}
            for e in self._msp:
                dtype = e.dxftype()
                by_type.setdefault(dtype, []).append(e)

            # Check LINE duplicates
            lines = by_type.get("LINE", [])
            for i in range(len(lines)):
                for j in range(i + 1, len(lines)):
                    a, b = lines[i], lines[j]
                    s1, e1 = a.dxf.start, a.dxf.end
                    s2, e2 = b.dxf.start, b.dxf.end
                    # Check both orientations
                    fwd = (abs(s1[0] - s2[0]) < tolerance and abs(s1[1] - s2[1]) < tolerance
                           and abs(e1[0] - e2[0]) < tolerance and abs(e1[1] - e2[1]) < tolerance)
                    rev = (abs(s1[0] - e2[0]) < tolerance and abs(s1[1] - e2[1]) < tolerance
                           and abs(e1[0] - s2[0]) < tolerance and abs(e1[1] - s2[1]) < tolerance)
                    if fwd or rev:
                        duplicates.append({
                            "type": "LINE",
                            "handle_a": a.dxf.handle,
                            "handle_b": b.dxf.handle,
                        })

            # Check CIRCLE duplicates
            circles = by_type.get("CIRCLE", [])
            for i in range(len(circles)):
                for j in range(i + 1, len(circles)):
                    a, b = circles[i], circles[j]
                    c1, c2 = a.dxf.center, b.dxf.center
                    if (abs(c1[0] - c2[0]) < tolerance
                            and abs(c1[1] - c2[1]) < tolerance
                            and abs(a.dxf.radius - b.dxf.radius) < tolerance):
                        duplicates.append({
                            "type": "CIRCLE",
                            "handle_a": a.dxf.handle,
                            "handle_b": b.dxf.handle,
                        })

            return CommandResult(ok=True, payload={"duplicates": duplicates})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def validate_zero_length(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            tolerance = 1e-9
            issues = []
            for e in self._msp:
                dtype = e.dxftype()
                if dtype == "LINE":
                    s = e.dxf.start
                    end = e.dxf.end
                    length = math.sqrt((end[0] - s[0]) ** 2 + (end[1] - s[1]) ** 2)
                    if length < tolerance:
                        issues.append({
                            "handle": e.dxf.handle,
                            "type": "LINE",
                            "issue": "zero_length",
                            "layer": e.dxf.get("layer", "0"),
                        })
                elif dtype == "CIRCLE":
                    if e.dxf.radius < tolerance:
                        issues.append({
                            "handle": e.dxf.handle,
                            "type": "CIRCLE",
                            "issue": "zero_radius",
                            "layer": e.dxf.get("layer", "0"),
                        })
            return CommandResult(ok=True, payload={"issues": issues})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def validate_qc_report(self, allowed_layers=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            checks = {}

            # Layer standards check
            if allowed_layers:
                layer_result = await self.validate_layer_standards(allowed_layers)
                checks["layer_standards"] = layer_result.payload if layer_result.ok else {"error": layer_result.error}
            else:
                checks["layer_standards"] = {"pass": True, "violations": [], "note": "No allowed_layers specified, skipped"}

            # Duplicates check
            dup_result = await self.validate_duplicates()
            checks["duplicates"] = dup_result.payload if dup_result.ok else {"error": dup_result.error}

            # Zero-length check
            zl_result = await self.validate_zero_length()
            checks["zero_length"] = zl_result.payload if zl_result.ok else {"error": zl_result.error}

            total_issues = 0
            if "violations" in checks["layer_standards"]:
                total_issues += len(checks["layer_standards"]["violations"])
            if "duplicates" in checks["duplicates"]:
                total_issues += len(checks["duplicates"]["duplicates"])
            if "issues" in checks["zero_length"]:
                total_issues += len(checks["zero_length"]["issues"])

            passed = total_issues == 0
            return CommandResult(ok=True, payload={
                "checks": checks,
                "total_issues": total_issues,
                "pass": passed,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Export: BOM / data extraction / reports ---

    async def export_bom(self, block_names=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            block_data: dict[str, dict] = {}
            for e in self._msp:
                if e.dxftype() != "INSERT":
                    continue
                bname = e.dxf.name
                if block_names and bname not in block_names:
                    continue

                if bname not in block_data:
                    block_data[bname] = {"count": 0, "attributes": []}
                block_data[bname]["count"] += 1

                attribs = {}
                for attrib in e.attribs:
                    attribs[attrib.dxf.tag] = attrib.dxf.text
                if attribs:
                    block_data[bname]["attributes"].append(attribs)

            items = []
            for bname, info in block_data.items():
                items.append({
                    "block": bname,
                    "count": info["count"],
                    "attributes": info["attributes"],
                })

            return CommandResult(ok=True, payload={"items": items})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def export_data_extract(self, entity_type=None, layer=None,
                                  properties=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            default_props = ["handle", "type", "layer", "color", "linetype", "lineweight"]
            columns = properties if properties else default_props
            rows = []
            for e in self._msp:
                dtype = e.dxftype()
                e_layer = e.dxf.get("layer", "0")
                if entity_type and dtype.upper() != entity_type.upper():
                    continue
                if layer and e_layer != layer:
                    continue

                row = []
                for prop in columns:
                    if prop == "handle":
                        row.append(e.dxf.handle)
                    elif prop == "type":
                        row.append(dtype)
                    elif prop == "layer":
                        row.append(e_layer)
                    elif prop == "color":
                        row.append(e.dxf.get("color", None))
                    elif prop == "linetype":
                        row.append(e.dxf.get("linetype", None))
                    elif prop == "lineweight":
                        row.append(e.dxf.get("lineweight", None))
                    elif prop == "text":
                        if dtype == "TEXT":
                            row.append(e.dxf.get("text", ""))
                        elif dtype == "MTEXT":
                            row.append(e.text)
                        else:
                            row.append(None)
                    elif prop == "block_name":
                        row.append(e.dxf.get("name", None) if dtype == "INSERT" else None)
                    else:
                        # Try generic dxf attribute
                        row.append(e.dxf.get(prop, None))
                rows.append(row)

            return CommandResult(ok=True, payload={
                "columns": columns,
                "rows": rows,
                "count": len(rows),
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def export_layer_report(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            layer_info: dict[str, dict[str, int]] = {}
            for e in self._msp:
                e_layer = e.dxf.get("layer", "0")
                dtype = e.dxftype()
                if e_layer not in layer_info:
                    layer_info[e_layer] = {}
                layer_info[e_layer][dtype] = layer_info[e_layer].get(dtype, 0) + 1

            layers = []
            for lname, types in layer_info.items():
                entity_count = sum(types.values())
                layers.append({
                    "name": lname,
                    "entity_count": entity_count,
                    "types": types,
                })
            # Sort by name
            layers.sort(key=lambda l: l["name"])

            return CommandResult(ok=True, payload={"layers": layers})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def export_block_count(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            counts: dict[str, int] = {}
            for e in self._msp:
                if e.dxftype() != "INSERT":
                    continue
                bname = e.dxf.name
                counts[bname] = counts.get(bname, 0) + 1

            blocks = [{"name": n, "count": c} for n, c in counts.items()]
            blocks.sort(key=lambda b: b["count"], reverse=True)

            return CommandResult(ok=True, payload={"blocks": blocks})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def export_drawing_statistics(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            by_type: dict[str, int] = {}
            entity_count = 0
            for e in self._msp:
                dtype = e.dxftype()
                by_type[dtype] = by_type.get(dtype, 0) + 1
                entity_count += 1

            layer_count = len(list(self._doc.layers))
            block_count = len([b for b in self._doc.blocks if not b.name.startswith("*")])
            style_count = len(list(self._doc.styles))

            return CommandResult(ok=True, payload={
                "entity_count": entity_count,
                "by_type": by_type,
                "layer_count": layer_count,
                "block_count": block_count,
                "style_count": style_count,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Extended Query ---

    async def query_text_styles(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            styles = []
            for style in self._doc.styles:
                styles.append({
                    "name": style.dxf.name,
                    "font": style.dxf.get("font", ""),
                    "height": style.dxf.get("height", 0),
                    "width_factor": style.dxf.get("width", 1.0),
                })
            return CommandResult(ok=True, payload={"styles": styles})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def query_dimension_styles(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            dimstyles = [{"name": ds.dxf.name} for ds in self._doc.dimstyles]
            return CommandResult(ok=True, payload={"dimstyles": dimstyles})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def query_linetypes(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            linetypes = [{"name": lt.dxf.name, "description": lt.dxf.get("description", "")} for lt in self._doc.linetypes]
            return CommandResult(ok=True, payload={"linetypes": linetypes})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def query_block_tree(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            blocks = []
            for block in self._doc.blocks:
                try:
                    is_xref = block.block.is_xref or block.block.is_xref_overlay
                except (AttributeError, Exception):
                    is_xref = False
                is_anonymous = block.name.startswith("*")
                entry = {"name": block.name, "is_xref": is_xref, "is_anonymous": is_anonymous, "entity_count": len(block)}
                if hasattr(block, 'attdefs'):
                    attdefs = [{"tag": ad.dxf.tag, "prompt": ad.dxf.get("prompt", "")} for ad in block.attdefs() if hasattr(ad.dxf, 'tag')]
                    if attdefs:
                        entry["attribute_definitions"] = attdefs
                blocks.append(entry)
            return CommandResult(ok=True, payload={"blocks": blocks})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def query_drawing_metadata(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            header = self._doc.header
            return CommandResult(ok=True, payload={
                "version": self._doc.dxfversion,
                "units": header.get("$LUNITS", 0),
                "limmin": list(header.get("$LIMMIN", (0, 0))),
                "limmax": list(header.get("$LIMMAX", (12, 9))),
                "dwgname": self._save_path or "untitled",
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Extended Search ---

    async def search_by_block_name(self, block_name) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            entities = []
            for e in self._msp:
                if e.dxftype() == "INSERT" and e.dxf.name == block_name:
                    entry = {"handle": e.dxf.handle, "block": e.dxf.name, "layer": e.dxf.get("layer", "0"), "position": list(e.dxf.insert)}
                    attribs = {a.dxf.tag: a.dxf.text for a in e.attribs}
                    if attribs:
                        entry["attributes"] = attribs
                    entities.append(entry)
            return CommandResult(ok=True, payload={"entities": entities, "count": len(entities)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def search_by_handle_list(self, handles) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            entities = []
            for h in handles:
                e = self._doc.entitydb.get(h)
                if e:
                    entry = {"handle": e.dxf.handle, "type": e.dxftype(), "layer": e.dxf.get("layer", "0")}
                    entities.append(entry)
            return CommandResult(ok=True, payload={"entities": entities, "count": len(entities)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Equipment Find / Inspect ---

    async def equipment_find(self, pattern, case_sensitive=False, search_scope="all",
                             zoom_to_first=True, zoom_height=600.0, max_results=50) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            import re
            import fnmatch

            # Build regex from pattern — support wildcards via fnmatch
            if "*" in pattern or "?" in pattern:
                regex_pat = fnmatch.translate(pattern)
            else:
                regex_pat = re.escape(pattern)
            flags = 0 if case_sensitive else re.IGNORECASE
            regex = re.compile(regex_pat, flags)

            results = []
            first_pos = None

            # Phase 1: Modelspace TEXT + MTEXT
            if search_scope in ("all", "modelspace"):
                for e in self._msp:
                    if len(results) >= max_results:
                        break
                    dtype = e.dxftype()
                    text_content = None
                    if dtype == "TEXT":
                        text_content = e.dxf.get("text", "")
                    elif dtype == "MTEXT":
                        text_content = e.text
                    if text_content and regex.search(text_content):
                        pos = list(e.dxf.insert) if e.dxf.hasattr("insert") else [0, 0, 0]
                        if first_pos is None:
                            first_pos = pos
                        results.append({
                            "type": dtype, "text": text_content,
                            "layer": e.dxf.get("layer", "0"),
                            "handle": e.dxf.handle,
                            "position": pos, "world_position": pos,
                            "context": "modelspace",
                        })

            # Phase 2: Attribute values on INSERTs
            if search_scope in ("all", "attributes"):
                for e in self._msp:
                    if len(results) >= max_results:
                        break
                    if e.dxftype() != "INSERT":
                        continue
                    ins_pos = list(e.dxf.insert) if e.dxf.hasattr("insert") else [0, 0, 0]
                    for attrib in e.attribs:
                        if len(results) >= max_results:
                            break
                        a_val = attrib.dxf.text
                        if a_val and regex.search(a_val):
                            a_pos = list(attrib.dxf.insert) if attrib.dxf.hasattr("insert") else ins_pos
                            if first_pos is None:
                                first_pos = a_pos
                            results.append({
                                "type": "ATTRIB", "text": a_val,
                                "tag": attrib.dxf.tag,
                                "layer": e.dxf.get("layer", "0"),
                                "handle": e.dxf.handle,
                                "position": a_pos, "world_position": a_pos,
                                "containing_block": e.dxf.name,
                                "insert_handle": e.dxf.handle,
                                "context": "attribute",
                            })

            # Phase 3: Block definition text
            if search_scope in ("all", "blocks"):
                for block in self._doc.blocks:
                    if len(results) >= max_results:
                        break
                    bname = block.name
                    if bname.startswith("*"):
                        continue  # skip anonymous blocks
                    for bent in block:
                        if len(results) >= max_results:
                            break
                        btype = bent.dxftype()
                        bcontent = None
                        if btype == "TEXT":
                            bcontent = bent.dxf.get("text", "")
                        elif btype == "MTEXT":
                            bcontent = bent.text
                        elif btype == "ATTDEF":
                            bcontent = bent.dxf.get("text", "")  # default value
                        if bcontent and regex.search(bcontent):
                            local_pos = list(bent.dxf.insert) if bent.dxf.hasattr("insert") else [0, 0, 0]
                            # Find all INSERTs of this block in modelspace
                            for msp_ent in self._msp:
                                if len(results) >= max_results:
                                    break
                                if msp_ent.dxftype() == "INSERT" and msp_ent.dxf.name == bname:
                                    ins_pt = list(msp_ent.dxf.insert) if msp_ent.dxf.hasattr("insert") else [0, 0, 0]
                                    sx = msp_ent.dxf.get("xscale", 1.0)
                                    sy = msp_ent.dxf.get("yscale", 1.0)
                                    rot = math.radians(msp_ent.dxf.get("rotation", 0.0))
                                    # Transform local to world
                                    lx = local_pos[0] * sx
                                    ly = local_pos[1] * sy
                                    cos_r = math.cos(rot)
                                    sin_r = math.sin(rot)
                                    wx = ins_pt[0] + lx * cos_r - ly * sin_r
                                    wy = ins_pt[1] + lx * sin_r + ly * cos_r
                                    world_pos = [wx, wy, 0.0]
                                    if first_pos is None:
                                        first_pos = world_pos
                                    results.append({
                                        "type": btype, "text": bcontent,
                                        "layer": bent.dxf.get("layer", "0"),
                                        "handle": msp_ent.dxf.handle,
                                        "position": local_pos, "world_position": world_pos,
                                        "containing_block": bname,
                                        "insert_handle": msp_ent.dxf.handle,
                                        "context": "block_definition",
                                    })

            payload = {"count": len(results), "results": results}
            if zoom_to_first and first_pos:
                payload["zoomed_to"] = first_pos
            return CommandResult(ok=True, payload=payload)
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def equipment_inspect(self, x, y, view_width=600.0, view_height=600.0,
                                infer_center=True, handle=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            from ezdxf import bbox as ezdxf_bbox

            half_w = view_width / 2.0
            half_h = view_height / 2.0
            x1, y1 = x - half_w, y - half_h
            x2, y2 = x + half_w, y + half_h

            # Collect entities in window
            total = 0
            by_type = {}
            blocks = []
            circles = []
            largest_circle = None
            nearest_insert = None
            nearest_insert_dist = float("inf")

            for e in self._msp:
                try:
                    box = ezdxf_bbox.extents([e])
                    if box.has_data:
                        ebox_min = box.extmin
                        ebox_max = box.extmax
                        # Check overlap with window
                        if ebox_max[0] < x1 or ebox_min[0] > x2 or ebox_max[1] < y1 or ebox_min[1] > y2:
                            continue
                    else:
                        continue
                except Exception:
                    continue

                total += 1
                dtype = e.dxftype()
                by_type[dtype] = by_type.get(dtype, 0) + 1

                if dtype == "CIRCLE":
                    center = list(e.dxf.center)
                    radius = e.dxf.radius
                    circles.append({"handle": e.dxf.handle, "center": center, "radius": radius})
                    if largest_circle is None or radius > largest_circle["radius"]:
                        largest_circle = {"center": center, "radius": radius}

                elif dtype == "INSERT":
                    ins_pt = list(e.dxf.insert) if e.dxf.hasattr("insert") else [0, 0, 0]
                    blocks.append({
                        "handle": e.dxf.handle,
                        "block_name": e.dxf.name,
                        "position": ins_pt,
                    })
                    # Only consider for nearest-insert if insertion point is inside view window
                    # (large blocks/xrefs may have insertion far from the viewed area)
                    if x1 <= ins_pt[0] <= x2 and y1 <= ins_pt[1] <= y2:
                        dist = math.hypot(ins_pt[0] - x, ins_pt[1] - y)
                        if dist < nearest_insert_dist:
                            nearest_insert_dist = dist
                            nearest_insert = ins_pt

            # Center inference
            eq_x, eq_y = x, y
            method = "fallback"
            confidence = "low"
            bbox_val = None

            if infer_center:
                # Priority 1: specific INSERT by handle
                if handle:
                    try:
                        target = self._doc.entitydb.get(handle)
                        if target and target.dxftype() == "INSERT":
                            box = ezdxf_bbox.extents([target])
                            if box.has_data:
                                eq_x = (box.extmin[0] + box.extmax[0]) / 2.0
                                eq_y = (box.extmin[1] + box.extmax[1]) / 2.0
                                bbox_val = [box.extmin[0], box.extmin[1], box.extmax[0], box.extmax[1]]
                                method = "insert_bbox"
                                confidence = "high"
                            else:
                                ins_pt = list(target.dxf.insert) if target.dxf.hasattr("insert") else [0, 0, 0]
                                eq_x, eq_y = ins_pt[0], ins_pt[1]
                                method = "insert_point"
                                confidence = "medium"
                    except Exception:
                        pass

                # Priority 2: largest circle
                if method == "fallback" and largest_circle:
                    eq_x = largest_circle["center"][0]
                    eq_y = largest_circle["center"][1]
                    r = largest_circle["radius"]
                    bbox_val = [eq_x - r, eq_y - r, eq_x + r, eq_y + r]
                    method = "largest_circle"
                    confidence = "high"

                # Priority 3: nearest INSERT
                if method == "fallback" and nearest_insert:
                    eq_x, eq_y = nearest_insert[0], nearest_insert[1]
                    method = "nearest_insert"
                    confidence = "medium"

            payload = {
                "view_center": [x, y, 0.0],
                "equipment_center": {
                    "x": eq_x, "y": eq_y,
                    "method": method, "confidence": confidence,
                    "bbox": bbox_val,
                },
                "nearby_entities": {
                    "total": total,
                    "by_type": by_type,
                    "blocks": blocks,
                    "circles": circles,
                },
            }
            return CommandResult(ok=True, payload=payload)
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Deep Text Search ---

    async def find_text(self, pattern, case_sensitive=False, max_results=50,
                        zoom_to_first=True, zoom_height=600.0) -> CommandResult:
        """Deep text search across modelspace and all block definitions."""
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            import re
            import fnmatch

            if "*" in pattern or "?" in pattern:
                regex_pat = fnmatch.translate(pattern)
            else:
                regex_pat = re.escape(pattern)
            flags = 0 if case_sensitive else re.IGNORECASE
            regex = re.compile(regex_pat, flags)

            results = []
            first_pos = None

            # Phase 1: Modelspace TEXT/MTEXT/DIMENSION
            for e in self._msp:
                if len(results) >= max_results:
                    break
                dtype = e.dxftype()
                text_content = None
                if dtype == "TEXT":
                    text_content = e.dxf.get("text", "")
                elif dtype == "MTEXT":
                    text_content = e.text
                elif dtype == "DIMENSION":
                    text_content = e.dxf.get("text", "")
                if text_content and regex.search(text_content):
                    pos = list(e.dxf.insert) if e.dxf.hasattr("insert") else [0, 0, 0]
                    if first_pos is None:
                        first_pos = pos
                    results.append({
                        "type": dtype, "text": text_content,
                        "layer": e.dxf.get("layer", "0"),
                        "handle": e.dxf.handle,
                        "position": pos, "world_position": pos,
                        "context": "modelspace",
                    })

            # Phase 2: Modelspace INSERT → walk ATTRIBs
            for e in self._msp:
                if len(results) >= max_results:
                    break
                if e.dxftype() != "INSERT":
                    continue
                ins_pos = list(e.dxf.insert) if e.dxf.hasattr("insert") else [0, 0, 0]
                for attrib in e.attribs:
                    if len(results) >= max_results:
                        break
                    a_val = attrib.dxf.text
                    if a_val and regex.search(a_val):
                        a_pos = list(attrib.dxf.insert) if attrib.dxf.hasattr("insert") else ins_pos
                        if first_pos is None:
                            first_pos = a_pos
                        results.append({
                            "type": "ATTRIB", "text": a_val,
                            "tag": attrib.dxf.tag,
                            "layer": e.dxf.get("layer", "0"),
                            "handle": attrib.dxf.handle,
                            "position": a_pos, "world_position": a_pos,
                            "containing_block": e.dxf.name,
                            "insert_handle": e.dxf.handle,
                            "context": "attribute",
                        })

            # Phase 3: Block definition text (TEXT/MTEXT/ATTDEF/DIMENSION)
            for block in self._doc.blocks:
                if len(results) >= max_results:
                    break
                bname = block.name
                if bname.startswith("*"):
                    continue
                for bent in block:
                    if len(results) >= max_results:
                        break
                    btype = bent.dxftype()
                    bcontent = None
                    if btype == "TEXT":
                        bcontent = bent.dxf.get("text", "")
                    elif btype == "MTEXT":
                        bcontent = bent.text
                    elif btype == "ATTDEF":
                        bcontent = bent.dxf.get("text", "")
                    elif btype == "DIMENSION":
                        bcontent = bent.dxf.get("text", "")
                    if bcontent and regex.search(bcontent):
                        local_pos = list(bent.dxf.insert) if bent.dxf.hasattr("insert") else [0, 0, 0]
                        # Find INSERT references in modelspace
                        found_insert = False
                        for msp_ent in self._msp:
                            if len(results) >= max_results:
                                break
                            if msp_ent.dxftype() == "INSERT" and msp_ent.dxf.name == bname:
                                found_insert = True
                                ins_pt = list(msp_ent.dxf.insert) if msp_ent.dxf.hasattr("insert") else [0, 0, 0]
                                sx = msp_ent.dxf.get("xscale", 1.0)
                                sy = msp_ent.dxf.get("yscale", 1.0)
                                rot = math.radians(msp_ent.dxf.get("rotation", 0.0))
                                lx = local_pos[0] * sx
                                ly = local_pos[1] * sy
                                cos_r = math.cos(rot)
                                sin_r = math.sin(rot)
                                wx = ins_pt[0] + lx * cos_r - ly * sin_r
                                wy = ins_pt[1] + lx * sin_r + ly * cos_r
                                world_pos = [wx, wy, 0.0]
                                if first_pos is None:
                                    first_pos = world_pos
                                results.append({
                                    "type": btype, "text": bcontent,
                                    "layer": bent.dxf.get("layer", "0"),
                                    "handle": msp_ent.dxf.handle,
                                    "position": local_pos, "world_position": world_pos,
                                    "containing_block": bname,
                                    "insert_handle": msp_ent.dxf.handle,
                                    "context": "block_definition",
                                })
                        if not found_insert:
                            if first_pos is None:
                                first_pos = local_pos
                            results.append({
                                "type": btype, "text": bcontent,
                                "layer": bent.dxf.get("layer", "0"),
                                "handle": bent.dxf.handle,
                                "position": local_pos, "world_position": local_pos,
                                "containing_block": bname,
                                "context": "nested_block_definition",
                            })

            payload = {"count": len(results), "results": results}
            if zoom_to_first and first_pos:
                payload["zoomed_to"] = first_pos
            return CommandResult(ok=True, payload=payload)
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Equipment Tag Placement ---

    async def place_equipment_tag(self, cx, cy, cz=0.0, tag="", cube_size=24.0,
                                  direction="right", text_height=8.0) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            half = cube_size / 2.0
            ds = -1.0 if direction == "left" else 1.0

            # Ensure layers
            self._ensure_layer("E-EQPM-N")
            self._ensure_layer("E-ANNO-TEXT")

            # Approximate text width for Arial Narrow
            text_width = text_height * 0.48 * len(tag)

            # Leader vertices
            v1x = cx + ds * half
            v1y = cy + half
            v2x = v1x + ds * 24.0
            v2y = v1y + 48.0
            v3x = v2x + ds * 4.0
            v3y = v2y

            # MTEXT position
            if direction == "left":
                mtx = v3x - 4.0 - text_width
            else:
                mtx = v3x + 4.0
            mty = v3y + 4.0

            # Underline
            uly = mty - 9.6
            ulx1 = mtx
            ulx2 = mtx + text_width

            # --- Create polyface mesh cube ---
            corners = [
                (cx - half, cy - half, cz + half),
                (cx + half, cy - half, cz + half),
                (cx + half, cy - half, cz - half),
                (cx - half, cy - half, cz - half),
                (cx - half, cy + half, cz + half),
                (cx + half, cy + half, cz + half),
                (cx + half, cy + half, cz - half),
                (cx - half, cy + half, cz - half),
            ]
            faces = [
                (0, 1, 2, 3), (4, 5, 6, 7), (0, 1, 5, 4),
                (3, 2, 6, 7), (0, 4, 7, 3), (1, 5, 6, 2),
            ]
            pface = self._msp.add_polyface()
            for corner in corners:
                pface.append_vertices([corner])
            for f in faces:
                pface.append_face([corners[i] for i in f])
            pface.close()
            pface.dxf.layer = "E-EQPM-N"
            cube_handle = pface.dxf.handle

            # --- Create LEADER ---
            leader = self._msp.new_entity(
                "LEADER",
                dxfattribs={
                    "layer": "E-ANNO-TEXT",
                    "dimstyle": "Standard",
                },
            )
            leader.set_vertices([(v1x, v1y, 0), (v2x, v2y, 0), (v3x, v3y, 0)])
            leader_handle = leader.dxf.handle

            # --- Create MTEXT ---
            mtext = self._msp.add_mtext(tag, dxfattribs={
                "insert": (mtx, mty, 0),
                "char_height": text_height,
                "width": text_width,
                "layer": "E-ANNO-TEXT",
                "style": "Standard",
                "attachment_point": 1,
            })
            mtext_handle = mtext.dxf.handle

            # --- Create underline LINE ---
            uline = self._msp.add_line(
                (ulx1, uly, 0), (ulx2, uly, 0),
                dxfattribs={"layer": "E-ANNO-TEXT"},
            )
            line_handle = uline.dxf.handle

            # Bounding box
            bbox_minx = min(cx - half, ulx1, mtx)
            bbox_miny = min(cy - half, uly)
            bbox_maxx = max(cx + half, ulx2, mtx + text_width)
            bbox_maxy = max(cy + half, mty + text_height)

            return CommandResult(ok=True, payload={
                "cube_handle": cube_handle,
                "leader_handle": leader_handle,
                "mtext_handle": mtext_handle,
                "line_handle": line_handle,
                "center": {"x": cx, "y": cy, "z": cz},
                "tag": tag,
                "text_width": text_width,
                "bbox": {
                    "min_x": bbox_minx, "min_y": bbox_miny,
                    "max_x": bbox_maxx, "max_y": bbox_maxy,
                },
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Batch Find and Tag ---

    async def batch_find_and_tag(self, tags, cube_size=24.0, direction="right",
                                  text_height=8.0) -> CommandResult:
        """Find multiple tags and place equipment tag groups at each location."""
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            results = []
            not_found = []
            placed = 0

            for tag in tags:
                if not tag:
                    continue
                # Reuse find_text to locate the tag (first match only)
                find_result = await self.find_text(
                    tag, case_sensitive=False, max_results=1,
                    zoom_to_first=False, zoom_height=600.0,
                )
                if find_result.ok and find_result.payload.get("count", 0) > 0:
                    match = find_result.payload["results"][0]
                    pos = match.get("world_position", match.get("position", [0, 0, 0]))
                    cx, cy = pos[0], pos[1]
                    cz = pos[2] if len(pos) > 2 else 0.0

                    place_result = await self.place_equipment_tag(
                        cx, cy, cz, tag, cube_size, direction, text_height,
                    )
                    if place_result.ok:
                        entry = place_result.payload.copy()
                        entry["tag"] = tag
                        entry["status"] = "placed"
                        entry["position"] = pos
                        results.append(entry)
                        placed += 1
                    else:
                        not_found.append(tag)
                        results.append({"tag": tag, "status": "place_failed",
                                        "error": place_result.error})
                else:
                    not_found.append(tag)
                    results.append({"tag": tag, "status": "not_found"})

            return CommandResult(ok=True, payload={
                "placed": placed,
                "not_found": not_found,
                "results": results,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Extended Entity Operations ---

    async def entity_explode(self, entity_id) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            e = self._doc.entitydb.get(entity_id)
            if not e:
                return CommandResult(ok=False, error=f"Entity not found: {entity_id}")
            dtype = e.dxftype()
            new_handles = []

            if dtype == "INSERT":
                # Explode block reference: copy block entities into modelspace
                block_name = e.dxf.name
                if block_name not in self._doc.blocks:
                    return CommandResult(ok=False, error=f"Block definition '{block_name}' not found")
                block = self._doc.blocks[block_name]
                insert_pt = e.dxf.insert
                xscale = e.dxf.get("xscale", 1.0)
                yscale = e.dxf.get("yscale", 1.0)
                rotation = e.dxf.get("rotation", 0.0)
                from ezdxf.math import Matrix44
                m = Matrix44.chain(
                    Matrix44.scale(xscale, yscale, 1.0),
                    Matrix44.z_rotate(math.radians(rotation)),
                    Matrix44.translate(insert_pt[0], insert_pt[1], insert_pt[2] if len(insert_pt) > 2 else 0),
                )
                for block_entity in block:
                    if block_entity.dxftype() == "ATTDEF":
                        continue  # Skip attribute definitions
                    try:
                        copy = block_entity.copy()
                        self._msp.add_entity(copy)
                        copy.transform(m)
                        new_handles.append(copy.dxf.handle)
                    except Exception:
                        pass
                # Delete the original INSERT
                try:
                    self._msp.delete_entity(e)
                except Exception:
                    pass

            elif dtype == "LWPOLYLINE":
                # Explode polyline into individual LINE segments
                pts = list(e.get_points(format="xyseb"))
                layer = e.dxf.get("layer", "0")
                n = len(pts)
                seg_count = n if e.closed else n - 1
                for i in range(seg_count):
                    j = (i + 1) % n
                    x1, y1 = pts[i][0], pts[i][1]
                    x2, y2 = pts[j][0], pts[j][1]
                    bulge = pts[i][4]
                    if abs(bulge) < 1e-10:
                        seg = self._msp.add_line((x1, y1), (x2, y2), dxfattribs={"layer": layer})
                    else:
                        # Convert bulge segment to arc
                        chord = math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
                        sagitta = abs(bulge) * chord / 2
                        radius = (chord ** 2 / 4 + sagitta ** 2) / (2 * sagitta)
                        # Center of arc
                        mx, my = (x1 + x2) / 2, (y1 + y2) / 2
                        dx, dy = x2 - x1, y2 - y1
                        d = math.sqrt(dx ** 2 + dy ** 2)
                        if d < 1e-10:
                            continue
                        # Perpendicular direction
                        nx, ny = -dy / d, dx / d
                        sign = 1 if bulge > 0 else -1
                        offset = sign * (radius - sagitta)
                        cx = mx + nx * offset
                        cy = my + ny * offset
                        sa = math.degrees(math.atan2(y1 - cy, x1 - cx))
                        ea = math.degrees(math.atan2(y2 - cy, x2 - cx))
                        if bulge > 0 and ea < sa:
                            ea += 360
                        elif bulge < 0 and sa < ea:
                            sa += 360
                        if bulge < 0:
                            sa, ea = ea, sa
                        seg = self._msp.add_arc((cx, cy), radius, sa, ea, dxfattribs={"layer": layer})
                    new_handles.append(seg.dxf.handle)
                # Delete original polyline
                try:
                    self._msp.delete_entity(e)
                except Exception:
                    pass

            else:
                return CommandResult(ok=True, payload={
                    "entity_id": entity_id,
                    "note": f"Explode not applicable for {dtype}",
                })

            return CommandResult(ok=True, payload={
                "exploded": entity_id,
                "new_entities": new_handles,
                "count": len(new_handles),
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_join(self, entity_ids) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            if not entity_ids or len(entity_ids) < 2:
                return CommandResult(ok=False, error="At least 2 entity IDs required for join")

            # Collect all LINE endpoints to build a polyline
            segments = []
            layer = None
            for eid in entity_ids:
                e = self._doc.entitydb.get(eid)
                if e is None:
                    return CommandResult(ok=False, error=f"Entity {eid} not found")
                dtype = e.dxftype()
                if layer is None:
                    layer = e.dxf.get("layer", "0")
                if dtype == "LINE":
                    segments.append((
                        (e.dxf.start[0], e.dxf.start[1]),
                        (e.dxf.end[0], e.dxf.end[1]),
                    ))
                elif dtype == "LWPOLYLINE":
                    pts = list(e.get_points(format="xy"))
                    for i in range(len(pts) - 1):
                        segments.append((pts[i], pts[i + 1]))
                    if e.closed and len(pts) > 1:
                        segments.append((pts[-1], pts[0]))
                else:
                    return CommandResult(ok=False, error=f"Cannot join entity type {dtype}")

            if not segments:
                return CommandResult(ok=False, error="No joinable segments found")

            # Chain segments by matching endpoints
            tolerance = 0.01
            chain = [segments[0][0], segments[0][1]]
            used = {0}
            changed = True
            while changed:
                changed = False
                for i, (sp, ep) in enumerate(segments):
                    if i in used:
                        continue
                    # Try to append to end of chain
                    dx1 = chain[-1][0] - sp[0]
                    dy1 = chain[-1][1] - sp[1]
                    if math.sqrt(dx1 * dx1 + dy1 * dy1) < tolerance:
                        chain.append(ep)
                        used.add(i)
                        changed = True
                        continue
                    dx2 = chain[-1][0] - ep[0]
                    dy2 = chain[-1][1] - ep[1]
                    if math.sqrt(dx2 * dx2 + dy2 * dy2) < tolerance:
                        chain.append(sp)
                        used.add(i)
                        changed = True
                        continue
                    # Try to prepend to start of chain
                    dx3 = chain[0][0] - ep[0]
                    dy3 = chain[0][1] - ep[1]
                    if math.sqrt(dx3 * dx3 + dy3 * dy3) < tolerance:
                        chain.insert(0, sp)
                        used.add(i)
                        changed = True
                        continue
                    dx4 = chain[0][0] - sp[0]
                    dy4 = chain[0][1] - sp[1]
                    if math.sqrt(dx4 * dx4 + dy4 * dy4) < tolerance:
                        chain.insert(0, ep)
                        used.add(i)
                        changed = True
                        continue

            # Check if closed
            d_close = math.sqrt((chain[0][0] - chain[-1][0]) ** 2 + (chain[0][1] - chain[-1][1]) ** 2)
            is_closed = d_close < tolerance

            # Create the new polyline
            new_e = self._msp.add_lwpolyline(chain, close=is_closed, dxfattribs={"layer": layer or "0"})

            # Delete original entities
            deleted = 0
            for eid in entity_ids:
                try:
                    e = self._doc.entitydb.get(eid)
                    if e:
                        self._msp.delete_entity(e)
                        deleted += 1
                except Exception:
                    pass

            return CommandResult(ok=True, payload={
                "handle": new_e.dxf.handle,
                "vertex_count": len(chain),
                "closed": is_closed,
                "segments_joined": len(used),
                "segments_total": len(segments),
                "originals_deleted": deleted,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_extend(self, entity_id, boundary_id) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            boundary = self._doc.entitydb.get(boundary_id)
            if boundary is None:
                return CommandResult(ok=False, error=f"Boundary entity {boundary_id} not found")

            if e.dxftype() != "LINE":
                return CommandResult(ok=False, error="Extend only supported for LINE entities in ezdxf backend")

            # Get line direction
            sx, sy = e.dxf.start[0], e.dxf.start[1]
            ex_pt, ey = e.dxf.end[0], e.dxf.end[1]
            dx, dy = ex_pt - sx, ey - sy
            line_len = math.sqrt(dx * dx + dy * dy)
            if line_len < 1e-10:
                return CommandResult(ok=False, error="Line has zero length")

            # Find intersection with boundary
            bdtype = boundary.dxftype()
            intersection = None

            if bdtype == "LINE":
                # Line-line intersection
                bsx, bsy = boundary.dxf.start[0], boundary.dxf.start[1]
                bex, bey = boundary.dxf.end[0], boundary.dxf.end[1]
                bdx, bdy = bex - bsx, bey - bsy

                denom = dx * bdy - dy * bdx
                if abs(denom) < 1e-10:
                    return CommandResult(ok=False, error="Line is parallel to boundary")

                t = ((bsx - sx) * bdy - (bsy - sy) * bdx) / denom
                s = ((bsx - sx) * dy - (bsy - sy) * dx) / denom

                if 0 <= s <= 1:  # Intersection is on the boundary segment
                    ix = sx + t * dx
                    iy = sy + t * dy
                    intersection = (ix, iy, t)

            elif bdtype == "CIRCLE":
                # Line-circle intersection
                ccx, ccy = boundary.dxf.center[0], boundary.dxf.center[1]
                r = boundary.dxf.radius
                # Parametric: P = S + t * D
                fx, fy = sx - ccx, sy - ccy
                a = dx * dx + dy * dy
                b = 2 * (fx * dx + fy * dy)
                c = fx * fx + fy * fy - r * r
                disc = b * b - 4 * a * c
                if disc < 0:
                    return CommandResult(ok=False, error="Line does not intersect boundary circle")
                sqrt_disc = math.sqrt(disc)
                t1 = (-b + sqrt_disc) / (2 * a)
                t2 = (-b - sqrt_disc) / (2 * a)
                # Pick the intersection beyond the line endpoint (t > 1 for extending end)
                best_t = None
                for t in [t1, t2]:
                    if t > 1:
                        if best_t is None or t < best_t:
                            best_t = t
                if best_t is None:
                    # Try extending from start (t < 0)
                    for t in [t1, t2]:
                        if t < 0:
                            if best_t is None or t > best_t:
                                best_t = t
                if best_t is not None:
                    intersection = (sx + best_t * dx, sy + best_t * dy, best_t)

            if intersection is None:
                return CommandResult(ok=False, error="No valid intersection found with boundary")

            ix, iy, t = intersection
            # Determine which end to extend
            if t > 1:
                e.dxf.end = (ix, iy, 0)
                extended_end = "end"
            elif t < 0:
                e.dxf.start = (ix, iy, 0)
                extended_end = "start"
            else:
                return CommandResult(ok=False, error="Intersection is within the line, no extension needed")

            return CommandResult(ok=True, payload={
                "handle": e.dxf.handle,
                "extended_end": extended_end,
                "new_point": [round(ix, 6), round(iy, 6)],
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_trim(self, entity_id, boundary_id) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            boundary = self._doc.entitydb.get(boundary_id)
            if boundary is None:
                return CommandResult(ok=False, error=f"Boundary entity {boundary_id} not found")

            if e.dxftype() != "LINE":
                return CommandResult(ok=False, error="Trim only supported for LINE entities in ezdxf backend")

            sx, sy = e.dxf.start[0], e.dxf.start[1]
            ex_pt, ey = e.dxf.end[0], e.dxf.end[1]
            dx, dy = ex_pt - sx, ey - sy

            # Find intersection with boundary
            bdtype = boundary.dxftype()
            intersection = None

            if bdtype == "LINE":
                bsx, bsy = boundary.dxf.start[0], boundary.dxf.start[1]
                bex, bey = boundary.dxf.end[0], boundary.dxf.end[1]
                bdx, bdy = bex - bsx, bey - bsy

                denom = dx * bdy - dy * bdx
                if abs(denom) < 1e-10:
                    return CommandResult(ok=False, error="Line is parallel to boundary")

                t = ((bsx - sx) * bdy - (bsy - sy) * bdx) / denom
                s = ((bsx - sx) * dy - (bsy - sy) * dx) / denom

                if 0 <= t <= 1 and 0 <= s <= 1:
                    intersection = (sx + t * dx, sy + t * dy, t)

            elif bdtype == "CIRCLE":
                ccx, ccy = boundary.dxf.center[0], boundary.dxf.center[1]
                r = boundary.dxf.radius
                fx, fy = sx - ccx, sy - ccy
                a = dx * dx + dy * dy
                b = 2 * (fx * dx + fy * dy)
                c = fx * fx + fy * fy - r * r
                disc = b * b - 4 * a * c
                if disc < 0:
                    return CommandResult(ok=False, error="Line does not intersect boundary")
                sqrt_disc = math.sqrt(disc)
                t1 = (-b + sqrt_disc) / (2 * a)
                t2 = (-b - sqrt_disc) / (2 * a)
                # Pick intersection on the line segment (0 <= t <= 1)
                best_t = None
                for t in [t1, t2]:
                    if 0 <= t <= 1:
                        if best_t is None or abs(t - 0.5) < abs(best_t - 0.5):
                            best_t = t
                if best_t is not None:
                    intersection = (sx + best_t * dx, sy + best_t * dy, best_t)

            if intersection is None:
                return CommandResult(ok=False, error="No intersection found on segment")

            ix, iy, t = intersection
            # Trim the longer portion (keep the shorter side)
            # If intersection is closer to end, trim end; closer to start, trim start
            if t >= 0.5:
                e.dxf.end = (ix, iy, 0)
                trimmed_end = "end"
            else:
                e.dxf.start = (ix, iy, 0)
                trimmed_end = "start"

            return CommandResult(ok=True, payload={
                "handle": e.dxf.handle,
                "trimmed_end": trimmed_end,
                "trim_point": [round(ix, 6), round(iy, 6)],
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def entity_break_at(self, entity_id, x, y) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            e = self._doc.entitydb.get(entity_id)
            if e is None:
                return CommandResult(ok=False, error=f"Entity {entity_id} not found")
            dtype = e.dxftype()
            layer = e.dxf.get("layer", "0")

            if dtype == "LINE":
                sx, sy = e.dxf.start[0], e.dxf.start[1]
                ex_pt, ey = e.dxf.end[0], e.dxf.end[1]

                # Create two new lines split at the break point
                line1 = self._msp.add_line((sx, sy), (x, y), dxfattribs={"layer": layer})
                line2 = self._msp.add_line((x, y), (ex_pt, ey), dxfattribs={"layer": layer})

                # Copy color/linetype properties
                for attr in ("color", "linetype", "lineweight"):
                    if e.dxf.hasattr(attr):
                        val = e.dxf.get(attr)
                        line1.dxf.set(attr, val)
                        line2.dxf.set(attr, val)

                # Delete original
                self._msp.delete_entity(e)

                return CommandResult(ok=True, payload={
                    "broken": entity_id,
                    "break_point": [x, y],
                    "new_handles": [line1.dxf.handle, line2.dxf.handle],
                })

            elif dtype == "LWPOLYLINE":
                pts = list(e.get_points(format="xy"))
                # Find closest segment to break point
                best_idx = 0
                best_dist = float("inf")
                n = len(pts)
                seg_count = n if e.closed else n - 1
                for i in range(seg_count):
                    j = (i + 1) % n
                    ax, ay = pts[i]
                    bx, by = pts[j]
                    # Point-to-segment distance
                    abx, aby = bx - ax, by - ay
                    apx, apy = x - ax, y - ay
                    ab_sq = abx * abx + aby * aby
                    if ab_sq > 0:
                        t = max(0, min(1, (apx * abx + apy * aby) / ab_sq))
                    else:
                        t = 0
                    px, py = ax + t * abx, ay + t * aby
                    dist = math.sqrt((x - px) ** 2 + (y - py) ** 2)
                    if dist < best_dist:
                        best_dist = dist
                        best_idx = i

                # Split polyline at break point after best_idx
                pts1 = list(pts[:best_idx + 1]) + [(x, y)]
                pts2 = [(x, y)] + list(pts[best_idx + 1:])

                poly1 = self._msp.add_lwpolyline(pts1, dxfattribs={"layer": layer})
                poly2 = self._msp.add_lwpolyline(pts2, dxfattribs={"layer": layer})

                self._msp.delete_entity(e)

                return CommandResult(ok=True, payload={
                    "broken": entity_id,
                    "break_point": [x, y],
                    "new_handles": [poly1.dxf.handle, poly2.dxf.handle],
                })

            else:
                return CommandResult(ok=False, error=f"Break not supported for {dtype}")
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Extended Validate ---

    async def validate_text_standards(self, allowed_styles=None, allowed_heights=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            violations = []
            for e in self._msp:
                dtype = e.dxftype()
                if dtype in ("TEXT", "MTEXT"):
                    style = e.dxf.get("style", "Standard")
                    height = e.dxf.get("char_height", 0) if dtype == "MTEXT" else e.dxf.get("height", 0)
                    if allowed_styles and style not in allowed_styles:
                        violations.append({"handle": e.dxf.handle, "type": dtype, "issue": f"non-standard style: {style}", "style": style, "height": height})
                    elif allowed_heights and height not in [float(h) for h in allowed_heights]:
                        violations.append({"handle": e.dxf.handle, "type": dtype, "issue": f"non-standard height: {height}", "style": style, "height": height})
            return CommandResult(ok=True, payload={"pass": len(violations) == 0, "violation_count": len(violations), "violations": violations})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def validate_orphaned_entities(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            issues = []
            for e in self._msp:
                layer_name = e.dxf.get("layer", "0")
                try:
                    layer = self._doc.layers.get(layer_name)
                    off = layer.is_off() if callable(layer.is_off) else layer.is_off
                    frozen = layer.is_frozen() if callable(layer.is_frozen) else layer.is_frozen
                    if off or frozen:
                        issues.append({"handle": e.dxf.handle, "layer": layer_name, "type": e.dxftype(), "issue": "layer off" if off else "layer frozen"})
                except Exception:
                    pass
            return CommandResult(ok=True, payload={"issue_count": len(issues), "issues": issues})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def validate_attribute_completeness(self, required_tags=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            issues = []
            for e in self._msp:
                if e.dxftype() == "INSERT":
                    if required_tags:
                        found_tags = {a.dxf.tag: a.dxf.text for a in e.attribs}
                        for tag in required_tags:
                            if tag not in found_tags:
                                issues.append({"handle": e.dxf.handle, "block": e.dxf.name, "tag": tag, "issue": "missing tag"})
                            elif not found_tags[tag]:
                                issues.append({"handle": e.dxf.handle, "block": e.dxf.name, "tag": tag, "issue": "empty value"})
            return CommandResult(ok=True, payload={"issue_count": len(issues), "issues": issues})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def validate_connectivity(self, layer=None, tolerance=0.01) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            endpoints = []
            for e in self._msp:
                if e.dxftype() == "LINE":
                    if layer and e.dxf.get("layer", "0") != layer:
                        continue
                    endpoints.append((e.dxf.handle, "start", (e.dxf.start[0], e.dxf.start[1])))
                    endpoints.append((e.dxf.handle, "end", (e.dxf.end[0], e.dxf.end[1])))
            dangling = []
            for h, end_type, pt in endpoints:
                connected = False
                for h2, end_type2, pt2 in endpoints:
                    if h == h2:
                        continue
                    dist = math.sqrt((pt[0] - pt2[0]) ** 2 + (pt[1] - pt2[1]) ** 2)
                    if dist < tolerance:
                        connected = True
                        break
                if not connected:
                    dangling.append({"handle": h, "end": end_type, "point": list(pt)})
            return CommandResult(ok=True, payload={"dangling_count": len(dangling), "total_endpoints": len(endpoints), "dangling": dangling[:100]})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Extended Select ---

    async def find_replace_attribute(self, tag, find, replace) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            replaced = []
            for e in self._msp:
                if e.dxftype() == "INSERT":
                    for a in e.attribs:
                        if a.dxf.tag == tag and find in a.dxf.text:
                            old = a.dxf.text
                            a.dxf.text = a.dxf.text.replace(find, replace)
                            replaced.append({"handle": e.dxf.handle, "tag": tag, "old": old, "new": a.dxf.text})
            return CommandResult(ok=True, payload={"replaced": len(replaced), "entities": replaced})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layer_rename(self, old_name, new_name) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            layer = self._doc.layers.get(old_name)
            layer.dxf.name = new_name
            for e in self._msp:
                if e.dxf.get("layer", "0") == old_name:
                    e.dxf.layer = new_name
            return CommandResult(ok=True, payload={"renamed": old_name, "to": new_name})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layer_merge(self, source_layer, target_layer) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            self._ensure_layer(target_layer)
            count = 0
            for e in self._msp:
                if e.dxf.get("layer", "0") == source_layer:
                    e.dxf.layer = target_layer
                    count += 1
            return CommandResult(ok=True, payload={"merged": count, "from": source_layer, "to": target_layer})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Enhanced View ---

    async def zoom_scale(self, factor) -> CommandResult:
        try:
            return CommandResult(ok=True, payload={"ok": True, "note": "zoom_scale is a no-op in ezdxf"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def pan(self, dx, dy) -> CommandResult:
        try:
            return CommandResult(ok=True, payload={"ok": True, "note": "pan is a no-op in ezdxf"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Enhanced Drawing ---

    async def drawing_audit(self, fix=False) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            from ezdxf import audit as ezdxf_audit
            auditor = self._doc.audit()
            issues = [{"severity": str(e.severity), "message": str(e)} for e in auditor.errors[:50]]
            if fix:
                auditor.fix_all()
            return CommandResult(ok=True, payload={"issues": issues, "issue_count": len(auditor.errors), "fixed": fix})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def drawing_units(self, units=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            if units is not None:
                self._doc.header['$LUNITS'] = int(units)
            return CommandResult(ok=True, payload={"units": self._doc.header.get('$LUNITS', 0)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def drawing_limits(self, x1=None, y1=None, x2=None, y2=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            if all(v is not None for v in [x1, y1, x2, y2]):
                self._doc.header['$LIMMIN'] = (x1, y1)
                self._doc.header['$LIMMAX'] = (x2, y2)
            limmin = self._doc.header.get('$LIMMIN', (0, 0))
            limmax = self._doc.header.get('$LIMMAX', (12, 9))
            return CommandResult(ok=True, payload={"limmin": list(limmin), "limmax": list(limmax)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def drawing_wblock(self, handles, path) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            new_doc = ezdxf.new(self._doc.dxfversion)
            new_msp = new_doc.modelspace()
            for h in handles:
                e = self._doc.entitydb.get(h)
                if e:
                    new_msp.add_entity(e.copy())
            new_doc.saveas(path)
            return CommandResult(ok=True, payload={"exported": len(handles), "path": path})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- XREF ---

    async def xref_list(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            xrefs = []
            for block_layout in self._doc.blocks:
                try:
                    block_entity = block_layout.block
                    is_xref = block_entity.is_xref or block_entity.is_xref_overlay
                    if is_xref:
                        xref_path = block_entity.dxf.get("xref_path", "")
                        xrefs.append({
                            "name": block_layout.name,
                            "path": xref_path,
                            "type": "overlay" if block_entity.is_xref_overlay else "attach",
                            "entity_count": len(block_layout),
                        })
                except (AttributeError, Exception):
                    continue
            return CommandResult(ok=True, payload={"xrefs": xrefs, "count": len(xrefs)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def xref_attach(self, path, x=0, y=0, scale=1.0, overlay=False) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            xref_name = Path(path).stem
            # Create xref block definition
            flags = 8  # XREF flag
            if overlay:
                flags |= 8  # overlay
            block = self._doc.blocks.new(name=xref_name)
            block.dxf.xref_path = path
            # Insert the xref
            e = self._msp.add_blockref(xref_name, (x, y), dxfattribs={
                "xscale": scale, "yscale": scale, "zscale": scale,
            })
            return CommandResult(ok=True, payload={
                "name": xref_name,
                "path": path,
                "handle": e.dxf.handle,
                "type": "overlay" if overlay else "attach",
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def xref_detach(self, name) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            if name not in self._doc.blocks:
                return CommandResult(ok=False, error=f"Block/xref '{name}' not found")
            block_layout = self._doc.blocks[name]
            try:
                is_xref = block_layout.block.is_xref or block_layout.block.is_xref_overlay
            except AttributeError:
                is_xref = False
            if not is_xref:
                return CommandResult(ok=False, error=f"'{name}' is not an xref")
            # Remove all INSERT references to this xref in modelspace
            inserts_removed = 0
            to_delete = [e for e in self._msp if e.dxftype() == "INSERT" and e.dxf.name == name]
            for e in to_delete:
                self._msp.delete_entity(e)
                inserts_removed += 1
            # Delete the block definition
            try:
                self._doc.blocks.delete_block(name, safe=False)
            except Exception:
                pass
            return CommandResult(ok=True, payload={
                "detached": name,
                "inserts_removed": inserts_removed,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def xref_reload(self, name) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            return CommandResult(ok=True, payload={"note": "xref_reload not supported in ezdxf"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def xref_bind(self, name, insert=False) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            return CommandResult(ok=True, payload={"note": "xref_bind not fully supported in ezdxf"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def xref_path_update(self, name, new_path) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            if name not in self._doc.blocks:
                return CommandResult(ok=False, error=f"Block/xref '{name}' not found")
            block = self._doc.blocks[name]
            old_path = block.dxf.get("xref_path", "")
            block.dxf.xref_path = new_path
            return CommandResult(ok=True, payload={
                "name": name,
                "old_path": old_path,
                "new_path": new_path,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def xref_query_entities(self, name, entity_type=None, layer=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            block = self._doc.blocks.get(name)
            entities = []
            for e in block:
                dtype = e.dxftype()
                elayer = e.dxf.get("layer", "0")
                if entity_type and dtype.upper() != entity_type.upper():
                    continue
                if layer and elayer != layer:
                    continue
                entities.append({"handle": e.dxf.handle, "type": dtype, "layer": elayer})
            return CommandResult(ok=True, payload={"entities": entities, "count": len(entities)})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Layout ---

    async def layout_list(self) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            layouts = [{"name": layout.name} for layout in self._doc.layouts]
            return CommandResult(ok=True, payload={"layouts": layouts})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_create(self, name) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            self._doc.layouts.new(name)
            return CommandResult(ok=True, payload={"created": name})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_switch(self, name) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            return CommandResult(ok=True, payload={"switched": name, "note": "layout switch is visual-only, no-op in ezdxf"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_delete(self, name) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            self._doc.layouts.delete(name)
            return CommandResult(ok=True, payload={"deleted": name})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_viewport_create(self, x, y, width, height, scale=1.0) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            layout = self._doc.layouts.get("Layout1") or list(self._doc.layouts)[1]  # First paper space
            vp = layout.add_viewport(center=(x, y), size=(width, height))
            return CommandResult(ok=True, payload={"created": True})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_viewport_set_scale(self, viewport_id, scale) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            return CommandResult(ok=True, payload={"note": "viewport_set_scale limited in ezdxf"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_viewport_lock(self, viewport_id, lock=True) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            return CommandResult(ok=True, payload={"note": "viewport_lock limited in ezdxf"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_page_setup(self, name, paper_size=None, orientation=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            return CommandResult(ok=True, payload={"layout": name})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_titleblock_fill(self, layout_name, attributes=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            if not attributes:
                return CommandResult(ok=True, payload={"updated": 0})
            layout = self._doc.layouts.get(layout_name)
            if not layout:
                return CommandResult(ok=False, error=f"Layout not found: {layout_name}")
            updated = 0
            for e in layout:
                if e.dxftype() == "INSERT":
                    for a in e.attribs:
                        if a.dxf.tag in attributes:
                            a.dxf.text = attributes[a.dxf.tag]
                            updated += 1
            return CommandResult(ok=True, payload={"layout": layout_name, "updated": updated})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def layout_batch_plot(self, layouts=None, output_path=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            return CommandResult(ok=True, payload={"status": "batch_plot not supported in ezdxf backend"})
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- Electrical ---

    async def electrical_nec_lookup(self, table, parameters=None) -> CommandResult:
        try:
            parameters = parameters or {}

            # NEC Table 310.16 - Ampacity for copper conductors, THHN insulation (90C column)
            ampacity_thhn = {
                "14": 15, "12": 20, "10": 30, "8": 40, "6": 55, "4": 70,
                "3": 85, "2": 95, "1": 110, "1/0": 125, "2/0": 145, "3/0": 165,
                "4/0": 195, "250": 215, "300": 240, "350": 260, "500": 320,
            }
            # NEC Table 310.16 - Ampacity for copper conductors, 75C column (THWN, XHHW)
            ampacity_75c = {
                "14": 15, "12": 20, "10": 30, "8": 40, "6": 55, "4": 65,
                "3": 75, "2": 85, "1": 100, "1/0": 115, "2/0": 130, "3/0": 150,
                "4/0": 180, "250": 205, "300": 230, "350": 250, "500": 310,
            }

            # NEC Table 9 - AC resistance per 1000ft (ohms, copper, in conduit)
            wire_resistance = {
                "14": 3.14, "12": 1.98, "10": 1.24, "8": 0.778, "6": 0.491,
                "4": 0.308, "3": 0.245, "2": 0.194, "1": 0.154, "1/0": 0.122,
                "2/0": 0.0967, "3/0": 0.0766, "4/0": 0.0608,
            }

            # NEC Chapter 9 Table 4 - Conduit areas (in^2) for EMT
            conduit_areas_emt = {
                "1/2": 0.304, "3/4": 0.533, "1": 0.864, "1-1/4": 1.496,
                "1-1/2": 2.036, "2": 3.356, "2-1/2": 4.866, "3": 7.499,
                "3-1/2": 9.521, "4": 12.554,
            }
            # Conduit areas for RMC (rigid metal conduit)
            conduit_areas_rmc = {
                "1/2": 0.314, "3/4": 0.549, "1": 0.887, "1-1/4": 1.526,
                "1-1/2": 2.071, "2": 3.408, "2-1/2": 4.866, "3": 7.499,
                "3-1/2": 9.521, "4": 12.554,
            }

            # Wire cross-section areas (in^2) - THHN insulation
            wire_areas_thhn = {
                "14": 0.0097, "12": 0.0133, "10": 0.0211, "8": 0.0366,
                "6": 0.0507, "4": 0.0824, "3": 0.0973, "2": 0.1158,
                "1": 0.1562, "1/0": 0.1855, "2/0": 0.2223, "3/0": 0.2679,
                "4/0": 0.3237,
            }

            # NEC Chapter 9 Table 1 - Conduit fill percentages
            conduit_fill_pct = {
                1: 53,  # 1 wire
                2: 31,  # 2 wires
            }
            # 3 or more wires: 40%

            if table == "wire_ampacity" or table == "ampacity":
                wire_gauge = parameters.get("wire_gauge")
                insulation = parameters.get("insulation", "THHN").upper()
                if insulation == "THHN":
                    tbl = ampacity_thhn
                    tbl_name = "NEC 310.16 Copper THHN (90C)"
                else:
                    tbl = ampacity_75c
                    tbl_name = "NEC 310.16 Copper 75C"
                if wire_gauge:
                    amp = tbl.get(str(wire_gauge))
                    if amp is None:
                        return CommandResult(ok=False, error=f"Wire gauge '{wire_gauge}' not found in table")
                    return CommandResult(ok=True, payload={
                        "table": tbl_name,
                        "wire_gauge": wire_gauge,
                        "ampacity": amp,
                    })
                return CommandResult(ok=True, payload={"table": tbl_name, "values": tbl})

            elif table == "wire_resistance":
                wire_gauge = parameters.get("wire_gauge")
                if wire_gauge:
                    r = wire_resistance.get(str(wire_gauge))
                    if r is None:
                        return CommandResult(ok=False, error=f"Wire gauge '{wire_gauge}' not found")
                    return CommandResult(ok=True, payload={
                        "table": "NEC Table 9 AC Resistance",
                        "wire_gauge": wire_gauge,
                        "resistance_per_1000ft": r,
                        "unit": "ohms/1000ft",
                    })
                return CommandResult(ok=True, payload={
                    "table": "NEC Table 9 AC Resistance",
                    "values": wire_resistance,
                    "unit": "ohms/1000ft",
                })

            elif table == "conduit_fill":
                conduit_size = parameters.get("conduit_size")
                conduit_type = parameters.get("conduit_type", "EMT").upper()
                areas = conduit_areas_emt if conduit_type == "EMT" else conduit_areas_rmc
                if conduit_size:
                    area = areas.get(str(conduit_size))
                    if area is None:
                        return CommandResult(ok=False, error=f"Conduit size '{conduit_size}' not found")
                    return CommandResult(ok=True, payload={
                        "table": f"NEC Chapter 9 Table 4 ({conduit_type})",
                        "conduit_size": conduit_size,
                        "area_sqin": area,
                    })
                return CommandResult(ok=True, payload={
                    "table": f"NEC Chapter 9 Table 4 ({conduit_type})",
                    "values": areas,
                    "unit": "sq inches",
                })

            elif table == "wire_area":
                wire_gauge = parameters.get("wire_gauge")
                if wire_gauge:
                    area = wire_areas_thhn.get(str(wire_gauge))
                    if area is None:
                        return CommandResult(ok=False, error=f"Wire gauge '{wire_gauge}' not found")
                    return CommandResult(ok=True, payload={
                        "table": "Wire Cross-Section THHN",
                        "wire_gauge": wire_gauge,
                        "area_sqin": area,
                    })
                return CommandResult(ok=True, payload={
                    "table": "Wire Cross-Section THHN",
                    "values": wire_areas_thhn,
                    "unit": "sq inches",
                })

            else:
                return CommandResult(ok=True, payload={
                    "error": f"Unknown table: {table}",
                    "available_tables": [
                        "wire_ampacity", "wire_resistance",
                        "conduit_fill", "wire_area",
                    ],
                })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def electrical_voltage_drop(self, voltage, current, wire_gauge, length, phase=1, power_factor=1.0) -> CommandResult:
        try:
            # Wire resistance per 1000ft (copper, approximate)
            resistance = {
                "14": 3.14, "12": 1.98, "10": 1.24, "8": 0.778, "6": 0.491,
                "4": 0.308, "3": 0.245, "2": 0.194, "1": 0.154, "1/0": 0.122,
                "2/0": 0.0967, "3/0": 0.0766, "4/0": 0.0608,
            }
            r = resistance.get(str(wire_gauge), 1.0)
            if phase == 1:
                vd = 2 * current * r * (length / 1000.0)
            else:
                vd = 1.732 * current * r * (length / 1000.0)
            vd_pct = (vd / voltage) * 100 if voltage > 0 else 0
            return CommandResult(ok=True, payload={
                "voltage_drop": round(vd, 2),
                "voltage_drop_percent": round(vd_pct, 2),
                "wire_gauge": wire_gauge,
                "length_ft": length,
                "current_a": current,
                "voltage": voltage,
                "acceptable": vd_pct <= 3.0,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def electrical_conduit_fill(self, conduit_size, conduit_type, wire_gauges=None) -> CommandResult:
        try:
            # Conduit area (sq in) for EMT
            conduit_areas = {
                "1/2": 0.304, "3/4": 0.533, "1": 0.864, "1-1/4": 1.496,
                "1-1/2": 2.036, "2": 3.356, "2-1/2": 5.858, "3": 8.846,
                "3-1/2": 11.545, "4": 14.753,
            }
            # Wire area (sq in) with insulation THHN
            wire_areas = {
                "14": 0.0097, "12": 0.0133, "10": 0.0211, "8": 0.0366,
                "6": 0.0507, "4": 0.0824, "3": 0.0973, "2": 0.1158,
                "1": 0.1562, "1/0": 0.1855, "2/0": 0.2223, "3/0": 0.2679,
                "4/0": 0.3237,
            }
            conduit_area = conduit_areas.get(conduit_size, 0)
            total_wire = sum(wire_areas.get(g, 0) for g in (wire_gauges or []))
            fill_pct = (total_wire / conduit_area * 100) if conduit_area > 0 else 0
            max_fill = 40 if len(wire_gauges or []) > 2 else (31 if len(wire_gauges or []) == 2 else 53)
            return CommandResult(ok=True, payload={
                "conduit_size": conduit_size,
                "conduit_area_sqin": conduit_area,
                "wire_area_sqin": round(total_wire, 4),
                "fill_percent": round(fill_pct, 1),
                "max_fill_percent": max_fill,
                "acceptable": fill_pct <= max_fill,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def electrical_load_calc(self, devices=None) -> CommandResult:
        try:
            if not devices:
                return CommandResult(ok=True, payload={
                    "total_watts": 0, "total_va": 0, "total_amps": 0,
                    "device_count": 0, "device_details": [],
                })

            device_details = []
            total_watts = 0
            total_va = 0

            for d in devices:
                name = d.get("name", "unnamed")
                watts = d.get("watts", 0)
                voltage = d.get("voltage", 120)
                pf = d.get("power_factor", 1.0)

                if watts == 0 and "hp" in d:
                    watts = d["hp"] * 746  # Convert HP to watts

                va = watts / pf if pf > 0 else watts
                amps = va / voltage if voltage > 0 else 0

                device_details.append({
                    "name": name,
                    "watts": round(watts, 2),
                    "va": round(va, 2),
                    "amps": round(amps, 2),
                    "voltage": voltage,
                    "power_factor": pf,
                })

                total_watts += watts
                total_va += va

            # Total amps (assume single-phase at dominant voltage)
            dominant_voltage = devices[0].get("voltage", 120) if devices else 120
            total_amps = total_va / dominant_voltage if dominant_voltage > 0 else 0

            return CommandResult(ok=True, payload={
                "total_watts": round(total_watts, 2),
                "total_va": round(total_va, 2),
                "total_kw": round(total_watts / 1000, 2),
                "total_kva": round(total_va / 1000, 2),
                "total_amps": round(total_amps, 2),
                "voltage": dominant_voltage,
                "device_count": len(devices),
                "device_details": device_details,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def electrical_symbol_insert(self, symbol_type, x, y, scale=1.0, rotation=0.0, layer=None) -> CommandResult:
        try:
            elec_layer = layer or "E-SYMBOLS"
            self._ensure_layer(elec_layer)
            s = scale
            handles = []

            symbol_type_lower = symbol_type.lower()

            if symbol_type_lower == "receptacle":
                # Circle with two parallel lines (duplex receptacle)
                e = self._msp.add_circle((x, y), 3 * s, dxfattribs={"layer": elec_layer})
                handles.append(e.dxf.handle)
                l1 = self._msp.add_line((x - 1.5 * s, y - 1 * s), (x - 1.5 * s, y + 1 * s), dxfattribs={"layer": elec_layer})
                l2 = self._msp.add_line((x + 1.5 * s, y - 1 * s), (x + 1.5 * s, y + 1 * s), dxfattribs={"layer": elec_layer})
                handles.extend([l1.dxf.handle, l2.dxf.handle])

            elif symbol_type_lower == "switch":
                # S with a line
                e = self._msp.add_text("S", dxfattribs={
                    "insert": (x - 1 * s, y - 1.25 * s), "height": 2.5 * s, "layer": elec_layer,
                })
                handles.append(e.dxf.handle)
                l = self._msp.add_line((x + 2 * s, y), (x + 5 * s, y), dxfattribs={"layer": elec_layer})
                handles.append(l.dxf.handle)

            elif symbol_type_lower == "light":
                # Circle with crosshairs
                e = self._msp.add_circle((x, y), 3 * s, dxfattribs={"layer": elec_layer})
                handles.append(e.dxf.handle)
                l1 = self._msp.add_line((x - 3 * s, y), (x + 3 * s, y), dxfattribs={"layer": elec_layer})
                l2 = self._msp.add_line((x, y - 3 * s), (x, y + 3 * s), dxfattribs={"layer": elec_layer})
                handles.extend([l1.dxf.handle, l2.dxf.handle])

            elif symbol_type_lower == "motor":
                # Circle with "M" inside
                e = self._msp.add_circle((x, y), 5 * s, dxfattribs={"layer": elec_layer})
                handles.append(e.dxf.handle)
                t = self._msp.add_text("M", dxfattribs={
                    "insert": (x - 1.5 * s, y - 1.5 * s), "height": 3 * s, "layer": elec_layer,
                })
                handles.append(t.dxf.handle)

            elif symbol_type_lower == "transformer":
                # Two circles side by side (primary/secondary coils)
                e1 = self._msp.add_circle((x - 3 * s, y), 3 * s, dxfattribs={"layer": elec_layer})
                e2 = self._msp.add_circle((x + 3 * s, y), 3 * s, dxfattribs={"layer": elec_layer})
                handles.extend([e1.dxf.handle, e2.dxf.handle])

            elif symbol_type_lower == "disconnect":
                # Two lines with a gap (knife switch)
                l1 = self._msp.add_line((x - 5 * s, y), (x - 1 * s, y), dxfattribs={"layer": elec_layer})
                l2 = self._msp.add_line((x + 1 * s, y), (x + 5 * s, y), dxfattribs={"layer": elec_layer})
                l3 = self._msp.add_line((x - 1 * s, y), (x + 1 * s, y + 3 * s), dxfattribs={"layer": elec_layer})
                handles.extend([l1.dxf.handle, l2.dxf.handle, l3.dxf.handle])

            elif symbol_type_lower == "panel":
                # Rectangle with internal lines (panel board)
                w, h = 8 * s, 12 * s
                pts = [(x - w / 2, y - h / 2), (x + w / 2, y - h / 2),
                       (x + w / 2, y + h / 2), (x - w / 2, y + h / 2)]
                e = self._msp.add_lwpolyline(pts, close=True, dxfattribs={"layer": elec_layer})
                handles.append(e.dxf.handle)
                # Center divider line
                l = self._msp.add_line((x, y - h / 2), (x, y + h / 2), dxfattribs={"layer": elec_layer})
                handles.append(l.dxf.handle)
                # Label
                t = self._msp.add_text("PNL", dxfattribs={
                    "insert": (x - 2 * s, y + h / 2 + 1 * s), "height": 2 * s, "layer": elec_layer,
                })
                handles.append(t.dxf.handle)

            elif symbol_type_lower == "junction_box" or symbol_type_lower == "jbox":
                # Small square
                half = 2 * s
                pts = [(x - half, y - half), (x + half, y - half),
                       (x + half, y + half), (x - half, y + half)]
                e = self._msp.add_lwpolyline(pts, close=True, dxfattribs={"layer": elec_layer})
                handles.append(e.dxf.handle)

            else:
                # Generic: circle with label
                e = self._msp.add_circle((x, y), 3 * s, dxfattribs={"layer": elec_layer})
                handles.append(e.dxf.handle)
                t = self._msp.add_text(symbol_type[:3].upper(), dxfattribs={
                    "insert": (x - 1.5 * s, y - 1 * s), "height": 2 * s, "layer": elec_layer,
                })
                handles.append(t.dxf.handle)

            # Apply rotation if needed
            if rotation != 0.0:
                from ezdxf.math import Matrix44
                m = Matrix44.z_rotate(math.radians(rotation))
                for h in handles:
                    ent = self._doc.entitydb.get(h)
                    if ent:
                        ent.translate(-x, -y, 0)
                        ent.transform(m)
                        ent.translate(x, y, 0)

            return CommandResult(ok=True, payload={
                "symbol_type": symbol_type,
                "position": [x, y],
                "handles": handles,
                "layer": elec_layer,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def electrical_circuit_trace(self, start_entity, layer=None) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            tolerance = 0.1

            # Collect all wire segments (LINE/LWPOLYLINE) on specified layer
            wire_segments = []  # list of (handle, [(x1,y1), (x2,y2), ...])
            for e in self._msp:
                if layer and e.dxf.get("layer", "0") != layer:
                    continue
                dtype = e.dxftype()
                if dtype == "LINE":
                    wire_segments.append((
                        e.dxf.handle,
                        [(e.dxf.start[0], e.dxf.start[1]), (e.dxf.end[0], e.dxf.end[1])],
                    ))
                elif dtype == "LWPOLYLINE":
                    pts = [(p[0], p[1]) for p in e.get_points(format="xy")]
                    if len(pts) >= 2:
                        wire_segments.append((e.dxf.handle, pts))

            if not wire_segments:
                return CommandResult(ok=False, error="No wire segments found on specified layer")

            # Build endpoint connectivity graph
            # For each segment, its endpoints connect to other segment endpoints
            def pts_match(p1, p2):
                return math.sqrt((p1[0] - p2[0]) ** 2 + (p1[1] - p2[1]) ** 2) < tolerance

            # Start from the given entity
            start_e = self._doc.entitydb.get(start_entity)
            if start_e is None:
                return CommandResult(ok=False, error=f"Start entity {start_entity} not found")

            # Get start entity endpoints
            start_points = []
            if start_e.dxftype() == "LINE":
                start_points = [(start_e.dxf.start[0], start_e.dxf.start[1]),
                                (start_e.dxf.end[0], start_e.dxf.end[1])]
            elif start_e.dxftype() == "LWPOLYLINE":
                pts = [(p[0], p[1]) for p in start_e.get_points(format="xy")]
                if pts:
                    start_points = [pts[0], pts[-1]]
            elif start_e.dxftype() == "INSERT":
                # Use insert point as connection point
                start_points = [(start_e.dxf.insert[0], start_e.dxf.insert[1])]

            if not start_points:
                return CommandResult(ok=False, error="Cannot determine connection points for start entity")

            # BFS traversal
            visited_handles = {start_entity}
            queue = list(start_points)
            traced_handles = [start_entity]
            traced_points = list(start_points)

            while queue:
                current_pt = queue.pop(0)
                for handle, pts in wire_segments:
                    if handle in visited_handles:
                        continue
                    endpoints = [pts[0], pts[-1]]
                    for ep in endpoints:
                        if pts_match(current_pt, ep):
                            visited_handles.add(handle)
                            traced_handles.append(handle)
                            # Add the OTHER endpoint to the queue
                            other = pts[-1] if ep == pts[0] else pts[0]
                            queue.append(other)
                            traced_points.append(other)
                            break

            return CommandResult(ok=True, payload={
                "start_entity": start_entity,
                "traced_handles": traced_handles,
                "traced_count": len(traced_handles),
                "connection_points": [[round(p[0], 4), round(p[1], 4)] for p in traced_points],
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def electrical_panel_schedule_gen(self, panel_block, x=0, y=0) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            self._ensure_layer("E-PANEL-SCHEDULE")

            # Panel schedule table dimensions
            col_widths = [8, 25, 10, 10, 25, 8]  # CKT, DESCRIPTION, AMPS, AMPS, DESCRIPTION, CKT
            headers = ["CKT", "DESCRIPTION", "AMPS", "AMPS", "DESCRIPTION", "CKT"]
            row_height = 5
            total_width = sum(col_widths)
            num_rows = 22  # 42 circuits (21 odd + 21 even), plus header

            handles = []

            # Draw title
            title = self._msp.add_text(f"PANEL: {panel_block}", dxfattribs={
                "insert": (x, y + (num_rows + 1) * row_height + 3),
                "height": 3.0, "layer": "E-PANEL-SCHEDULE",
            })
            handles.append(title.dxf.handle)

            # Draw header row
            header_y = y + num_rows * row_height
            cx = x
            for i, hdr in enumerate(headers):
                t = self._msp.add_text(hdr, dxfattribs={
                    "insert": (cx + 1, header_y + 1),
                    "height": 1.8, "layer": "E-PANEL-SCHEDULE",
                })
                handles.append(t.dxf.handle)
                cx += col_widths[i]

            # Draw grid
            # Horizontal lines
            for row in range(num_rows + 1):
                ry = y + row * row_height
                l = self._msp.add_line((x, ry), (x + total_width, ry), dxfattribs={"layer": "E-PANEL-SCHEDULE"})
                handles.append(l.dxf.handle)

            # Vertical lines
            cx = x
            for i in range(len(col_widths) + 1):
                l = self._msp.add_line((cx, y), (cx, y + num_rows * row_height), dxfattribs={"layer": "E-PANEL-SCHEDULE"})
                handles.append(l.dxf.handle)
                if i < len(col_widths):
                    cx += col_widths[i]

            # Fill circuit numbers
            for row in range(num_rows - 1):  # -1 for header
                ry = y + row * row_height
                odd_ckt = row * 2 + 1
                even_ckt = row * 2 + 2
                # Left circuit number
                t1 = self._msp.add_text(str(odd_ckt), dxfattribs={
                    "insert": (x + 2, ry + 1), "height": 1.5, "layer": "E-PANEL-SCHEDULE",
                })
                # Right circuit number
                t2 = self._msp.add_text(str(even_ckt), dxfattribs={
                    "insert": (x + total_width - col_widths[-1] + 2, ry + 1),
                    "height": 1.5, "layer": "E-PANEL-SCHEDULE",
                })
                handles.extend([t1.dxf.handle, t2.dxf.handle])

            # Outline
            outline = self._msp.add_lwpolyline([
                (x, y), (x + total_width, y),
                (x + total_width, y + num_rows * row_height),
                (x, y + num_rows * row_height),
            ], close=True, dxfattribs={"layer": "E-PANEL-SCHEDULE"})
            handles.append(outline.dxf.handle)

            return CommandResult(ok=True, payload={
                "panel": panel_block,
                "position": [x, y],
                "size": [total_width, num_rows * row_height],
                "circuits": (num_rows - 1) * 2,
                "handles": handles,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    async def electrical_wire_number_assign(self, layer, prefix="W", start_num=1) -> CommandResult:
        if not self._doc:
            return CommandResult(ok=False, error="No document open")
        try:
            self._ensure_layer("E-WIRE-NUMBERS")

            # Find all wire segments (LINE/LWPOLYLINE) on the specified layer
            wire_entities = []
            for e in self._msp:
                if e.dxf.get("layer", "0") != layer:
                    continue
                dtype = e.dxftype()
                if dtype in ("LINE", "LWPOLYLINE"):
                    wire_entities.append(e)

            if not wire_entities:
                return CommandResult(ok=False, error=f"No wire segments found on layer '{layer}'")

            assignments = []
            num = start_num

            for e in wire_entities:
                wire_number = f"{prefix}{num}"
                dtype = e.dxftype()

                # Calculate midpoint for label placement
                if dtype == "LINE":
                    mx = (e.dxf.start[0] + e.dxf.end[0]) / 2
                    my = (e.dxf.start[1] + e.dxf.end[1]) / 2
                elif dtype == "LWPOLYLINE":
                    pts = list(e.get_points(format="xy"))
                    if len(pts) >= 2:
                        mid_idx = len(pts) // 2
                        mx = (pts[mid_idx - 1][0] + pts[mid_idx][0]) / 2
                        my = (pts[mid_idx - 1][1] + pts[mid_idx][1]) / 2
                    else:
                        mx, my = pts[0][0], pts[0][1]
                else:
                    continue

                # Place wire number text near midpoint (offset slightly above)
                t = self._msp.add_text(wire_number, dxfattribs={
                    "insert": (mx, my + 1.5),
                    "height": 1.5,
                    "layer": "E-WIRE-NUMBERS",
                })

                assignments.append({
                    "handle": e.dxf.handle,
                    "wire_number": wire_number,
                    "label_handle": t.dxf.handle,
                    "position": [round(mx, 4), round(my, 4)],
                })
                num += 1

            return CommandResult(ok=True, payload={
                "layer": layer,
                "wire_count": len(assignments),
                "assignments": assignments,
                "prefix": prefix,
                "start_num": start_num,
                "end_num": num - 1,
            })
        except Exception as ex:
            return CommandResult(ok=False, error=str(ex))

    # --- View ---

    async def get_screenshot(self) -> CommandResult:
        data = self._screenshot.capture()
        if data:
            return CommandResult(ok=True, payload=data)
        return CommandResult(ok=False, error="Screenshot render failed")

    # --- Helpers ---

    @staticmethod
    def _segment_length(x1: float, y1: float, x2: float, y2: float, bulge: float) -> float:
        """Calculate the length of a polyline segment, accounting for bulge (arc)."""
        chord = math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
        if abs(bulge) < 1e-10:
            return chord
        # Bulge = tan(included_angle / 4)
        angle = 4 * math.atan(abs(bulge))
        # radius = chord / (2 * sin(angle/2))
        sin_half = math.sin(angle / 2)
        if abs(sin_half) < 1e-10:
            return chord
        radius = chord / (2 * sin_half)
        # arc length = radius * angle
        return radius * angle

    @staticmethod
    def _color_to_int(color: str | int) -> int:
        if isinstance(color, int):
            return color
        color_map = {
            "red": 1, "yellow": 2, "green": 3, "cyan": 4,
            "blue": 5, "magenta": 6, "white": 7, "grey": 8, "gray": 8,
        }
        return color_map.get(color.lower(), 7)

    # --- Batch Drawing Operations ---

    async def batch_draw_lines(self, lines: list[dict]) -> CommandResult:
        if not self._msp:
            return CommandResult(ok=False, error="No drawing open")
        results = []
        for i, spec in enumerate(lines):
            try:
                r = await self.create_line(
                    spec.get("x1", 0), spec.get("y1", 0),
                    spec.get("x2", 0), spec.get("y2", 0),
                    layer=spec.get("layer"),
                )
                results.append({"index": i, "ok": r.ok, "handle": r.payload.get("handle") if r.ok else None})
            except Exception as e:
                results.append({"index": i, "ok": False, "error": str(e)})
        return CommandResult(ok=True, payload={"total": len(lines), "created": sum(1 for r in results if r["ok"]), "results": results})

    async def batch_draw_circles(self, circles: list[dict]) -> CommandResult:
        if not self._msp:
            return CommandResult(ok=False, error="No drawing open")
        results = []
        for i, spec in enumerate(circles):
            try:
                r = await self.create_circle(
                    spec.get("cx", 0), spec.get("cy", 0), spec.get("radius", 1),
                    layer=spec.get("layer"),
                )
                results.append({"index": i, "ok": r.ok, "handle": r.payload.get("handle") if r.ok else None})
            except Exception as e:
                results.append({"index": i, "ok": False, "error": str(e)})
        return CommandResult(ok=True, payload={"total": len(circles), "created": sum(1 for r in results if r["ok"]), "results": results})

    async def batch_draw_rectangles(self, rectangles: list[dict]) -> CommandResult:
        if not self._msp:
            return CommandResult(ok=False, error="No drawing open")
        results = []
        for i, spec in enumerate(rectangles):
            try:
                r = await self.create_rectangle(
                    spec.get("x1", 0), spec.get("y1", 0),
                    spec.get("x2", 0), spec.get("y2", 0),
                    layer=spec.get("layer"),
                )
                results.append({"index": i, "ok": r.ok, "handle": r.payload.get("handle") if r.ok else None})
            except Exception as e:
                results.append({"index": i, "ok": False, "error": str(e)})
        return CommandResult(ok=True, payload={"total": len(rectangles), "created": sum(1 for r in results if r["ok"]), "results": results})

    async def batch_draw_polylines(self, polylines: list[dict]) -> CommandResult:
        if not self._msp:
            return CommandResult(ok=False, error="No drawing open")
        results = []
        for i, spec in enumerate(polylines):
            try:
                r = await self.create_polyline(
                    spec.get("points", []),
                    closed=spec.get("closed", False),
                    layer=spec.get("layer"),
                )
                results.append({"index": i, "ok": r.ok, "handle": r.payload.get("handle") if r.ok else None})
            except Exception as e:
                results.append({"index": i, "ok": False, "error": str(e)})
        return CommandResult(ok=True, payload={"total": len(polylines), "created": sum(1 for r in results if r["ok"]), "results": results})

    async def batch_draw_texts(self, texts: list[dict]) -> CommandResult:
        if not self._msp:
            return CommandResult(ok=False, error="No drawing open")
        results = []
        for i, spec in enumerate(texts):
            try:
                r = await self.create_text(
                    spec.get("x", 0), spec.get("y", 0), spec.get("text", ""),
                    height=spec.get("height", 2.5),
                    rotation=spec.get("rotation", 0),
                    layer=spec.get("layer"),
                )
                results.append({"index": i, "ok": r.ok, "handle": r.payload.get("handle") if r.ok else None})
            except Exception as e:
                results.append({"index": i, "ok": False, "error": str(e)})
        return CommandResult(ok=True, payload={"total": len(texts), "created": sum(1 for r in results if r["ok"]), "results": results})

    # --- NLP ---

    async def execute_natural_command(self, command: str) -> CommandResult:
        from autocad_mcp.nlp.processor import NLPProcessor
        processor = NLPProcessor()
        parsed = processor.parse_command(command)
        return CommandResult(ok=True, payload={
            "operation": parsed.operation,
            "parameters": parsed.parameters,
            "confidence": parsed.confidence,
            "note": "Parsed only — execution is handled by the server tool layer",
        })

    # --- Excel Export ---

    async def export_to_excel(self, filename: str = "drawing_data.xlsx",
                               output_dir: str | None = None) -> CommandResult:
        if not self._msp:
            return CommandResult(ok=False, error="No drawing open")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            out = Path(output_dir or "~/Documents/AutoCAD MCP Exports").expanduser()
            out.mkdir(parents=True, exist_ok=True)
            filepath = out / filename

            wb = Workbook()
            ws = wb.active
            ws.title = "Entities"

            columns = ["Handle", "Type", "Layer", "Color"]
            hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF")
            for ci, cn in enumerate(columns, 1):
                cell = ws.cell(row=1, column=ci, value=cn)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")

            row = 2
            for entity in self._msp:
                handle = entity.dxf.handle if entity.dxf.handle else f"ezdxf_{row}"
                ws.cell(row=row, column=1, value=handle)
                ws.cell(row=row, column=2, value=entity.dxftype())
                ws.cell(row=row, column=3, value=entity.dxf.layer if hasattr(entity.dxf, 'layer') else "0")
                ws.cell(row=row, column=4, value=entity.dxf.color if hasattr(entity.dxf, 'color') else 7)
                row += 1

            for ci in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 18
            ws.freeze_panes = "A2"
            wb.save(str(filepath))

            return CommandResult(ok=True, payload={"file": str(filepath), "count": row - 2})
        except ImportError:
            return CommandResult(ok=False, error="openpyxl not installed. Install with: pip install openpyxl")
        except Exception as e:
            return CommandResult(ok=False, error=str(e))
